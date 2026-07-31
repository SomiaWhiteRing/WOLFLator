import tempfile
import unittest
import json
from pathlib import Path
from unittest import mock

from openpyxl import Workbook

from formats import ARTIFACT_EPOCH
from fonts import BUNDLED_FONT_FAMILY, BUNDLED_FONT_ID, load_font_scheme
from models import (
    AppSettings,
    ImportProtectionRules,
    ImportScope,
    RunMode,
    Stage,
    StageStatus,
    ToolResult,
    TranslationItem,
)
from pipeline import Pipeline, create_project, load_manifest
from wolf_editor import EditorInfo, analyze_auto_export
from wolf_analysis import ANALYSIS_ENGINE, write_program_cache
from wolf_tools import (
    OfficialToolDialogError,
    UberWolfRunner,
    dump_items,
    load_items,
)


def make_game(root: Path) -> Path:
    root.mkdir(parents=True)
    (root / "Game.exe").write_bytes(b"game")
    (root / "Data" / "BasicData").mkdir(parents=True)
    (root / "Data" / "BasicData" / "Game.dat").write_bytes(b"data")
    return root


class FakePipeline(Pipeline):
    executed = None

    def _stage_artifacts_valid(self, _stage: Stage) -> bool:
        return True

    def _execute(self, stage: Stage) -> dict[str, str]:
        if self.executed is not None:
            self.executed.append(stage)
        if stage is Stage.COPY:
            return self._copy()
        return {"artifact": str(self.artifacts_dir / f"{stage.value}.ok")}


class FailingPipeline(FakePipeline):
    def _execute(self, stage: Stage) -> dict[str, str]:
        if stage is Stage.GLOSSARY:
            raise RuntimeError("simulated failure")
        return super()._execute(stage)


class PipelineTests(unittest.TestCase):
    _LEGACY_DIALOG = (
        "Warning! | The process completed, but the Editor.exe version used "
        "to create the game data seems to be old!"
    )

    def test_legacy_export_respects_auto_conversion_setting(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            manifest_path = create_project(root / "projects", make_game(root / "game"))
            messages = []
            pipeline = Pipeline(
                manifest_path,
                AppSettings(auto_convert_legacy_games=False),
                "",
                root / "cache",
                glossary_api_key="",
                log=messages.append,
            )
            runner = mock.Mock(scope=ImportScope(external=False))
            runner.extract.side_effect = OfficialToolDialogError([self._LEGACY_DIALOG])
            with self.assertRaisesRegex(RuntimeError, "自动转换 Ver2"):
                pipeline._run_scoped_export(runner, "EXTRACT")
            self.assertTrue(any("请在设置中开启" in message for message in messages))

            pipeline.settings.auto_convert_legacy_games = True
            workbook = root / "source.xlsx"
            runner.extract.side_effect = [
                OfficialToolDialogError([self._LEGACY_DIALOG]),
                workbook,
            ]
            conversion = mock.Mock()
            with mock.patch("pipeline.convert_legacy_game", return_value=conversion) as convert:
                self.assertEqual(workbook, pipeline._run_scoped_export(runner, "EXTRACT"))
            convert.assert_called_once()
            self.assertIs(conversion, pipeline._legacy_conversion_result)

    def test_unpack_excludes_mtool_trsdata_with_hash_evidence(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            game = make_game(root / "game")
            manifest_path = create_project(root / "projects", game)
            pipeline = Pipeline(
                manifest_path,
                AppSettings(),
                "",
                root / "cache",
                glossary_api_key="",
            )
            pipeline.work_dir.mkdir(parents=True)
            (pipeline.work_dir / "Game.exe").write_bytes(b"game")
            mtool = pipeline.work_dir / "TrsData.bin"
            mtool.write_bytes(b"MTool data")
            variant = pipeline.work_dir / "TrsData_ChatGPT_2023108 124849.bin"
            variant.write_bytes(b"MTool translated data")
            with mock.patch("pipeline.prepare_uberwolf", return_value=root / "UberWolfCli.exe"):
                with mock.patch.object(UberWolfRunner, "unpack"):
                    artifacts = pipeline._unpack()
            self.assertFalse(mtool.exists())
            self.assertFalse(variant.exists())
            evidence = json.loads(Path(artifacts["excluded_files"]).read_text(encoding="utf-8"))
            self.assertEqual(
                {"TrsData.bin", "TrsData_ChatGPT_2023108 124849.bin"},
                {item["relative_path"] for item in evidence["files"]},
            )
            self.assertEqual({10, 21}, {item["bytes"] for item in evidence["files"]})

    def test_uberwolf_merges_incomplete_loose_data_over_archive(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            game = root / "game"
            (game / "Data" / "Online").mkdir(parents=True)
            (game / "Data" / "Online" / "decode.csv").write_text(
                "loose", encoding="utf-8"
            )
            (game / "Game.exe").write_bytes(b"game")
            (game / "Data.wolf").write_bytes(b"archive")
            executable = root / "UberWolfCli.exe"
            executable.write_bytes(b"tool")

            def unpack(command, **_kwargs):
                sandbox = Path(command[1]).parent
                basic = sandbox / "Data" / "BasicData"
                basic.mkdir(parents=True)
                (basic / "Game.dat").write_bytes(b"unpacked")
                (sandbox / "Data" / "Online").mkdir()
                (sandbox / "Data" / "Online" / "decode.csv").write_text(
                    "archive", encoding="utf-8"
                )
                return ToolResult(command, 0)

            with mock.patch("wolf_tools.run_process", side_effect=unpack):
                UberWolfRunner(executable).unpack(game)

            self.assertEqual(
                b"unpacked", (game / "Data" / "BasicData" / "Game.dat").read_bytes()
            )
            self.assertEqual(
                "loose",
                (game / "Data" / "Online" / "decode.csv").read_text(encoding="utf-8"),
            )

    def _attach_editor_analysis(self, pipeline: Pipeline) -> Path:
        path = pipeline.artifacts_dir / "editor-analysis.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        auto_dir = pipeline.artifacts_dir / "editor-auto"
        auto_dir.mkdir(parents=True, exist_ok=True)
        basic = auto_dir / "BasicData"
        basic.mkdir()
        (basic / "CommonEvent.dat.Auto.txt").write_text(
            "\n".join(
                (
                    "[COMMON_EVENT_TEXT_OUTPUT]",
                    "COMMON_EVENT_NUM=1",
                    "COMMON_ID=1",
                    "COMMON_NAME=Fixture",
                    "COMMAND_NUM=3",
                    "WoditorEvCOMMAND_START",
                    '[101][0,1]<0>()("甲")',
                    '[101][0,1]<0>()("\\\\C[1]乙")',
                    '[101][0,1]<0>()("原文")',
                    "WoditorEvCOMMAND_END",
                )
            ),
            encoding="utf-8",
        )
        items_path = pipeline.manifest.version.stage(Stage.EXTRACT).artifacts.get("items")
        items = load_items(items_path) if items_path else []
        editor_path = pipeline.artifacts_dir / "Editor.exe"
        editor_path.write_bytes(b"editor")
        report = analyze_auto_export(
            auto_dir,
            items,
            EditorInfo(
                editor_path,
                "3.713.2026.718",
                (3, 713, 2026, 718),
                "a" * 64,
            ),
            input_hash="fixture",
        )
        path.write_text(json.dumps(report, ensure_ascii=False), encoding="utf-8")
        pipeline.manifest.version.stage(Stage.EXTRACT).artifacts["editor_analysis"] = str(path)
        pipeline.manifest.version.stage(Stage.EXTRACT).artifacts["editor_auto_dir"] = str(auto_dir)
        return path

    def _translation_pipeline(self, root: Path) -> Pipeline:
        manifest_path = create_project(root / "projects", make_game(root / "game"))
        pipeline = Pipeline(
            manifest_path,
            AppSettings(translation_rounds=6),
            "secret",
            root / "cache",
            glossary_api_key="",
        )
        make_game(pipeline.work_dir)
        items = [
            TranslationItem(key="plain", original="甲", code="COMMON-1-0-0"),
            TranslationItem(
                key="control",
                original=r"\C[1]乙",
                code="COMMON-1-1-0",
                control_signature=[r"\C[1]"],
            ),
        ]
        items_path = dump_items(pipeline.artifacts_dir / "items-extracted.json", items)
        pipeline.manifest.version.stage(Stage.EXTRACT).artifacts["items"] = str(items_path)
        self._attach_editor_analysis(pipeline)
        (pipeline.project_dir / "glossary.json").write_text("{}", encoding="utf-8")
        return pipeline

    def test_corrupt_editor_program_cache_rebuilds_from_auto(self):
        with tempfile.TemporaryDirectory() as directory:
            pipeline = self._translation_pipeline(Path(directory))
            analysis_path = Path(
                pipeline.manifest.version.stage(Stage.EXTRACT).artifacts[
                    "editor_analysis"
                ]
            )
            items = load_items(
                pipeline.manifest.version.stage(Stage.EXTRACT).artifacts["items"]
            )
            cache_path = analysis_path.with_name("editor-program.json")
            write_program_cache(cache_path, analysis_path, items)
            cache_path.write_text("{}", encoding="utf-8")
            expected = json.loads(analysis_path.read_text(encoding="utf-8"))
            editor = EditorInfo(
                Path("Editor.exe"),
                "3.713.2026.718",
                (3, 713, 2026, 718),
                "a" * 64,
            )

            with mock.patch("pipeline.inspect_wolf_editor", return_value=editor), mock.patch(
                "pipeline.analyze_auto_export", return_value=expected
            ) as analyze, mock.patch.object(pipeline, "save"):
                pipeline.manifest.version.stage(Stage.EXTRACT).artifacts.update(
                    {
                        "editor_program": str(cache_path),
                        "editor_version": "3.713.2026.718",
                        "editor_sha256": "a" * 64,
                    }
                )
                self.assertEqual(expected, pipeline._editor_analysis(items))

            analyze.assert_called_once()
            cache = json.loads(cache_path.read_text(encoding="utf-8"))
            self.assertEqual("editor-program-cache", cache["kind"])
            self.assertEqual(ARTIFACT_EPOCH, cache["epoch"])

    def _attach_import_protection(
        self, pipeline: Pipeline, protected_keys: tuple[str, ...] = ()
    ) -> Path:
        path = pipeline.artifacts_dir / "import-protection.json"
        path.write_text(
            json.dumps(
                {
                    "kind": "import-protection",
                    "epoch": ARTIFACT_EPOCH,
                    "protected_keys": list(protected_keys),
                    "safe_to_translate": [],
                    "keep_original": [],
                    "translation_overrides": {},
                    "approvals": {},
                    "unresolved_scopes": [],
                    "translated_replay": {},
                    "structural_diff": {"status": "passed", "differences": []},
                    "entries": [],
                    "summary": {},
                    "middle_dot_normalized": [],
                }
            ),
            encoding="utf-8",
        )
        pipeline.manifest.version.stage(Stage.IMPORT).artifacts[
            "import_protection"
        ] = str(path)
        return path

    def test_translation_retries_only_failed_rows_in_one_fresh_session(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            pipeline = self._translation_pipeline(root)
            calls = []

            def fake_translation(_runtime, input_json, output_dir, *_args, **_kwargs):
                rows = json.loads(Path(input_json).read_text(encoding="utf-8"))
                calls.append((rows, Path(output_dir)))
                if len(calls) == 1:
                    return [{**rows[0], "translation": "译文甲", "stage": 1}]
                self.assertEqual(["control"], [row["key"] for row in rows])
                return [
                    {
                        **rows[0],
                        "translation": chr(0xE100) + "译文乙",
                        "stage": 1,
                    }
                ]

            with mock.patch("pipeline.require_managed_runtime", return_value=root / "runtime"), mock.patch(
                "pipeline.run_translation", side_effect=fake_translation
            ):
                artifacts = pipeline._translate()

            self.assertEqual(2, len(calls))
            self.assertEqual("ainiee-output", calls[0][1].name)
            self.assertEqual("ainiee-retry-output", calls[1][1].name)
            merged = load_items(artifacts["items"])
            self.assertEqual(["译文甲", r"\C[1]译文乙"], [item.translation for item in merged])
            retry_input = json.loads(Path(artifacts["ainiee_retry_input"]).read_text(encoding="utf-8"))
            self.assertEqual(["control"], [row["key"] for row in retry_input])
            report = json.loads(Path(artifacts["ainiee_retry_result"]).read_text(encoding="utf-8"))
            self.assertEqual(1, report["first_pass_failed"])
            self.assertEqual(0, report["remaining_failed"])

    def test_translation_stops_after_one_failed_only_retry(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            pipeline = self._translation_pipeline(root)
            with mock.patch("pipeline.require_managed_runtime", return_value=root / "runtime"), mock.patch(
                "pipeline.run_translation", return_value=[]
            ) as run:
                with self.assertRaisesRegex(ValueError, "missing=2"):
                    pipeline._translate()
            self.assertEqual(2, run.call_count)




    def test_font_release_uses_official_workbook_and_verifies_output(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            manifest_path = create_project(root / "projects", make_game(root / "game"))
            logs = []
            pipeline = Pipeline(
                manifest_path,
                AppSettings(),
                "",
                root / "cache",
                glossary_api_key="",
                log=logs.append,
            )
            items = [
                TranslationItem(key=f"font-{index}", original=f"原字体{index}", code=f"BASICDATA-{index + 3}")
                for index in range(4)
            ]
            items.append(TranslationItem(key="text", original="原文", translation="中文𠀀", code="COMMON-1-2-0"))
            items.append(
                TranslationItem(
                    key="protected",
                    original="·隣𠀁",
                    translation="不会导入",
                    code="COMMON-1-3-0",
                )
            )
            items_path = dump_items(pipeline.artifacts_dir / "items-translated.json", items)
            pipeline.manifest.version.stage(Stage.VALIDATE).artifacts["items"] = str(items_path)
            workbook_path = pipeline.artifacts_dir / "source.xlsx"
            workbook_path.parent.mkdir(parents=True, exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "Code (No Change)",
                    "Flag (No Change)",
                    "Type",
                    "Info",
                    "Your notes",
                    "Original text (No Change)",
                    "Translated text 1 / Chinese (Simplified)",
                ]
            )
            for index in range(4):
                sheet.append(
                    [
                        f"BASICDATA-{index + 3}",
                        "",
                        "Basic Game Settings",
                        f"Font {index}",
                        "",
                        f"原字体{index}",
                        "",
                    ]
                )
            sheet.append(
                [
                    "COMMON-2",
                    "<Half-Width Characters Only>\nCOPY-FROM-BASICDATA-3",
                    "Event",
                    "Message",
                    "",
                    "原字体0",
                    "",
                ]
            )
            workbook.save(workbook_path)
            pipeline.manifest.version.stage(Stage.EXTRACT).artifacts["workbook"] = str(workbook_path)
            pipeline.manifest.version.stage(Stage.EXTRACT).artifacts["items"] = str(items_path)
            self._attach_editor_analysis(pipeline)
            self._attach_import_protection(pipeline, ("protected",))

            verification = root / "verification.xlsx"
            verify_book = Workbook()
            verify_sheet = verify_book.active
            verify_sheet.append(list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0])
            verify_sheet.append(
                [
                    "COMMON-2",
                    "<Half-Width Characters Only>",
                    "Event",
                    "Message",
                    "",
                    "原字体0",
                    "",
                ]
            )
            for index in range(4):
                verify_sheet.append(
                    [
                        f"BASICDATA-{index + 3}",
                        "",
                        "Basic Game Settings",
                        f"Font {index}",
                        "",
                        BUNDLED_FONT_FAMILY,
                        "",
                    ]
                )
            verify_book.save(verification)

            translated = make_game(root / "translated")
            extra_map = translated / "Data" / "MapData" / "Map0EX.mps"
            extra_map.parent.mkdir(parents=True)
            extra_map.write_bytes(b"extra-map")
            runner = mock.Mock()

            def translate(game_root, **_kwargs):
                generated = Path(game_root) / "Translated1_Chinese (Simplified)"
                make_game(generated)
                return generated

            runner.translate.side_effect = translate
            runner.extract.side_effect = [workbook_path, verification]
            runner.console_outputs = []
            temporary = pipeline.version_dir / ".release-ready"
            with mock.patch.object(
                pipeline,
                "_translation_safety",
                side_effect=AssertionError("release must reuse import protection"),
            ), mock.patch.object(pipeline, "_official_runner", return_value=runner):
                artifacts = pipeline._build_font_release(
                    translated, temporary, load_font_scheme(manifest_path.parent)
                )
            self.assertTrue((temporary / BUNDLED_FONT_ID).is_file())
            self.assertEqual(b"extra-map", (temporary / "Data" / "MapData" / "Map0EX.mps").read_bytes())
            self.assertEqual("4", artifacts["font_warning_count"])
            self.assertTrue(
                any(
                    line.startswith("[WARNING] 字体缺字：主字体") and '样例 "𠀀"' in line
                    for line in logs
                )
            )
            result = json.loads(Path(artifacts["font_result"]).read_text(encoding="utf-8"))
            self.assertEqual("font-result", result["kind"])
            self.assertEqual(ARTIFACT_EPOCH, result["epoch"])
            self.assertEqual([BUNDLED_FONT_FAMILY] * 4, result["applied_slots"])
            self.assertEqual("approved_import_translations", result["coverage_scope"])
            warnings = json.loads(Path(artifacts["font_warnings"]).read_text(encoding="utf-8"))
            self.assertTrue(
                all("𠀁" not in warning["missing"] for warning in warnings["warnings"])
            )
            runner.translate.assert_called_once()
            self.assertEqual(2, runner.extract.call_count)

    def test_font_release_rejects_non_font_text_changes(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            manifest_path = create_project(root / "projects", make_game(root / "game"))
            pipeline = Pipeline(manifest_path, AppSettings(), "", root / "cache", glossary_api_key="")
            items = [
                TranslationItem(key=f"font-{index}", original=f"原字体{index}", code=f"BASICDATA-{index + 3}")
                for index in range(4)
            ]
            items_path = dump_items(pipeline.artifacts_dir / "items.json", items)
            pipeline.manifest.version.stage(Stage.VALIDATE).artifacts["items"] = str(items_path)

            def workbook(path, text):
                book = Workbook()
                sheet = book.active
                sheet.append([
                    "Code (No Change)", "Flag (No Change)", "Type", "Info",
                    "Your notes", "Original text (No Change)",
                    "Translated text 1 / Chinese (Simplified)",
                ])
                for index in range(4):
                    sheet.append([
                        f"BASICDATA-{index + 3}", "", "Basic Game Settings", f"Font {index}",
                        "", BUNDLED_FONT_FAMILY, "",
                    ])
                sheet.append(["COMMON-1", "", "Event", "Message", "", text, ""])
                book.save(path)
                return path

            baseline = workbook(root / "baseline.xlsx", "未变化")
            changed = workbook(root / "changed.xlsx", "被改动")
            pipeline.manifest.version.stage(Stage.EXTRACT).artifacts = {
                "workbook": str(baseline),
                "items": str(items_path),
            }
            self._attach_editor_analysis(pipeline)
            self._attach_import_protection(pipeline)
            translated = make_game(root / "translated")
            generated = root / "generated"
            runner = mock.Mock()
            runner.extract.side_effect = [baseline, changed]
            runner.translate.side_effect = lambda *_args, **_kwargs: make_game(generated)
            runner.console_outputs = []
            with mock.patch.object(pipeline, "_official_runner", return_value=runner):
                with self.assertRaisesRegex(RuntimeError, "字体字段以外"):
                    pipeline._build_font_release(
                        translated,
                        pipeline.version_dir / ".release-ready",
                        load_font_scheme(manifest_path.parent),
                    )

    def test_font_release_failure_keeps_previous_release(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            manifest_path = create_project(root / "projects", make_game(root / "game"))
            pipeline = Pipeline(manifest_path, AppSettings(), "", root / "cache", glossary_api_key="")
            translated = make_game(root / "translated")
            pipeline.manifest.version.stage(Stage.IMPORT).artifacts["translated_game"] = str(translated)
            pipeline.release_dir.mkdir(parents=True)
            (pipeline.release_dir / "old.txt").write_text("keep", encoding="utf-8")
            with mock.patch.object(pipeline, "_build_font_release", side_effect=RuntimeError("font failed")):
                with self.assertRaisesRegex(RuntimeError, "font failed"):
                    pipeline._release()
            self.assertEqual("keep", (pipeline.release_dir / "old.txt").read_text(encoding="utf-8"))
            with mock.patch("pipeline.load_font_scheme", return_value=None), mock.patch(
                "pipeline.replace_with_retry",
                side_effect=PermissionError(13, "sharing violation"),
            ):
                with self.assertRaisesRegex(RuntimeError, "发布目录正在使用"):
                    pipeline._release()
            self.assertEqual("keep", (pipeline.release_dir / "old.txt").read_text(encoding="utf-8"))

    def test_import_uses_the_same_full_structure_as_export(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            pipeline = self._translation_pipeline(root)
            map_path = pipeline.work_dir / "Data" / "MapData" / "Map0EX.mps"
            map_path.parent.mkdir(parents=True)
            map_path.write_bytes(b"map")
            (pipeline.work_dir / "Data.wolf").write_bytes(b"packed")
            items = [
                TranslationItem(
                    key="plain",
                    original="甲",
                    translation="中・文",
                    code="COMMON-1-0-0",
                )
            ]
            items_path = dump_items(pipeline.artifacts_dir / "items-translated.json", items)
            pipeline.manifest.version.stage(Stage.VALIDATE).artifacts = {
                "full_workbook": str(pipeline.artifacts_dir / "translated-full.xlsx"),
                "items": str(items_path),
            }
            scoped = root / "import-scoped.xlsx"
            scoped.write_bytes(b"xlsx")
            runner = mock.Mock()
            runner.translate.side_effect = lambda game_root, **_kwargs: make_game(
                Path(game_root) / "Translated1_Chinese (Simplified)"
            )
            runner.diagnostics = []
            runner.console_outputs = []
            stale_diagnostics = pipeline.artifacts_dir / "official-diagnostics.json"
            stale_diagnostics.parent.mkdir(parents=True, exist_ok=True)
            stale_diagnostics.write_text("stale", encoding="utf-8")

            post_editor = mock.Mock(auto_dir=root / "post-auto", analysis_path=root / "post-analysis.json")

            def verify_merged(_editor, game_root, *_args, **_kwargs):
                merged = Path(game_root)
                self.assertTrue((merged / "Data" / "MapData" / "Map0EX.mps").is_file())
                self.assertFalse((merged / "Data.wolf").exists())
                return post_editor

            def write_scoped(_full, _output, _scope, _game, scoped_items, **_kwargs):
                self.assertEqual("中·文", scoped_items[0].translation)
                self.assertEqual({"plain": "安全混合译文"}, _kwargs["translation_overrides"])
                return scoped

            safety = {
                "engine": ANALYSIS_ENGINE,
                "safe_to_translate": ["plain"],
                "keep_original": [],
                "translation_overrides": {"plain": "安全混合译文"},
                "approvals": {},
                "unresolved_scopes": [],
                "replay": {},
                "reasons": {},
            }

            with mock.patch.object(pipeline, "_official_runner", return_value=runner) as factory, mock.patch(
                "pipeline.write_scoped_workbook", side_effect=write_scoped
            ), mock.patch.object(
                pipeline, "_translation_safety", return_value=safety
            ), mock.patch("pipeline.export_and_analyze", side_effect=verify_merged), mock.patch(
                "pipeline.compare_auto_structure", return_value={"status": "passed", "differences": []}
            ):
                artifacts = pipeline._import()

            runner_scope = factory.call_args.args[0]
            self.assertTrue(runner_scope.external)
            self.assertTrue(runner_scope.display)
            self.assertTrue(runner_scope.optional_name)
            self.assertTrue(runner_scope.halfwidth)
            self.assertTrue(runner_scope.filename)
            runner.translate.assert_called_once()
            self.assertEqual(
                str(pipeline.work_dir / "Translated1_Chinese (Simplified)"),
                artifacts["translated_game"],
            )
            self.assertFalse(any(pipeline.artifacts_dir.glob(".import-game-*")))
            self.assertFalse(stale_diagnostics.exists())
            protection = json.loads(
                Path(artifacts["import_protection"]).read_text(encoding="utf-8")
            )
            self.assertEqual(["plain"], protection["middle_dot_normalized"])

    def test_import_structure_failure_keeps_previous_translated_game(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            pipeline = self._translation_pipeline(root)
            items = [
                TranslationItem(
                    key="plain", original="甲", translation="译文", code="COMMON-1-0-0"
                )
            ]
            items_path = dump_items(
                pipeline.artifacts_dir / "items-translated.json", items
            )
            pipeline.manifest.version.stage(Stage.VALIDATE).artifacts = {
                "full_workbook": str(pipeline.artifacts_dir / "translated-full.xlsx"),
                "items": str(items_path),
            }
            old = make_game(pipeline.work_dir / "Translated1_Chinese (Simplified)")
            (old / "old.txt").write_text("keep", encoding="utf-8")
            scoped = root / "import-scoped.xlsx"
            scoped.write_bytes(b"xlsx")
            runner = mock.Mock()
            runner.translate.side_effect = lambda game_root, **_kwargs: make_game(
                Path(game_root) / "Translated1_Chinese (Simplified)"
            )
            runner.diagnostics = []
            runner.console_outputs = []
            post_editor = mock.Mock(
                auto_dir=root / "post-auto", analysis_path=root / "post-analysis.json"
            )
            with mock.patch.object(
                pipeline, "_official_runner", return_value=runner
            ), mock.patch(
                "pipeline.write_scoped_workbook", return_value=scoped
            ), mock.patch(
                "pipeline.export_and_analyze", return_value=post_editor
            ), mock.patch(
                "pipeline.compare_auto_structure",
                return_value={
                    "status": "failed",
                    "differences": [{"location": "event=1", "kind": "opcode"}],
                },
            ):
                with self.assertRaisesRegex(RuntimeError, "已拒绝本次导入"):
                    pipeline._import()

            self.assertEqual("keep", (old / "old.txt").read_text(encoding="utf-8"))
            self.assertFalse(any(pipeline.artifacts_dir.glob(".import-game-*")))


    def test_manifest_rejects_missing_translation_scope(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            manifest_path = create_project(root / "projects", make_game(root / "game"))
            data = json.loads(Path(manifest_path).read_text(encoding="utf-8"))
            data.pop("translation_scope")
            Path(manifest_path).write_text(json.dumps(data), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "项目格式不兼容"):
                load_manifest(manifest_path)

    def test_manifest_rejects_missing_kind_and_extra_fields(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            manifest_path = create_project(root / "projects", make_game(root / "game"))
            data = json.loads(manifest_path.read_text(encoding="utf-8"))
            data.pop("kind")
            manifest_path.write_text(json.dumps(data), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "项目格式不兼容"):
                load_manifest(manifest_path)

            data["kind"] = "project"
            data["unexpected"] = True
            manifest_path.write_text(json.dumps(data), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "项目格式不兼容"):
                load_manifest(manifest_path)

    def test_manifest_rejects_non_boolean_scope(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            manifest_path = create_project(root / "projects", make_game(root / "game"))
            data = json.loads(Path(manifest_path).read_text(encoding="utf-8"))
            data["translation_scope"]["display"] = "true"
            Path(manifest_path).write_text(json.dumps(data), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "项目格式不兼容"):
                load_manifest(manifest_path)


    def test_run_stages_executes_only_contiguous_selection(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            manifest_path = create_project(root / "projects", make_game(root / "game"))
            executed = []
            progress = []
            pipeline = FakePipeline(
                manifest_path, AppSettings(), "", root / "cache", glossary_api_key=""
            )
            pipeline.executed = executed
            pipeline.progress = lambda current, total, stage: progress.append(
                (current, total, stage)
            )
            stages = (Stage.UNPACK, Stage.EXTRACT, Stage.GLOSSARY)

            self.assertEqual("completed", pipeline.run_stages(stages))
            self.assertEqual(list(stages), executed)
            self.assertEqual(
                [
                    (0, 3, Stage.UNPACK),
                    (1, 3, Stage.UNPACK),
                    (1, 3, Stage.EXTRACT),
                    (2, 3, Stage.EXTRACT),
                    (2, 3, Stage.GLOSSARY),
                    (3, 3, Stage.GLOSSARY),
                ],
                progress,
            )
            current = load_manifest(manifest_path)
            self.assertEqual(
                [StageStatus.COMPLETED] * 3,
                [current.version.stage(stage).status for stage in stages],
            )
            self.assertEqual(StageStatus.PENDING, current.version.stage(Stage.COPY).status)
            self.assertEqual(StageStatus.PENDING, current.version.stage(Stage.TRANSLATE).status)

            with self.assertRaisesRegex(ValueError, "相邻"):
                pipeline.run_stages((Stage.COPY, Stage.EXTRACT))


    def test_failure_is_persisted_and_retryable(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            game = make_game(root / "game")
            manifest_path = create_project(root / "projects", game)
            settings = AppSettings(
                api_base_url="https://user:password@example.com/v1/secret-token?token=hidden",
                api_model="test-model",
                glossary_api_base_url="https://glossary-user:glossary-password@example.net/v1?key=glossary-hidden",
                glossary_api_model="glossary-model",
            )
            app_log = []
            pipeline = FailingPipeline(
                manifest_path,
                settings,
                "secret-token",
                root / "cache",
                glossary_api_key="glossary-secret",
                log=app_log.append,
            )
            with self.assertRaisesRegex(RuntimeError, "simulated"):
                pipeline.run()
            current = load_manifest(manifest_path)
            self.assertEqual(StageStatus.FAILED, current.version.stage(Stage.GLOSSARY).status)
            self.assertEqual(StageStatus.PENDING, current.version.stage(Stage.TRANSLATE).status)
            logs = list((Path(manifest_path).parent / "versions" / current.active_version / "artifacts" / "logs").glob("*.log"))
            self.assertEqual(1, len(logs))
            pipeline.log("credential=secret-token")
            pipeline.detail(
                "tool echoed https://user:password@example.com/v1/secret-token?token=hidden"
            )
            pipeline.detail(
                "glossary echoed https://glossary-user:glossary-password@example.net/v1?key=glossary-hidden glossary-secret"
            )
            log_text = logs[0].read_text(encoding="utf-8-sig")
            self.assertIn("simulated failure", log_text)
            self.assertIn("credential=[REDACTED]", log_text)
            self.assertIn("[DETAIL] stage.exception stage=glossary", log_text)
            self.assertIn("Traceback", log_text)
            self.assertIn("manifest.save.complete", log_text)
            self.assertIn("tool echoed https://example.com/v1/[REDACTED]", log_text)
            self.assertFalse(any("Traceback" in line for line in app_log))
            self.assertIn("api_url=https://example.com/v1/[REDACTED]", log_text)
            self.assertIn("glossary_api_url=https://example.net/v1", log_text)
            self.assertNotIn("secret-token", log_text)
            self.assertNotIn("glossary-secret", log_text)
            self.assertNotIn("glossary-password", log_text)
            self.assertNotIn("glossary-hidden", log_text)
            self.assertNotIn("password", log_text)
            self.assertNotIn("token=hidden", log_text)
            pipeline.retry_failed()
            self.assertEqual(StageStatus.PENDING, load_manifest(manifest_path).version.stage(Stage.GLOSSARY).status)

    def test_source_change_is_blocked(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            game = make_game(root / "game")
            manifest_path = create_project(root / "projects", game)
            pipeline = FakePipeline(
                manifest_path, AppSettings(), "", root / "cache", glossary_api_key=""
            )
            pipeline.set_run_mode(RunMode.STEP)
            pipeline.run()
            (game / "changed.txt").write_text("changed", encoding="utf-8")
            with self.assertRaisesRegex(RuntimeError, "新的源版本"):
                FakePipeline(
                    manifest_path, AppSettings(), "", root / "cache", glossary_api_key=""
                ).run()


    def test_import_protection_resets_only_affected_stages(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            manifest_path = create_project(root / "projects", make_game(root / "game"))
            pipeline = FakePipeline(
                manifest_path, AppSettings(), "", root / "cache", glossary_api_key=""
            )
            self.assertEqual("completed", pipeline.run())
            current = pipeline.manifest.import_protection
            pipeline.set_import_protection(
                ImportProtectionRules(
                    **{**current.__dict__, "protect_paths_and_commands": False}
                )
            )
            changed = load_manifest(manifest_path)
            self.assertEqual(StageStatus.COMPLETED, changed.version.stage(Stage.VALIDATE).status)
            self.assertEqual(StageStatus.PENDING, changed.version.stage(Stage.IMPORT).status)
            self.assertEqual(StageStatus.PENDING, changed.version.stage(Stage.RELEASE).status)

            with pipeline._mutation("restore-completed"):
                pipeline.manifest = changed
                for stage in Stage:
                    pipeline.manifest.version.stage(stage).status = StageStatus.COMPLETED
                pipeline.save()
            current = pipeline.manifest.import_protection
            pipeline.set_import_protection(
                ImportProtectionRules(
                    **{**current.__dict__, "allow_copy_condition_groups": False}
                )
            )
            changed = load_manifest(manifest_path)
            self.assertEqual(StageStatus.COMPLETED, changed.version.stage(Stage.EXTRACT).status)
            self.assertEqual(StageStatus.PENDING, changed.version.stage(Stage.GLOSSARY).status)



if __name__ == "__main__":
    unittest.main()
