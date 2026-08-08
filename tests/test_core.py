import http.server
import hashlib
import json
import os
import subprocess
import sys
import tempfile
import threading
import time
import unittest
import zipfile
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

import ainiee
from formats import ARTIFACT_EPOCH
from fonts import (
    BUNDLED_FONT_FAMILY,
    BUNDLED_FONT_SHA256,
    FontError,
    bundled_font_path,
    default_font_scheme,
    font_file_info,
    font_file_faces,
    load_font_scheme,
    load_original_fonts,
    record_original_fonts,
    save_font_scheme,
    validate_font_scheme,
)
from models import (
    AppSettings,
    ImportCategory,
    ImportProtectionRules,
    ImportScope,
    ToolResult,
    TranslationItem,
)
from wolf_tools import (
    CancelledError,
    _dismiss_process_dialogs,
    _official_config_text,
    _normalize_xlsx_shared_strings,
    _silent_official_executable,
    analyze_import_protection,
    classify_optional_name_delta,
    dump_items,
    full_export_scope,
    imported_display_texts,
    load_items,
    merge_ainiee_output,
    name_baseline_scope,
    parse_official_map_failures,
    official_dialogs_indicate_legacy_game,
    protect_control_tokens,
    read_translation_items,
    read_font_slots,
    reconcile_incremental,
    restore_control_tokens,
    run_process,
    selected_translation_requirements,
    temporary_external_filter_view,
    to_paratranz,
    write_font_workbook,
    write_full_workbook,
    write_scoped_workbook,
)
from wolf_editor import (
    _AnalysisAudit,
    _AnalysisState,
    _BlockAnalyzer,
    _address_variables_for_block,
    _Command,
    _CommandBlock,
    _DatabaseType,
    EditorRelease,
    EditorInfo,
    _NumberValue,
    _StringValue,
    _copy_editor_sandbox,
    _calculate_numbers,
    _loop_identity,
    _legacy_conversion_action,
    _database_index,
    _editor_execution_lock,
    _external_text_observer_report,
    _inspect_matching_runtime,
    _merge_states,
    _merge_numbers,
    _restore_editor_map_paths,
    _translation_usage_report,
    analyze_auto_export,
    analyze_translation_safety,
    compare_auto_structure,
    inspect_wolf_editor,
    install_supported_editor,
)
from wolf_analysis import (
    ANALYSIS_ENGINE,
    load_program_cache,
    source_structure_fingerprint,
    write_program_cache,
)


HEADERS = [
    "Code (No Change)",
    "Flag (No Change)",
    "Type",
    "Info",
    "Your notes",
    "Original text (No Change)",
    "Translated text 1 / Chinese (Simplified)",
]


class LegacyConversionContractTests(unittest.TestCase):
    def test_old_editor_dialog_and_conversion_sequence_are_exact(self):
        self.assertTrue(
            official_dialogs_indicate_legacy_game(
                [
                    "Warning! | The process completed, but the Editor.exe version used "
                    "to create the game data seems to be old!"
                ]
            )
        )
        self.assertFalse(official_dialogs_indicate_legacy_game(["unknown warning"]))
        self.assertEqual(
            ("start", "start"),
            _legacy_conversion_action(
                "【Ver2以前のWOLF RPGエディターをご利用だった方へ】 最初にファイルのコンバートを行います",
                "",
                True,
                started=False,
            ),
        )
        self.assertEqual(
            ("legacy-behavior", "no"),
            _legacy_conversion_action(
                "確認",
                "【注意！ Ver3では挙動が大きく変わります！】",
                True,
                started=True,
            ),
        )
        self.assertEqual(
            ("conversion-complete", "ok"),
            _legacy_conversion_action(
                "完了",
                "ファイルのコンバート作業が完了しました。",
                True,
                started=True,
            ),
        )

    def test_legacy_runtime_must_match_editor_version(self):
        with tempfile.TemporaryDirectory() as directory:
            runtime = Path(directory) / "Game.exe"
            runtime.write_bytes(b"runtime")
            editor = EditorInfo(
                Path(directory) / "Editor.exe",
                "3.713.2026.718",
                (3, 713, 2026, 718),
                "a" * 64,
            )
            with mock.patch(
                "wolf_editor._windows_version_resource",
                return_value=(
                    "3.713.2026.718",
                    (3, 713, 2026, 718),
                    "Game / WOLF RPG Editor",
                ),
            ):
                self.assertEqual(
                    hashlib.sha256(b"runtime").hexdigest(),
                    _inspect_matching_runtime(runtime, editor),
                )


def make_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    sheet.append(["COMMON-1", "", "Event", "Message", "", r"こんにちは\C[1]", ""])
    sheet.append(["NAME-D-SDB-1-0", "", "SDB info", "Data name", "", "主人公", ""])
    sheet.append(["SDB-1-0", "<FILENAME>", "Image", "File", "", "Picture/顔.png", ""])
    sheet.append(["COMMON-2", "<Half-Width Characters Only>", "Event", "Code", "", "ABC", ""])
    sheet.append(["COMMON-3", "COPY-FROM-COMMON-1", "Event", "Copy", "", r"こんにちは\C[1]", ""])
    sheet.append(["TXT-1", "", "TXT File", "Line", "", "外部テキスト", ""])
    sheet.append(["DUP", "", "Event", "A", "", "重複", ""])
    sheet.append(["DUP", "", "Event", "B", "", "重複", ""])
    sheet["A1"].font = Font(bold=True, color="FF112233")
    table = Table(displayName="WolfTranslation", ref=f"A1:G{sheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)
    workbook.save(path)
    return path


class WorkbookTests(unittest.TestCase):
    def test_xlsx_shared_strings_preserve_excel_control_escape_contract(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "control.xlsx"
            workbook = Workbook()
            workbook.active["A1"] = "before_x0008_after"
            workbook.save(path)
            _normalize_xlsx_shared_strings(path)

            with zipfile.ZipFile(path) as archive:
                self.assertIn("xl/sharedStrings.xml", archive.namelist())
                shared = archive.read("xl/sharedStrings.xml")
                sheet = archive.read("xl/worksheets/sheet1.xml")
            self.assertIn(b"_x0008_", shared)
            self.assertIn(b't="s"', sheet)
            self.assertNotIn(b"inlineStr", sheet)
            loaded = load_workbook(path)
            self.assertEqual("before_x0008_after", loaded.active["A1"].value)
            loaded.close()

    def test_calibrated_database_and_array_commands_have_conservative_effects(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            basic = root / "Auto" / "BasicData"
            basic.mkdir(parents=True)
            (basic / "CommonEvent.dat.Auto.txt").write_text(
                "\n".join(
                    (
                        "[COMMON_EVENT_TEXT_OUTPUT]",
                        "COMMON_EVENT_NUM=1",
                        "COMMON_ID=1",
                        "COMMON_NAME=Safety",
                        "COMMAND_NUM=4",
                        "WoditorEvCOMMAND_START",
                        '[252][5,4]<0>(0,0,0,0,1)("","","","")',
                        '[255][5,4]<0>(0,0,0,66304,1)("","CAL-ARRAY-A","","")',
                        '[257][5,4]<0>(0,0,0,66304,1)("","CAL-ARRAY-A","","")',
                        '[300][5,3]<0>(0,12321,0,0,0)("選択肢の用意","A","")',
                        "WoditorEvCOMMAND_END",
                    )
                ),
                encoding="utf-8",
            )
            editor = EditorInfo(
                root / "Editor.exe",
                "3.713.2026.718",
                (3, 713, 2026, 718),
                "a" * 64,
            )
            report = analyze_auto_export(
                root / "Auto", [], editor, input_hash="input"
            )
            catalog = report["command_catalog"]
            for name in (
                "shape_coverage",
                "semantic_coverage",
                "cfg_coverage",
                "call_target_coverage",
                "data_effect_coverage",
            ):
                self.assertEqual(1.0, catalog[name]["ratio"], name)
            self.assertEqual(0, catalog["opaque_effects"])
            effects = report["runtime_semantics"]["data_effects"]
            self.assertEqual(4, len(effects))
            self.assertTrue(
                all(
                    effect["status"] == "conservative"
                    for command_id, effect in effects.items()
                    if not command_id.endswith(":4")
                )
            )
            last_effect = next(
                value for key, value in effects.items() if key.endswith(":4")
            )
            self.assertEqual("exact", last_effect["status"])

    def test_known_shape_with_illegal_database_flags_is_not_semantically_covered(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            basic = root / "Auto" / "BasicData"
            basic.mkdir(parents=True)
            (basic / "CommonEvent.dat.Auto.txt").write_text(
                "\n".join(
                    (
                        "[COMMON_EVENT_TEXT_OUTPUT]",
                        "COMMON_EVENT_NUM=1",
                        "COMMON_ID=1",
                        "COMMON_NAME=Invalid",
                        "COMMAND_NUM=1",
                        "WoditorEvCOMMAND_START",
                        '[250][5,4]<0>(0,0,0,3840,0)("","","","")',
                        "WoditorEvCOMMAND_END",
                    )
                ),
                encoding="utf-8",
            )
            editor = EditorInfo(
                root / "Editor.exe", "3.713.2026.718", (3, 713, 2026, 718), "a" * 64
            )
            report = analyze_auto_export(root / "Auto", [], editor, input_hash="input")
            catalog = report["command_catalog"]
            self.assertEqual(1.0, catalog["shape_coverage"]["ratio"])
            self.assertEqual(0.0, catalog["semantic_coverage"]["ratio"])
            self.assertEqual(0.0, catalog["data_effect_coverage"]["ratio"])
            self.assertEqual(1, catalog["opaque_effects"])


    def test_classification_stable_keys_and_controls(self):
        with tempfile.TemporaryDirectory() as directory:
            path = make_workbook(Path(directory) / "source.xlsx")
            items = read_translation_items(path)
            self.assertEqual(8, len(items))
            self.assertEqual(
                [
                    ImportCategory.DISPLAY,
                    ImportCategory.OPTIONAL_NAME,
                    ImportCategory.FILENAME,
                    ImportCategory.HALFWIDTH,
                    ImportCategory.COPY,
                    ImportCategory.EXTERNAL,
                    ImportCategory.DISPLAY,
                    ImportCategory.DISPLAY,
                ],
                [item.category for item in items],
            )
            self.assertNotEqual(items[-1].key, items[-2].key)
            payload = to_paratranz(items, full_export_scope())
            self.assertEqual(6, len(payload))
            self.assertNotIn("外部テキスト", {row["original"] for row in payload})
            protected = payload[0]["original"]
            self.assertIn(chr(0xE100), protected)
            self.assertNotIn(r"\C[1]", protected)

    def test_safe_scope_excludes_optional_rows_from_ai(self):
        with tempfile.TemporaryDirectory() as directory:
            items = read_translation_items(make_workbook(Path(directory) / "source.xlsx"))
            payload = to_paratranz(items, ImportScope())
            self.assertEqual(3, len(payload))
            self.assertNotIn("主人公", {row["original"] for row in payload})
            self.assertNotIn("Picture/顔.png", {row["original"] for row in payload})
            translated = [
                {
                    **row,
                    "translation": "译文" + "".join(
                        char for char in row["original"] if 0xE100 <= ord(char) <= 0xF7FF
                    ),
                    "stage": 1,
                }
                for row in payload
            ]
            merged = merge_ainiee_output(items, translated, ImportScope())
            self.assertEqual("", merged[1].translation)
            self.assertEqual("", merged[2].translation)


    def test_control_failure_identifies_the_wolf_row(self):
        with tempfile.TemporaryDirectory() as directory:
            items = read_translation_items(make_workbook(Path(directory) / "source.xlsx"))
            payload = to_paratranz(items, ImportScope())
            translated = [{**row, "translation": "译文", "stage": 1} for row in payload]
            with self.assertRaisesRegex(ValueError, "COMMON-1.*占位序列"):
                merge_ainiee_output(items, translated, ImportScope())



    def test_official_config_excludes_external_files_and_font_rows_from_translation(self):
        config = _official_config_text(full_export_scope())
        self.assertIn("Tool_A_Get_CommonEvent_Name=1\r\n", config)
        self.assertIn("Tool_A_Get_DB_DataName=1\r\n", config)
        self.assertIn("Tool_A_Get_TXT=0\r\n", config)
        self.assertIn("Tool_A_Get_CSV=0\r\n", config)
        baseline = _official_config_text(name_baseline_scope())
        self.assertIn("Tool_A_Get_CommonEvent_Name=0\r\n", baseline)
        self.assertIn("Tool_A_Get_DB_DataName=0\r\n", baseline)
        self.assertIn("Tool_A_Get_TXT=0\r\n", baseline)
        self.assertIn("Tool_A_Get_CSV=0\r\n", baseline)
        external_baseline = _official_config_text(
            name_baseline_scope(ImportScope(external=True))
        )
        self.assertIn("Tool_A_Get_TXT=1\r\n", external_baseline)
        self.assertIn("Tool_A_Get_CSV=1\r\n", external_baseline)

        with tempfile.TemporaryDirectory() as directory:
            source = make_workbook(Path(directory) / "source.xlsx")
            workbook = load_workbook(source)
            workbook.active["A9"] = "BASICDATA-3"
            workbook.save(source)
            items = read_translation_items(source)
            items[-1].translation = "旧版字体译文"
            payload = to_paratranz(items, full_export_scope())
            self.assertNotIn(items[-1].key, {row["key"] for row in payload})
            full = write_full_workbook(source, Path(directory) / "full.xlsx", items)
            self.assertIsNone(load_workbook(full).active["G9"].value)

    def test_font_workbook_changes_fonts_and_pins_copy_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "source.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(HEADERS)
            for index, code in enumerate(("BASICDATA-3", "BASICDATA-4", "BASICDATA-5", "BASICDATA-6")):
                sheet.append([code, "", "Basic Game Settings", f"Font {index}", "", f"原字体{index}", "旧译"])
            sheet.append(
                [
                    "COMMON-1",
                    "<Half-Width Characters Only>\nCOPY-FROM-BASICDATA-3",
                    "Event",
                    "Message",
                    "",
                    "原字体0",
                    "旧译",
                ]
            )
            sheet.append(
                ["COMMON-2", "COPY-FROM-COMMON-1", "Event", "Message", "", "原字体0", "旧译"]
            )
            sheet.append(["COMMON-3", "", "Event", "Message", "", "原文", "旧译"])
            sheet.append(
                ["COMMON-4", "COPY-FROM-COMMON-3", "Event", "Message", "", "原文", "旧译"]
            )
            workbook.save(source)
            slots = ["字体一", "字体二", "字体三", "字体四"]
            output = write_font_workbook(source, root / "font.xlsx", slots)
            sheet = load_workbook(output).active
            self.assertEqual(slots, [sheet.cell(row, 7).value for row in range(2, 6)])
            self.assertEqual("原字体0", sheet.cell(6, 7).value)
            self.assertEqual("原字体0", sheet.cell(7, 7).value)
            self.assertEqual("<Half-Width Characters Only>", sheet.cell(6, 2).value)
            self.assertIsNone(sheet.cell(7, 2).value)
            self.assertIsNone(sheet.cell(8, 7).value)
            self.assertIsNone(sheet.cell(9, 7).value)
            self.assertEqual("COPY-FROM-COMMON-3", sheet.cell(9, 2).value)
            self.assertEqual([f"原字体{index}" for index in range(4)], read_font_slots(read_translation_items(output)))


    def test_imported_display_texts_exclude_preserved_originals(self):
        source = TranslationItem(
            key="source",
            original="原文甲",
            translation="译文乙\\C[1]",
            code="COMMON-1",
        )
        copied = TranslationItem(
            key="copy",
            original="原文甲",
            code="COMMON-2",
            flag="COPY-FROM-COMMON-1",
            category=ImportCategory.COPY,
            copy_category=ImportCategory.DISPLAY,
        )
        protected = TranslationItem(
            key="protected",
            original="·隣",
            translation="相邻",
            code="COMMON-3",
        )
        unchanged = TranslationItem(
            key="unchanged",
            original="原文丙",
            translation="原文丙",
            code="COMMON-4",
        )

        self.assertEqual(
            ["译文乙", "译文乙"],
            imported_display_texts(
                [source, copied, protected, unchanged],
                ImportScope(),
                protected_keys={protected.key},
            ),
        )


class WorkbookAndFontTests(unittest.TestCase):
    def test_editor_program_cache_roundtrip_and_validation(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            report_path = root / "editor-analysis.json"
            cache_path = root / "editor-program.json"
            item = TranslationItem(
                key="key",
                original="原文",
                translation="译文",
                code="COMMON-1-0-0",
            )
            report = {
                "kind": "editor-analysis",
                "epoch": ARTIFACT_EPOCH,
                "engine": ANALYSIS_ENGINE,
                "input_hash": "input",
                "output_hash": "output",
                "editor": {"version": "3.713", "sha256": "a" * 64},
            }
            report_path.write_text(
                json.dumps(report, ensure_ascii=False), encoding="utf-8"
            )

            write_program_cache(cache_path, report_path, [item])
            self.assertEqual(
                report,
                load_program_cache(
                    cache_path,
                    items=[item],
                    input_hash="input",
                    editor_version="3.713",
                    editor_sha256="a" * 64,
                ),
            )
            translated_again = TranslationItem(
                key="key",
                original="原文",
                translation="另一候选",
                stage=1,
                code="COMMON-1-0-0",
            )
            self.assertEqual(
                source_structure_fingerprint([item]),
                source_structure_fingerprint([translated_again]),
            )
            with self.assertRaisesRegex(ValueError, "源结构不匹配"):
                load_program_cache(
                    cache_path,
                    items=[TranslationItem(key="other", original="原文")],
                )
            report_path.write_text("{}", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "哈希不匹配"):
                load_program_cache(cache_path)


    def test_external_filter_view_excludes_only_files_over_the_kb_limit(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            game = root / "game"
            data = game / "Data" / "nested"
            data.mkdir(parents=True)
            (game / "Game.exe").write_bytes(b"game")
            exact = data / "exact.TXT"
            large = data / "large.csv"
            binary = data / "large.bin"
            exact.write_bytes(b"x" * (128 * 1024))
            large.write_bytes(b"y" * (128 * 1024 + 1))
            binary.write_bytes(b"z" * (128 * 1024 + 1))
            original = {path: path.read_bytes() for path in (exact, large, binary)}
            stale = root / "version" / ".wolflator-export-view-stale"
            stale.mkdir(parents=True)

            with temporary_external_filter_view(game, root / "version", 128) as (view, excluded):
                view_path = view
                self.assertTrue((view / "Data" / "nested" / "exact.TXT").is_file())
                self.assertFalse((view / "Data" / "nested" / "large.csv").exists())
                self.assertTrue((view / "Data" / "nested" / "large.bin").is_file())
                self.assertEqual(
                    [(str(Path("Data") / "nested" / "large.csv"), 128 * 1024 + 1)],
                    excluded,
                )
                self.assertFalse(stale.exists())

            self.assertFalse(view_path.exists())
            self.assertEqual(original, {path: path.read_bytes() for path in original})

    def test_original_fonts_are_immutable_per_version(self):
        with tempfile.TemporaryDirectory() as directory:
            project = Path(directory)
            workbook = project / "source.xlsx"
            workbook.write_bytes(b"source")
            first = ["主字体", "副字体一", "副字体二", "副字体三"]
            path = record_original_fonts(project, "v1", first, "a" * 64, workbook)
            self.assertEqual(first, load_original_fonts(project, "v1")["slots"])
            self.assertEqual(path, record_original_fonts(project, "v1", first, "a" * 64, workbook))
            with self.assertRaisesRegex(FontError, "请新建游戏版本"):
                record_original_fonts(
                    project,
                    "v1",
                    ["被修改", *first[1:]],
                    "a" * 64,
                    workbook,
                )
            second = ["新主字体", *first[1:]]
            record_original_fonts(project, "v2", second, "b" * 64, workbook)
            self.assertEqual(second, load_original_fonts(project, "v2")["slots"])

    def test_bundled_font_and_default_scheme_are_verified(self):
        path = bundled_font_path()
        families, codepoints = font_file_info(path)
        self.assertIn(BUNDLED_FONT_FAMILY, families)
        self.assertIn(ord("中"), codepoints)
        self.assertEqual(BUNDLED_FONT_SHA256, __import__("hashlib").sha256(path.read_bytes()).hexdigest())
        with tempfile.TemporaryDirectory() as directory:
            save_font_scheme(directory, default_font_scheme())
            scheme = load_font_scheme(directory)
            self.assertEqual([BUNDLED_FONT_FAMILY] * 4, [slot["family"] for slot in scheme["slots"]])

    def test_font_scheme_rejects_project_path_escape(self):
        scheme = default_font_scheme()
        scheme["slots"][0] = {
            "mode": "font",
            "family": "Bad Font",
            "provenance": "system",
            "files": [
                {
                    "kind": "project",
                    "path": "../bad.ttf",
                    "filename": "bad.ttf",
                    "sha256": "0" * 64,
                }
            ],
        }
        with tempfile.TemporaryDirectory() as directory, self.assertRaisesRegex(FontError, "越界"):
            validate_font_scheme(directory, scheme, check_files=False)


    def test_font_parser_rejects_truncated_and_unsupported_files(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            truncated = root / "bad.ttf"
            truncated.write_bytes(b"\x00\x01")
            with self.assertRaises(FontError):
                font_file_info(truncated)
            unsupported = root / "font.woff"
            unsupported.write_bytes(b"font")
            with self.assertRaisesRegex(FontError, "不支持"):
                font_file_info(unsupported)


    @unittest.skipUnless(Path(r"C:\Windows\Fonts\msyh.ttc").is_file(), "Windows YaHei TTC unavailable")

    @unittest.skipUnless(Path(r"C:\Windows\Fonts\msyhl.ttc").is_file(), "Windows YaHei Light unavailable")
    def test_font_faces_preserve_gdi_family_style_and_weight(self):
        faces = font_file_faces(Path(r"C:\Windows\Fonts\msyhl.ttc"))
        face = next(item for item in faces if item.family == "Microsoft YaHei Light")
        self.assertEqual("Microsoft YaHei", face.preview_family)
        self.assertEqual("Light", face.style)
        self.assertEqual(290, face.weight)

    @unittest.skipUnless(Path(r"C:\Windows\Fonts\msyhl.ttc").is_file(), "Windows YaHei Light unavailable")

    def test_full_baseline_delta_and_cross_category_copy_are_scope_safe(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            full_path = root / "full.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(HEADERS)
            sheet.append(["NAME-D-UDB-1-0", "", "UDB info", "Data name", "", "攻撃力", ""])
            sheet.append(
                [
                    "UDB-1-0-0",
                    "COPY-FROM-NAME-D-UDB-1-0",
                    "Status",
                    "Label",
                    "",
                    "攻撃力",
                    "",
                ]
            )
            sheet.append(["COMMON-1-0-0", "", "Event", "(Common Event)", "", "内部名", ""])
            sheet.append(["DISPLAY-1", "", "Event", "Message", "", "顔", ""])
            sheet.append(["FILE-1", "<FILENAME>\nCOPY-FROM-DISPLAY-1", "Image", "File", "", "顔", ""])
            sheet.append(["NAME-D-SDB-0-9", "<FILENAME>", "SDB info", "Data name", "", "トイレ", ""])
            sheet.append(["DISPLAY-2", "COPY-FROM-NAME-D-SDB-0-9", "Event", "Message", "", "トイレ", ""])
            workbook.save(full_path)

            baseline_path = root / "baseline.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(HEADERS)
            sheet.append(["UDB-1-0-0", "", "Status", "Label", "", "攻撃力", ""])
            sheet.append(["DISPLAY-1", "", "Event", "Message", "", "顔", ""])
            sheet.append(["FILE-1", "<FILENAME>\nCOPY-FROM-DISPLAY-1", "Image", "File", "", "顔", ""])
            sheet.append(["DISPLAY-2", "COPY-FROM-NAME-D-SDB-0-9", "Event", "Message", "", "トイレ", ""])
            workbook.save(baseline_path)

            items = read_translation_items(full_path)
            baseline_items = read_translation_items(baseline_path)
            self.assertEqual(3, classify_optional_name_delta(items, baseline_items))
            self.assertEqual(ImportCategory.OPTIONAL_NAME, items[0].category)
            self.assertEqual(ImportCategory.DISPLAY, items[1].copy_category)
            self.assertEqual(ImportCategory.OPTIONAL_NAME, items[2].category)

            payload = to_paratranz(items, ImportScope())
            self.assertEqual([], payload)
            translated_full = write_full_workbook(full_path, root / "translated.xlsx", items)
            scoped = write_scoped_workbook(
                translated_full,
                root / "scoped.xlsx",
                ImportScope(),
                root / "game",
                items,
            )
            output = load_workbook(scoped)
            self.assertIsNone(output.active["G2"].value)
            self.assertEqual("攻撃力", output.active["G3"].value)
            self.assertIsNone(output.active["G4"].value)
            self.assertIsNone(output.active["G5"].value)
            self.assertEqual("顔", output.active["G6"].value)
            self.assertIsNone(output.active["G7"].value)
            self.assertEqual("トイレ", output.active["G8"].value)

    def test_database_name_and_all_copy_references_change_together(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source_path = root / "source.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(HEADERS)
            sheet.append(["NAME-T-UDB-17", "", "UDB info", "Type name", "", "システム設定", ""])
            sheet.append(
                [
                    "MAP-0-Ev014-Page1-72-2",
                    "COPY-FROM-NAME-T-UDB-17",
                    "Event",
                    "Show Choice",
                    "",
                    "システム設定",
                    "",
                ]
            )
            sheet.append(
                [
                    "COMMON-48-87-1",
                    "COPY-FROM-NAME-T-UDB-17",
                    "Event",
                    "DB Management",
                    "",
                    "システム設定",
                    "",
                ]
            )
            workbook.save(source_path)

            items = read_translation_items(source_path)
            items[1].copy_category = ImportCategory.DISPLAY
            items[2].copy_category = ImportCategory.OPTIONAL_NAME
            items[0].translation = "系统设置"
            translated = write_full_workbook(source_path, root / "translated.xlsx", items)

            default_scoped = write_scoped_workbook(
                translated,
                root / "default.xlsx",
                ImportScope(),
                root / "game",
                items,
            )
            default_sheet = load_workbook(default_scoped).active
            self.assertIsNone(default_sheet["G2"].value)
            self.assertEqual("システム設定", default_sheet["G3"].value)
            self.assertEqual("システム設定", default_sheet["G4"].value)

            names_scoped = write_scoped_workbook(
                translated,
                root / "names.xlsx",
                ImportScope(optional_name=True),
                root / "game",
                items,
            )
            names_sheet = load_workbook(names_scoped).active
            self.assertEqual("系统设置", names_sheet["G2"].value)
            self.assertEqual("系统设置", names_sheet["G3"].value)
            self.assertEqual("系统设置", names_sheet["G4"].value)

    def test_workbook_writes_equals_prefix_as_literal_text(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source_path = root / "source.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(HEADERS)
            sheet.append(
                ["NAME-T-UDB-1", "", "UDB", "Type", "", "==literal==", ""]
            )
            sheet.append(
                [
                    "UDB-1-0-0",
                    "COPY-FROM-NAME-T-UDB-1",
                    "UDB",
                    "Name",
                    "",
                    "==literal==",
                    "",
                ]
            )
            workbook.save(source_path)

            items = read_translation_items(source_path)
            items[0].translation = "=translated"
            full = write_full_workbook(source_path, root / "full.xlsx", items)
            full_sheet = load_workbook(full, data_only=False).active
            self.assertEqual("=translated", full_sheet["G2"].value)
            self.assertEqual("s", full_sheet["G2"].data_type)

            default_scoped = write_scoped_workbook(
                full, root / "default.xlsx", ImportScope(), root / "game", items
            )
            default_sheet = load_workbook(default_scoped, data_only=False).active
            self.assertEqual("==literal==", default_sheet["G3"].value)
            self.assertEqual("s", default_sheet["G3"].data_type)

            names_scoped = write_scoped_workbook(
                full,
                root / "names.xlsx",
                ImportScope(optional_name=True),
                root / "game",
                items,
            )
            names_sheet = load_workbook(names_scoped, data_only=False).active
            self.assertEqual("=translated", names_sheet["G3"].value)
            self.assertEqual("s", names_sheet["G3"].data_type)

    def test_import_protection_keeps_conditions_and_external_references_original(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "source.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(HEADERS)
            sheet.append(
                [
                    "COMMON-63-167-2",
                    "",
                    "◆操作盤3",
                    r"CEv63 [index167](Condition[String])\n\cself[9] to compare",
                    "",
                    "HPバー",
                    "",
                ]
            )
            sheet.append(["UDB-35-92-0", "", "◆スチル", "名称", "", "再起動_1", ""])
            sheet.append(["UDB-35-93-0", "", "◆スチル", "名称", "", "SHORT1", ""])
            sheet.append(["DISPLAY-1", "", "Event", "Show Message", "", "通常表示", ""])
            workbook.save(source)
            items = read_translation_items(source)
            items[0].translation = "HP条"
            items[1].translation = "重新启动_1"
            items[2].translation = "短片1"
            items[3].translation = "正常显示"
            game = root / "game"
            scenario = game / "Data" / "textfile" / "01_scenario.md"
            scenario.parent.mkdir(parents=True)
            scenario.write_text("@s再起動_1\n@sSHORT1\n", encoding="utf-8")

            report = analyze_import_protection(
                items,
                ImportScope(),
                game,
                ImportProtectionRules(protect_logic_references=False),
            )
            protected_codes = {
                entry["code"]
                for entry in report["entries"]
                if entry["action"] == "keep_original"
            }
            self.assertEqual(
                {
                    "COMMON-63-167-2",
                    "DISPLAY-1",
                    "UDB-35-92-0",
                    "UDB-35-93-0",
                },
                protected_codes,
            )

            full = write_full_workbook(source, root / "full.xlsx", items)
            scoped = write_scoped_workbook(
                full,
                root / "scoped.xlsx",
                ImportScope(),
                game,
                items,
                allow_copy_condition_groups=True,
                protected_keys=set(report["protected_keys"]),
            )
            output = load_workbook(scoped).active
            self.assertIsNone(output["G2"].value)
            self.assertIsNone(output["G3"].value)
            self.assertIsNone(output["G4"].value)
            self.assertIsNone(output["G5"].value)

    def test_import_protection_uses_editor_logic_evidence(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            game = root / "game"
            source = root / "source.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(HEADERS)
            sheet.append(["UDB-7-1-1", "", "◆[rb]操作盤名", "[1:操作盤1] [リソース名1]", "", "戦_HPバー", ""])
            sheet.append(["UDB-7-1-2", "", "◆[rb]操作盤名", "[1:操作盤1] [リソース名2]", "", "戦_HPバー2", ""])
            sheet.append(["CDB-1-11-0", "COPY-FROM-UDB-7-1-1", "汎用リソース管理", "[11:戦_HPバー] [name]", "", "戦_HPバー", ""])
            sheet.append(["COMMON-63-167-2", "", "◆操作盤3", "(Condition[String])", "", "HPバー", ""])
            sheet.append(["DISPLAY-1", "", "画面", "名称", "", "再起動_1", ""])
            workbook.save(source)
            items = read_translation_items(source)
            for item in items:
                if item.code.startswith("UDB-7-1"):
                    item.translation = "翻译后的键"
                elif item.code == "COMMON-63-167-2":
                    item.translation = "HP 条"
                elif item.code == "DISPLAY-1":
                    item.translation = "重新启动_1"
            by_code = {item.code: item for item in items}
            with self.assertRaisesRegex(ValueError, "Editor 分析报告格式不兼容"):
                analyze_import_protection(
                    items, ImportScope(), game, ImportProtectionRules(), {}
                )
            dependency = {
                "auto_file": "BasicData/CommonEvent.dat.Auto.txt",
                "event_type": "common",
                "event_id": 63,
                "event_name": "◆[rb]操作盤3",
                "page": 1,
                "command": 168,
                "operator": "contains",
                "literal": "HPバー",
                "condition_keys": [by_code["COMMON-63-167-2"].key],
                "source_keys": [
                    by_code["UDB-7-1-1"].key,
                    by_code["UDB-7-1-2"].key,
                ],
                "right_source_keys": [],
                "database_cells": [
                    {"database": "UDB", "type": 7, "data": 1, "field": 1},
                    {"database": "UDB", "type": 7, "data": 1, "field": 2},
                ],
                "status": "resolved",
                "reason": "",
            }
            analysis = {
                "kind": "editor-analysis",
                "epoch": ARTIFACT_EPOCH,
                "engine": ANALYSIS_ENGINE,
                "input_hash": "input",
                "output_hash": "output",
                "editor": {"version": "3.713", "sha256": "a" * 64},
                "unknown_commands": [],
                "blocking_issues": [],
                "dependencies": [dependency],
                "safe_to_translate": [by_code["DISPLAY-1"].key],
            }
            report = analyze_import_protection(
                items,
                ImportScope(),
                game,
                ImportProtectionRules(),
                analysis,
            )
            protected_codes = {
                entry["code"]
                for entry in report["entries"]
                if entry["action"] == "keep_original"
            }
            self.assertIn("UDB-7-1-1", protected_codes)
            self.assertIn("UDB-7-1-2", protected_codes)
            self.assertIn("COMMON-63-167-2", protected_codes)
            self.assertNotIn("DISPLAY-1", protected_codes)
            resource_dependency = {
                **dependency,
                "kind": "resource",
                "operator": "resource_reference",
                "resource_role": "resource_path",
                "condition_keys": [],
                "source_keys": [by_code["COMMON-63-167-2"].key],
                "unresolved_scopes": ["project"],
            }
            resource_report = analyze_import_protection(
                items,
                ImportScope(),
                game,
                ImportProtectionRules(),
                {**analysis, "dependencies": [resource_dependency]},
            )
            self.assertIn(
                by_code["COMMON-63-167-2"].key, resource_report["protected_keys"]
            )
            self.assertNotIn(by_code["DISPLAY-1"].key, resource_report["protected_keys"])
            resource_entry = next(
                entry
                for entry in resource_report["entries"]
                if entry["reason"] == "resource_reference"
            )
            self.assertEqual("resource_path", resource_entry["resource_role"])
            call_dependency = {
                **dependency,
                "kind": "call",
                "operator": "event_call",
                "literal": "X[共]アイテム増減",
                "condition_keys": [],
                "source_keys": [by_code["COMMON-63-167-2"].key],
                "unresolved_scopes": ["common:63"],
            }
            call_report = analyze_import_protection(
                items,
                ImportScope(),
                game,
                ImportProtectionRules(),
                {**analysis, "dependencies": [call_dependency]},
            )
            self.assertIn(
                by_code["COMMON-63-167-2"].key, call_report["protected_keys"]
            )
            self.assertTrue(
                any(
                    entry["reason"] == "event_call_target"
                    for entry in call_report["entries"]
                )
            )
            variable_side = {
                **dependency,
                "status": "untracked",
                "reason": "字符串变量比较的一侧来源未知",
                "right_is_variable": True,
                "condition_keys": [],
                "source_keys": [],
                "right_source_keys": [by_code["UDB-7-1-1"].key],
            }
            variable_report = analyze_import_protection(
                items,
                ImportScope(),
                game,
                ImportProtectionRules(),
                {**analysis, "dependencies": [variable_side]},
            )
            self.assertIn(by_code["UDB-7-1-1"].key, variable_report["protected_keys"])
            for operator, literal in (
                ("equals", "戦_HPバー"),
                ("not_equals", "戦_HPバー"),
                ("contains", "HPバー"),
                ("starts_with", "戦_"),
            ):
                trial = analyze_import_protection(
                    items,
                    ImportScope(),
                    game,
                    ImportProtectionRules(),
                    {**analysis, "dependencies": [{**dependency, "operator": operator, "literal": literal}]},
                )
                self.assertIn(by_code["UDB-7-1-1"].key, trial["protected_keys"])
            derived = analyze_import_protection(
                items,
                ImportScope(),
                game,
                ImportProtectionRules(),
                {
                    **analysis,
                    "dependencies": [
                        {
                            **dependency,
                            "source_keys": [by_code["UDB-7-1-1"].key],
                            "literal": "永不命中",
                            "left_values": ["前缀/戦_HPバー"],
                        }
                    ],
                },
            )
            self.assertTrue(
                any(
                    entry["reason"] == "logic_derived_value"
                    for entry in derived["entries"]
                )
            )
            blocking_dependency = {
                **dependency,
                "status": "blocking",
                "reason": "来源经过未支持命令 opcode=999",
            }
            blocking_analysis = {
                **analysis,
                "dependencies": [blocking_dependency],
                "blocking_issues": [blocking_dependency],
            }
            with self.assertRaisesRegex(RuntimeError, "保守：保留风险原文后继续"):
                analyze_import_protection(
                    items,
                    ImportScope(),
                    game,
                    ImportProtectionRules(logic_unknown_policy="block"),
                    blocking_analysis,
                )
            permissive = analyze_import_protection(
                items,
                ImportScope(),
                game,
                ImportProtectionRules(logic_unknown_policy="warn"),
                blocking_analysis,
            )
            self.assertEqual(1, permissive["summary"]["logic_permissive_warnings"])
            self.assertEqual(0, permissive["summary"]["logic_risk"])
            self.assertGreaterEqual(permissive["summary"]["logic_auto_preserved"], 1)
            self.assertTrue(
                any(entry["reason"] == "logic_unresolved_scope" for entry in permissive["entries"])
            )
            forced = analyze_import_protection(
                items,
                ImportScope(),
                game,
                ImportProtectionRules(protect_logic_references=False),
                blocking_analysis,
            )
            self.assertEqual(1, forced["summary"]["logic_risk"])
            full = write_full_workbook(source, root / "full.xlsx", items)
            scoped = write_scoped_workbook(
                full,
                root / "scoped.xlsx",
                ImportScope(),
                game,
                items,
                protected_keys=set(report["protected_keys"]),
            )
            output = load_workbook(scoped).active
            self.assertIsNone(output["G2"].value)
            self.assertIsNone(output["G3"].value)
            self.assertEqual("戦_HPバー", output["G4"].value)
            self.assertIsNone(output["G5"].value)
            self.assertEqual("重新启动_1", output["G6"].value)

    def test_copy_mixed_scope_group_can_follow_the_display_source(self):
        source = TranslationItem(
            key="source",
            original="装填レバー",
            translation="装填杆",
            code="UDB-11-21-0",
            type="武装一覧",
        )
        condition = TranslationItem(
            key="condition",
            original="装填レバー",
            code="COMMON-75-1016-2",
            flag="COPY-FROM-UDB-11-21-0",
            context="Event | CEv75 [index1016](DB Management)",
            category=ImportCategory.COPY,
            copy_category=ImportCategory.OPTIONAL_NAME,
        )
        self.assertEqual({}, selected_translation_requirements([source, condition], ImportScope()))
        selected = selected_translation_requirements(
            [source, condition],
            ImportScope(),
            allow_copy_condition_groups=True,
        )
        self.assertEqual({"source"}, set(selected))
        source.translation = ""
        report = analyze_import_protection(
            [source, condition],
            ImportScope(),
            Path("missing-game"),
            ImportProtectionRules(
                allow_copy_condition_groups=True,
                protect_logic_references=False,
            ),
        )
        self.assertEqual(1, report["summary"]["atomic_groups"])

    def test_scoped_workbook_rewrites_stale_formula_in_selected_copy_source(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            full = root / "full.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(HEADERS)
            sheet.append(
                ["UDB-20-15-0", "", "Effect", "Name", "", "==▼仮連撃▼==", "==▼仮連撃▼=="]
            )
            sheet.append(
                [
                    "NAME-D-UDB-21-10",
                    "COPY-FROM-UDB-20-15-0",
                    "UDB info",
                    "Data name",
                    "",
                    "==▼仮連撃▼==",
                    "",
                ]
            )
            workbook.save(full)
            items = read_translation_items(full)
            items[0].translation = items[0].original
            scoped = write_scoped_workbook(
                full,
                root / "scoped.xlsx",
                ImportScope(),
                root / "game",
                items,
                allow_copy_condition_groups=True,
            )

            loaded = load_workbook(scoped, data_only=False)
            self.assertIsNone(loaded.active["G2"].value)
            loaded.close()
            with zipfile.ZipFile(scoped) as archive:
                sheet_xml = archive.read("xl/worksheets/sheet1.xml")
            self.assertIn(b'r="G2"', sheet_xml)
            self.assertNotIn(b"<f>", sheet_xml)

    def test_editor_auto_analysis_contract_and_runtime_match(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            auto = root / "Auto"
            basic = auto / "BasicData"
            maps = auto / "MapData"
            basic.mkdir(parents=True)
            maps.mkdir()
            common = basic / "CommonEvent.dat.Auto.txt"
            common.write_text(
                "\n".join(
                    [
                        "[COMMON_EVENT_TEXT_OUTPUT]",
                        "COMMON_EVENT_NUM=2",
                        "COMMON_ID=63",
                        "COMMON_NAME=◆[rb]操作盤3",
                        # Editor 3.713 calibration: sample CEv70 command 333 emitted
                        # 0x301 for first-line cut; RIMWING CEv92 command 156 emitted
                        # 0xA01 for cut-up-to-text; RIMWING CEv52 command 865
                        # emitted 0x900 and described replacement in pretty output.
                        "COMMAND_NUM=22",
                        "WoditorEvCOMMAND_START",
                        # Official 3.713 omits source_raw for literal assignment.
                        '[122][2,1]<0>(1600005,0)("校准")',
                        '[122][3,0]<0>(1600006,1,1600005)()',
                        # Official 3.713 can encode condition count as 0x10 | count
                        # while retaining four padded string slots.
                        '[112][2,4]<0>(17,1600006)("校准","","","")',
                        '[250][5,4]<0>(7,-3,0,332288,1600096)("","任意类型","","任意字段甲")',
                        '[121][4,0]<0>(1600016,0,0,0)()',
                        '[179][1,0]<0>(2)()',
                        '[121][4,0]<1>(1600081,1600096,1600016,0)()',
                        '[250][5,4]<1>(7,1600071,1600081,70144,1600007)("","任意类型","","")',
                        # Editor 3.713 call flags expose four numeric slots (the
                        # command selector plus three inputs) and two string inputs.
                        # Opcode 210 reserves string slot 0 for its target just like 300.
                        '[210][9,3]<1>(500008,16785444,151,0,0,0,1600007,0,1600008)("","","")',
                        # Picture operations can share a numeric CSelf slot with a
                        # tracked string without assigning that string namespace.
                        '[150][3,0]<1>(2,1600008,4)()',
                        '[122][3,0]<1>(1600006,769,1600008)()',
                        '[122][3,0]<1>(1600006,2561,1600007)()',
                        '[122][3,1]<1>(1600006,2560,0)("\\cself[7]")',
                        '[122][3,2]<1>(1600006,2304,0)("HP","")',
                        '[170][0,0]<1>()()',
                        '[122][3,1]<2>(1600006,0,0)("\\cself[6]")',
                        '[171][0,0]<2>()()',
                        '[498][0,0]<1>()()',
                        # RIMWING CEv92 command 279 used flags=36 and had no
                        # assignment in pretty output, so its inputs stay intact.
                        '[210][8,0]<1>(600100,36,11,0,0,0,1600006,1600007)()',
                        '[112][2,1]<1>(1,538470918)("HPバー")',
                        '[121][4,0]<1>(1600016,1,0,256)()',
                        '[498][0,0]<0>()()',
                        "WoditorEvCOMMAND_END",
                        "COMMON_ID=8",
                        "COMMON_NAME=Formatter",
                        "VALINPUT_NUM=4",
                        "STRINPUT_NUM=2",
                        "RETURN_VAL_TARGET=5",
                        "COMMAND_NUM=2",
                        "WoditorEvCOMMAND_START",
                        '[212][0,1]<0>()("cmd:151")',
                        '[213][0,1]<0>()("END")',
                        "WoditorEvCOMMAND_END",
                    ]
                ),
                encoding="utf-8",
            )
            (maps / "Test.mps.Auto.txt").write_text(
                "\n".join(
                    [
                        "[MAPDATA_TEXT_OUTPUT]",
                        "EVENT_NUM=1",
                        "EVENT_ID=1",
                        "EVENT_NAME=Map event",
                        "COMMAND_NUM=11",
                        "WoditorEvCOMMAND_START",
                        '[179][1,0]<0>(2)()',
                        '[999][0,1]<1>()("未知,\\路径")',
                        '[101][0,1]<1>()("文章")',
                        '[103][0,1]<1>()("注释")',
                        '[106][0,1]<1>()("调试")',
                        '[140][6,1]<1>(33554465,0,0,0,100,100)("\\cself[8]")',
                        '[150][11,1]<1>(0,1600033,0,2,1,1,255,1600031,1600032,100,0)("picture.png")',
                        '[212][0,1]<1>()("label")',
                        # A familiar opcode with an uncalibrated shape must not
                        # enter the specialized string-assignment transfer.
                        '[122][1,0]<1>(1600099)()',
                        '[213][0,1]<1>()("END")',
                        '[498][0,0]<0>()()',
                        "WoditorEvCOMMAND_END",
                    ]
                ),
                encoding="utf-8",
            )
            (basic / "DataBase.Auto.txt").write_text(
                "\n".join(
                    [
                        "[DATABASE_TEXT_OUTPUT]",
                        "TYPE_NUM=1",
                        "TYPE_ID=7",
                        "ITEM_NUM=3",
                        "DATATYPE_0=2000",
                        "DATATYPE_1=2001",
                        "DATATYPE_2=2002",
                        "DATA_NUM=2",
                        "TYPENAME=任意类型",
                        "ITEMNAME_NUM=3",
                        "ITEMNAME0=名称",
                        "ITEMNAME1=任意字段甲",
                        "ITEMNAME2=无关字段",
                        "<<--CSV_START-->>",
                        '"名称","任意字段甲","无关字段",',
                        '"一","戦_HPバー",",\\空",',
                        '"二","戦_HPバー2","通常",',
                        "<<--CSV_END-->>",
                    ]
                ),
                encoding="utf-8",
            )
            items = [
                TranslationItem(
                    key="hp1",
                    original="戦_HPバー",
                    code="UDB-7-0-1",
                    type="任意类型",
                    info="[0:一] [任意字段甲]",
                ),
                TranslationItem(
                    key="hp2",
                    original="戦_HPバー2",
                    code="UDB-7-1-1",
                    type="任意类型",
                    info="[1:二] [任意字段甲]",
                ),
                TranslationItem(
                    key="condition",
                    original="HPバー",
                    code="COMMON-63-19-0",
                    type="事件",
                    info="条件",
                ),
                TranslationItem(
                    key="calibration_source",
                    original="校准",
                    code="COMMON-63-0-0",
                    type="事件",
                    info="字符串赋值",
                ),
                TranslationItem(
                    key="calibration_condition",
                    original="校准",
                    code="COMMON-63-2-0",
                    type="事件",
                    info="字符串条件",
                ),
                TranslationItem(key="plain", original="通常表示", type="画面", info="名称"),
            ]
            editor_path = root / "Editor.exe"
            editor_path.write_bytes(b"editor")
            editor = EditorInfo(editor_path, "3.713.2026.718", (3, 713, 2026, 718), "a" * 64)
            report = analyze_auto_export(auto, items, editor, input_hash="input")
            self.assertEqual("editor-analysis", report["kind"])
            self.assertEqual(ARTIFACT_EPOCH, report["epoch"])
            self.assertIn("event_summaries", report)
            self.assertIn("call_graph", report)
            self.assertNotIn("translated_replay", report)
            dependency = next(
                item for item in report["dependencies"] if item["literal"] == "HPバー"
            )
            self.assertEqual({"hp1", "hp2"}, set(dependency["source_keys"]))
            self.assertEqual(["condition"], dependency["condition_keys"])
            self.assertEqual("contains", dependency["operator"])
            self.assertEqual("BasicData/CommonEvent.dat.Auto.txt", dependency["auto_file"])
            self.assertEqual({1}, {cell["field"] for cell in dependency["database_cells"]})
            self.assertTrue(any("common=8 cmd=151" in entry for entry in dependency["trace"]))
            self.assertEqual(1, report["counts"]["map_maps"])
            self.assertEqual(
                [(122, 1), (999, 1)],
                [(entry["opcode"], entry["count"]) for entry in report["unknown_commands"]],
            )
            opaque = next(
                item
                for item in report["blocking_issues"]
                if item.get("kind") == "opaque" and "opcode=122" in item["reason"]
            )
            self.assertEqual(["project"], opaque["unresolved_scopes"])

            common.write_text(common.read_text(encoding="utf-8").replace("任意类型", "English Type").replace("任意字段甲", "field_name"), encoding="utf-8")
            database = basic / "DataBase.Auto.txt"
            database.write_text(database.read_text(encoding="utf-8").replace("任意类型", "English Type").replace("任意字段甲", "field_name"), encoding="utf-8")
            renamed = analyze_auto_export(auto, items, editor, input_hash="input")
            renamed_hp = next(
                item for item in renamed["dependencies"] if item["literal"] == "HPバー"
            )
            self.assertEqual({"hp1", "hp2"}, set(renamed_hp["source_keys"]))

            database_text = database.read_text(encoding="utf-8")
            database.write_text(database_text.replace("DATA_NUM=2", "DATA_NUM=3"), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "DATA_NUM"):
                analyze_auto_export(auto, items, editor, input_hash="input")
            database.write_text(database_text, encoding="utf-8")

            common.write_text(common.read_text(encoding="utf-8").replace("COMMAND_NUM=22", "COMMAND_NUM=23", 1), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "COMMAND_NUM"):
                analyze_auto_export(auto, items, editor, input_hash="input")

    def test_editor_cfg_handles_cross_structure_gotos_without_recursion(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            basic = root / "Auto" / "BasicData"
            basic.mkdir(parents=True)
            (basic / "CommonEvent.dat.Auto.txt").write_text(
                "\n".join(
                    [
                        "[COMMON_EVENT_TEXT_OUTPUT]",
                        "COMMON_EVENT_NUM=2",
                        "COMMON_ID=1",
                        "COMMON_NAME=Nested back edge",
                        "COMMAND_NUM=10",
                        "WoditorEvCOMMAND_START",
                        '[212][0,1]<0>()("start")',
                        '[112][2,1]<0>(1,1600000)("outer")',
                        '[401][1,0]<0>(0)()',
                        '[170][0,0]<1>()()',
                        '[112][2,1]<2>(1,1600001)("inner")',
                        '[401][1,0]<2>(0)()',
                        '[213][0,1]<3>()("start")',
                        '[499][0,0]<2>()()',
                        '[498][0,0]<1>()()',
                        '[499][0,0]<0>()()',
                        "WoditorEvCOMMAND_END",
                        "COMMON_ID=2",
                        "COMMON_NAME=Jump into loop",
                        "COMMAND_NUM=5",
                        "WoditorEvCOMMAND_START",
                        '[213][0,1]<0>()("inside")',
                        '[179][1,0]<0>(2)()',
                        '[212][0,1]<1>()("inside")',
                        '[171][0,0]<1>()()',
                        '[498][0,0]<0>()()',
                        "WoditorEvCOMMAND_END",
                    ]
                ),
                encoding="utf-8",
            )
            editor_path = root / "Editor.exe"
            editor_path.write_bytes(b"editor")
            editor = EditorInfo(
                editor_path,
                "3.713.2026.718",
                (3, 713, 2026, 718),
                "a" * 64,
            )

            report = analyze_auto_export(root / "Auto", [], editor, input_hash="input")

            self.assertEqual("editor-analysis", report["kind"])
            self.assertEqual(ARTIFACT_EPOCH, report["epoch"])
            self.assertFalse(
                any("固定点超过" in issue["reason"] for issue in report["blocking_issues"])
            )

    def test_editor_cfg_widens_large_forward_join_without_blocking(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            basic = root / "Auto" / "BasicData"
            basic.mkdir(parents=True)
            commands: list[str] = []
            items: list[TranslationItem] = []
            for branch in range(70):
                start = len(commands)
                commands.extend((
                    "[111][4,0]<0>(1,1600000,0,0)()",
                    "[401][1,0]<0>(0)()",
                    f'[122][2,1]<1>(1600005,0)("Text {branch}")',
                    '[213][0,1]<1>()("join")',
                    "[420][1,0]<0>(0)()",
                    "[499][0,0]<0>()()",
                ))
                items.append(TranslationItem(
                    key=f"branch-{branch}",
                    original=f"Text {branch}",
                    translation=f"译文 {branch}",
                    code=f"COMMON-1-{start + 2}-0",
                ))
            commands.extend(('[212][0,1]<0>()("join")', '[101][0,1]<0>()("\\s[5]")'))
            (basic / "CommonEvent.dat.Auto.txt").write_text(
                "\n".join((
                    "[COMMON_EVENT_TEXT_OUTPUT]",
                    "COMMON_EVENT_NUM=1",
                    "COMMON_ID=1",
                    "COMMON_NAME=Large join",
                    f"COMMAND_NUM={len(commands)}",
                    "WoditorEvCOMMAND_START",
                    *commands,
                    "WoditorEvCOMMAND_END",
                )),
                encoding="utf-8",
            )
            editor_path = root / "Editor.exe"
            editor_path.write_bytes(b"editor")
            report = analyze_auto_export(
                root / "Auto",
                items,
                EditorInfo(
                    editor_path,
                    "3.713.2026.718",
                    (3, 713, 2026, 718),
                    "a" * 64,
                ),
                input_hash="input",
            )

            self.assertEqual(1.0, report["command_catalog"]["cfg_coverage"]["ratio"])
            self.assertFalse(report["blocking_issues"])

    def test_editor_cfg_queues_a_changing_forward_join_once(self):
        block = _CommandBlock(
            "Auto",
            "common",
            1,
            "join",
            1,
            tuple(_Command(0, (), (), 0, "") for _ in range(4)),
        )
        analyzer = _BlockAnalyzer(block, {}, {}, {}, audit=_AnalysisAudit.empty())
        visits: dict[int, int] = {}

        def transfer(index, state, _exits):
            visits[index] = visits.get(index, 0) + 1
            if index in {1, 2}:
                state.numbers[1_600_000] = _NumberValue(frozenset({index}))

        successors = {
            0: ((1, 4), (2, 4)),
            1: ((3, 4),),
            2: ((3, 4),),
            3: ((None, 4),),
        }
        analyzer._transfer_command = transfer
        analyzer._cfg_successors = lambda index, _limit, _state, _exits: successors[index]
        analyzer._basic_block_starts = frozenset({0, 1, 2, 3})
        state = _AnalysisState({}, {}, {})

        self.assertTrue(analyzer._execute(0, 4, state))
        self.assertEqual(1, visits[3])
        self.assertEqual(frozenset({1, 2}), state.numbers[1_600_000].values)

    def test_semantic_ledger_covers_choice_database_and_no_return_call(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            basic = root / "Auto" / "BasicData"
            basic.mkdir(parents=True)
            (basic / "CommonEvent.dat.Auto.txt").write_text(
                "\n".join(
                    (
                        "[COMMON_EVENT_TEXT_OUTPUT]",
                        "COMMON_EVENT_NUM=2",
                        "COMMON_ID=1",
                        "COMMON_NAME=Caller",
                        "COMMAND_NUM=8",
                        "WoditorEvCOMMAND_START",
                        '[102][1,2]<0>(0)("甲","乙")',
                        "[401][1,0]<0>(0)()",
                        "[250][5,0]<1>(0,0,0,0,1600000)()",
                        "[210][2,0]<1>(500002,0)()",
                        "[104][0,0]<1>()()",
                        "[401][1,0]<0>(1)()",
                        '[101][0,1]<1>()("分支乙")',
                        "[499][0,0]<0>()()",
                        "WoditorEvCOMMAND_END",
                        "COMMON_ID=2",
                        "COMMON_NAME=Callee",
                        "COMMAND_NUM=1",
                        "WoditorEvCOMMAND_START",
                        '[122][2,1]<0>(2000000,0)("全局值")',
                        "WoditorEvCOMMAND_END",
                    )
                ),
                encoding="utf-8",
            )
            (basic / "CDataBase.Auto.txt").write_text(
                "\n".join(
                    (
                        "[DATABASE_TEXT_OUTPUT]",
                        "TYPE_NUM=1",
                        "TYPE_ID=0",
                        "ITEM_NUM=1",
                        "DATATYPE_0=0",
                        "DATA_NUM=1",
                        "TYPENAME=Numbers",
                        "ITEMNAME_NUM=1",
                        "ITEMNAME0=Value",
                        "<<--CSV_START-->>",
                        '"Value",',
                        '"7","row",',
                        "<<--CSV_END-->>",
                    )
                ),
                encoding="utf-8",
            )
            editor_path = root / "Editor.exe"
            editor_path.write_bytes(b"editor")
            report = analyze_auto_export(
                root / "Auto",
                [],
                EditorInfo(
                    editor_path,
                    "3.713.2026.718",
                    (3, 713, 2026, 718),
                    "a" * 64,
                ),
                input_hash="input",
            )

            catalog = report["command_catalog"]
            for field in (
                "shape_coverage",
                "semantic_coverage",
                "cfg_coverage",
                "call_target_coverage",
                "data_effect_coverage",
            ):
                self.assertEqual(1.0, catalog[field]["ratio"], field)
            self.assertEqual(0, catalog["opaque_effects"])
            choice_edges = [
                edge
                for edge in report["runtime_semantics"]["cfg_edges"]
                if edge[0].endswith(":1")
            ]
            self.assertGreaterEqual(len(choice_edges), 2)
            self.assertIn(
                "END",
                {
                    edge[1]
                    for edge in report["runtime_semantics"]["cfg_edges"]
                    if edge[0].endswith(":5")
                },
            )
            self.assertFalse(
                any(
                    dependency["kind"] == "call"
                    and "无返回公共事件副作用" in dependency["reason"]
                    for dependency in report["dependencies"]
                )
            )

    def test_call_ledger_merges_repeated_contexts_order_independently(self):
        block = _CommandBlock("Auto", "common", 1, "caller", 1, ())
        left = _BlockAnalyzer(block, {}, {}, {}, audit=_AnalysisAudit.empty())
        right = _BlockAnalyzer(block, {}, {}, {}, audit=_AnalysisAudit.empty())
        left._record_call("call", "exact", ("common:1",))
        left._record_call("call", "exact", ("common:2",))
        right._record_call("call", "exact", ("common:2",))
        right._record_call("call", "exact", ("common:1",))
        self.assertEqual(
            ("exact", ("common:1", "common:2")), left.audit.calls["call"]
        )
        self.assertEqual(left.audit.calls, right.audit.calls)
        self.assertEqual(left.audit.data_effects, right.audit.data_effects)

    def test_no_return_numeric_alias_call_uses_public_event_analysis(self):
        database = _DatabaseType(
            "CDB", 0, "", {0: "text"}, {0: 2000}, (("",),), ("row",)
        )
        callee = _CommandBlock(
            "Auto",
            "common",
            2,
            "callee",
            1,
            (
                _Command(
                    250,
                    (0, 1_600_000, 0, 0x1000, 1_600_001),
                    ("",) * 4,
                    0,
                    "",
                ),
            ),
        )
        caller = _CommandBlock("Auto", "common", 1, "caller", 1, ())
        analyzer = _BlockAnalyzer(
            caller,
            {"CDB": {0: database}},
            {},
            {},
            {2: callee},
            audit=_AnalysisAudit.empty(),
        )
        call = _Command(210, (500_002, 1, 1_600_000), (), 0, "")

        with mock.patch.object(
            _BlockAnalyzer,
            "_execute",
            side_effect=AssertionError("numeric-only no-return call was expanded"),
        ):
            analyzer._call_event(call, 0, _AnalysisState({}, {}, {}))

        self.assertEqual(
            ("exact", ("common:2",)), analyzer.audit.calls["Auto:common:1:1:1"]
        )

    def test_literal_no_return_calls_keep_per_call_display_provenance(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            basic = root / "Auto" / "BasicData"
            basic.mkdir(parents=True)
            calls = [
                f'[300][3,2]<0>(0,4112,0)("Callee","Text {index}")'
                for index in range(260)
            ]
            (basic / "CommonEvent.dat.Auto.txt").write_text(
                "\n".join(
                    [
                        "[COMMON_EVENT_TEXT_OUTPUT]",
                        "COMMON_EVENT_NUM=2",
                        "COMMON_ID=1",
                        "COMMON_NAME=Caller",
                        f"COMMAND_NUM={len(calls)}",
                        "WoditorEvCOMMAND_START",
                        *calls,
                        "WoditorEvCOMMAND_END",
                        "COMMON_ID=2",
                        "COMMON_NAME=Callee",
                        "COMMAND_NUM=1",
                        "WoditorEvCOMMAND_START",
                        '[101][0,1]<0>()("\\cself[5]")',
                        "WoditorEvCOMMAND_END",
                    ]
                ),
                encoding="utf-8",
            )
            items = [
                TranslationItem(
                    key=f"call-{index}",
                    original=f"Text {index}",
                    translation=f"译文 {index}",
                    code=f"COMMON-1-{index}-1",
                )
                for index in range(260)
            ]
            editor_path = root / "Editor.exe"
            editor_path.write_bytes(b"editor")
            report = analyze_auto_export(
                root / "Auto",
                items,
                EditorInfo(
                    editor_path,
                    "3.713.2026.718",
                    (3, 713, 2026, 718),
                    "a" * 64,
                ),
                input_hash="input",
            )

            self.assertEqual(260, len(set(report["safe_to_translate"])))
            display_sources = {
                tuple(dependency["source_keys"])
                for dependency in report["dependencies"]
                if dependency["kind"] == "display" and dependency["source_keys"]
            }
            self.assertEqual(
                {(f"call-{index}",) for index in range(260)}, display_sources
            )
            self.assertEqual([], report["blocking_issues"])
            self.assertEqual(0, report["command_catalog"]["opaque_effects"])

    def test_translation_safety_protects_picture_path_from_string_variable(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            basic = root / "Auto" / "BasicData"
            basic.mkdir(parents=True)
            (basic / "CommonEvent.dat.Auto.txt").write_text(
                "\n".join(
                    (
                        "[COMMON_EVENT_TEXT_OUTPUT]",
                        "COMMON_EVENT_NUM=1",
                        "COMMON_ID=391",
                        "COMMON_NAME=Picture",
                        "COMMAND_NUM=3",
                        "WoditorEvCOMMAND_START",
                        '[122][3,1]<0>(1600009,0,0)("000すぽっと.png")',
                        '[122][3,1]<0>(1600009,0,0)("1枚絵マップ/\\cself[9]")',
                        "[150][12,0]<0>(16,1600010,0,1,1,1,1600014,0,0,1600011,0,1600009)()",
                        "WoditorEvCOMMAND_END",
                    )
                ),
                encoding="utf-8",
            )
            item = TranslationItem(
                key="picture_variable",
                original="1枚絵マップ/\\cself[9]",
                translation="静止画地图/\\cself[9]",
                code="COMMON-391-1-0",
            )
            editor_path = root / "Editor.exe"
            editor_path.write_bytes(b"editor")
            editor = EditorInfo(
                editor_path,
                "3.713.2026.718",
                (3, 713, 2026, 718),
                "a" * 64,
            )

            analysis = analyze_auto_export(
                root / "Auto", [item], editor, input_hash="input"
            )
            self.assertEqual(["resource"], analysis["usage_by_key"][item.key])
            dependency = next(
                dependency
                for dependency in analysis["dependencies"]
                if dependency.get("resource_role") == "resource_path_variable"
            )
            self.assertEqual([item.key], dependency["source_keys"])
            self.assertEqual(
                ["1枚絵マップ/000すぽっと.png"], dependency["source_values"]
            )

            safety = analyze_translation_safety(
                root / "Auto",
                [item],
                {item.key: item.translation},
                "warn",
                analysis=analysis,
            )
            self.assertEqual([item.key], safety["keep_original"])
            self.assertIn("resource", safety["reasons"][item.key])
            self.assertNotIn(
                item.key, safety["approvals"]["official_display_contract"]
            )

    def test_translation_safety_uses_official_display_contract_and_auto_proof(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            basic = root / "Auto" / "BasicData"
            basic.mkdir(parents=True)
            (basic / "CommonEvent.dat.Auto.txt").write_text(
                "\n".join(
                    (
                        "[COMMON_EVENT_TEXT_OUTPUT]",
                        "COMMON_EVENT_NUM=1",
                        "COMMON_ID=1",
                        "COMMON_NAME=Safety",
                        "COMMAND_NUM=8",
                        "WoditorEvCOMMAND_START",
                        '[101][0,1]<0>()("表示文本")',
                        '[122][2,1]<0>(1600005,0)("HP")',
                        '[112][2,1]<0>(1,1600005)("HP")',
                        '[122][2,1]<0>(1600006,0)("操作レバー")',
                        '[150][11,1]<0>(32,0,0,0,0,0,0,0,0,0,0)("\\cself[6]")',
                        '[122][2,1]<0>(1600007,0)("Picture/file.png")',
                        '[150][11,1]<0>(0,0,0,0,0,0,0,0,0,0,0)("\\cself[7]")',
                        '[122][2,1]<0>(2000000,0)("全局文本")',
                        "WoditorEvCOMMAND_END",
                    )
                ),
                encoding="utf-8",
            )
            items = [
                TranslationItem(
                    key="display",
                    original="表示文本",
                    translation="显示文本",
                    code="COMMON-1-0-0",
                ),
                TranslationItem(
                    key="logic_source",
                    original="HP",
                    translation="生命",
                    code="COMMON-1-1-0",
                ),
                TranslationItem(
                    key="logic_literal",
                    original="HP",
                    translation="生命",
                    code="COMMON-1-2-0",
                ),
                TranslationItem(
                    key="unmapped",
                    original="系统参数",
                    translation="系统参数译文",
                    code="UDB-9-9-9",
                ),
                TranslationItem(
                    key="unchanged",
                    original="无需改动",
                    translation="无需改动",
                    code="UDB-9-9-10",
                ),
                TranslationItem(
                    key="picture_text",
                    original="操作レバー",
                    translation="操作杆",
                    code="COMMON-1-3-0",
                ),
                TranslationItem(
                    key="picture_file",
                    original="Picture/file.png",
                    translation="Picture/translated.png",
                    code="COMMON-1-5-0",
                ),
                TranslationItem(
                    key="dynamic_safe",
                    original="装填レバー",
                    translation="装填杆",
                    code="UDB-7-0-0",
                ),
                TranslationItem(
                    key="dynamic_unsafe",
                    original="戦_HPバー",
                    translation="战斗_HP槽",
                    code="UDB-7-1-0",
                ),
                TranslationItem(
                    key="global_state",
                    original="全局文本",
                    translation="全局译文",
                    code="COMMON-1-7-0",
                ),
            ]
            editor_path = root / "Editor.exe"
            editor_path.write_bytes(b"editor")
            editor = EditorInfo(
                editor_path,
                "3.713.2026.718",
                (3, 713, 2026, 718),
                "a" * 64,
            )
            analysis = analyze_auto_export(
                root / "Auto", items, editor, input_hash="input"
            )
            analysis["usage_by_key"].update({
                "dynamic_safe": ["display_only", "logic"],
                "dynamic_unsafe": ["display_only", "logic"],
            })
            dynamic_dependency = {
                "kind": "condition",
                "condition_keys": [],
                "source_keys": [],
                "right_source_keys": [],
                "source_scopes": ["database:UDB:7:*:0"],
                "right_source_scopes": [],
                "unresolved_scopes": ["database:UDB:7:*:0"],
                "right_is_variable": False,
                "left_values": [],
                "right_values": [],
                "status": "dynamic",
                "reason": "数据库字符串来源集合超过 256 项",
            }
            analysis["dependencies"].extend((
                {**dynamic_dependency, "operator": "not_equals", "literal": ""},
                {**dynamic_dependency, "operator": "contains", "literal": "HPバー"},
            ))
            safety = analyze_translation_safety(
                root / "Auto",
                items,
                {item.key: item.translation for item in items},
                "warn",
                analysis=analysis,
            )
            self.assertEqual(
                [
                    "display",
                    "dynamic_safe",
                    "global_state",
                    "picture_text",
                    "unmapped",
                ],
                safety["safe_to_translate"],
            )
            self.assertEqual(
                {
                    "dynamic_unsafe",
                    "logic_literal",
                    "logic_source",
                    "picture_file",
                },
                set(safety["keep_original"]),
            )
            self.assertIn(
                "unmapped", safety["approvals"]["official_display_contract"]
            )
            self.assertEqual(2, safety["replay"]["iterations"])
            self.assertTrue(safety["replay"]["control_flow_equivalent"])
            protection = analyze_import_protection(
                items,
                ImportScope(),
                root,
                ImportProtectionRules(),
                analysis,
                logic_safety=safety,
            )
            self.assertEqual("import-protection", protection["kind"])
            self.assertEqual(ARTIFACT_EPOCH, protection["epoch"])
            self.assertEqual(
                [
                    "display",
                    "dynamic_safe",
                    "global_state",
                    "picture_text",
                    "unmapped",
                ],
                protection["safe_to_translate"],
            )
            self.assertNotIn("unchanged", protection["keep_original"])

    def test_translation_safety_allows_display_database_transport(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            basic = root / "Auto" / "BasicData"
            basic.mkdir(parents=True)
            (basic / "CommonEvent.dat.Auto.txt").write_text(
                "\n".join(
                    (
                        "[COMMON_EVENT_TEXT_OUTPUT]",
                        "COMMON_EVENT_NUM=1",
                        "COMMON_ID=1",
                        "COMMON_NAME=Display transport",
                        "COMMAND_NUM=4",
                        "WoditorEvCOMMAND_START",
                        # Write a display payload into a text field, read it back,
                        # then pass it to a verified message command.
                        '[122][3,1]<0>(1600002,0,0)("状态提示")',
                        '[250][5,4]<0>(0,0,0,328192,1600002)("","可视文本","","正文")',
                        '[250][5,4]<0>(0,0,0,332288,1600001)("","可视文本","","正文")',
                        '[101][0,1]<0>()("\\cself[1]")',
                        "WoditorEvCOMMAND_END",
                    )
                ),
                encoding="utf-8",
            )
            (basic / "DataBase.Auto.txt").write_text(
                "\n".join(
                    (
                        "[DATABASE_TEXT_OUTPUT]",
                        "TYPE_NUM=1",
                        "TYPE_ID=7",
                        "ITEM_NUM=2",
                        "DATATYPE_0=2000",
                        "DATATYPE_1=2000",
                        "DATA_NUM=1",
                        "TYPENAME=可视文本",
                        "ITEMNAME_NUM=2",
                        "ITEMNAME0=名称",
                        "ITEMNAME1=正文",
                        "<<--CSV_START-->>",
                        '"名称","正文",',
                        '"展示条目","",',
                        "<<--CSV_END-->>",
                    )
                ),
                encoding="utf-8",
            )
            item = TranslationItem(
                key="display_payload",
                original="状态提示",
                translation="状态说明",
                code="COMMON-1-0-0",
                category=ImportCategory.EXTERNAL,
            )
            editor_path = root / "Editor.exe"
            editor_path.write_bytes(b"editor")
            editor = EditorInfo(
                editor_path, "3.713.2026.718", (3, 713, 2026, 718), "a" * 64
            )
            analysis = analyze_auto_export(
                root / "Auto", [item], editor, input_hash="input"
            )
            self.assertEqual(
                ["display_only", "display_storage"],
                analysis["usage_by_key"][item.key],
            )
            self.assertTrue(
                any(
                    dependency.get("resource_role") == "database_string_write"
                    and dependency["source_keys"] == [item.key]
                    for dependency in analysis["dependencies"]
                )
            )
            safety = analyze_translation_safety(
                root / "Auto",
                [item],
                {item.key: item.translation},
                "warn",
                analysis=analysis,
            )
            self.assertEqual([item.key], safety["safe_to_translate"])
            self.assertEqual([], safety["keep_original"])
            self.assertTrue(safety["replay"]["control_flow_equivalent"])

    def test_database_storage_tracks_display_sink_and_direct_logic(self):
        item = TranslationItem(
            key="display_payload",
            original="状态提示",
            translation="状态说明",
            code="COMMON-1-0-0",
        )
        writer = {
            "kind": "resource",
            "resource_role": "database_string_write",
            "source_keys": [item.key],
            "right_source_keys": [],
            "condition_keys": [],
            "unresolved_scopes": ["database:CDB:58:*:1"],
            "target_database_cells": [
                {"database": "CDB", "type": 58, "data": 0, "field": 1}
            ],
        }
        display = {
            "kind": "display",
            "source_keys": [],
            "right_source_keys": [],
            "condition_keys": [],
            "database_cells": [
                {"database": "CDB", "type": 58, "data": 0, "field": 1}
            ],
        }
        usage, _ = _translation_usage_report((), [item], [writer, display])
        self.assertEqual(
            ["display_only", "display_storage"], usage[item.key]
        )
        transport = {
            "kind": "state",
            "source_keys": [],
            "right_source_keys": [],
            "condition_keys": [],
            "database_cells": display["database_cells"],
        }
        usage, _ = _translation_usage_report((), [item], [writer, display, transport])
        self.assertEqual(
            ["display_only", "display_storage"], usage[item.key]
        )
        self.assertTrue(writer["display_sink_proven"])
        logic = {
            "kind": "condition",
            "source_keys": [item.key],
            "right_source_keys": [],
            "condition_keys": [],
            "database_cells": display["database_cells"],
        }
        usage, _ = _translation_usage_report((), [item], [writer, display, logic])
        self.assertEqual(
            ["display_storage", "logic"], usage[item.key]
        )
        self.assertFalse(writer["display_sink_proven"])

    def test_database_display_sink_accepts_same_field_across_rows(self):
        item = TranslationItem(
            key="display_payload",
            original="状态提示",
            translation="状态说明",
            code="COMMON-1-0-0",
        )
        writer = {
            "kind": "resource",
            "resource_role": "database_string_write",
            "source_keys": [item.key],
            "right_source_keys": [],
            "condition_keys": [],
            "target_database_cells": [
                {"database": "CDB", "type": 58, "data": 1, "field": 1}
            ],
        }
        display = {
            "kind": "display",
            "source_keys": [],
            "right_source_keys": [],
            "condition_keys": [],
            "database_cells": [
                {"database": "CDB", "type": 58, "data": 0, "field": 1}
            ],
        }

        usage, _ = _translation_usage_report((), [item], [writer, display])

        self.assertEqual(["display_only", "display_storage"], usage[item.key])
        self.assertTrue(writer["display_sink_proven"])
        self.assertEqual("database_field", writer["display_sink_basis"])

    def test_database_storage_splits_unproven_sink_by_source_contract(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            basic = root / "Auto" / "BasicData"
            basic.mkdir(parents=True)
            (basic / "CommonEvent.dat.Auto.txt").write_text(
                "[COMMON_EVENT_TEXT_OUTPUT]\nCOMMON_EVENT_NUM=0\n",
                encoding="utf-8",
            )
            display = TranslationItem(
                key="display_name",
                original="ウィリアム",
                translation="威廉",
                code="UDB-16-1-0",
                category=ImportCategory.DISPLAY,
            )
            internal = TranslationItem(
                key="internal_name",
                original="InternalKey",
                translation="内部键",
                code="COMMON-1-0-0",
                category=ImportCategory.EXTERNAL,
            )
            items = [display, internal]
            analysis = analyze_auto_export(
                root / "Auto",
                items,
                EditorInfo(
                    root / "Editor.exe",
                    "3.713.2026.718",
                    (3, 713, 2026, 718),
                    "a" * 64,
                ),
                input_hash="input",
            )
            analysis["usage_by_key"] = {
                display.key: ["display_only", "display_storage"],
                internal.key: ["display_storage"],
            }
            analysis["dependencies"] = [{
                "kind": "resource",
                "resource_role": "database_string_write",
                "source_keys": [display.key, internal.key],
                "right_source_keys": [],
                "condition_keys": [],
                "display_sink_proven": False,
            }]

            with mock.patch("wolf_editor.analyze_auto_export", return_value=analysis):
                safety = analyze_translation_safety(
                    root / "Auto",
                    items,
                    {item.key: item.translation for item in items},
                    "warn",
                    analysis=analysis,
                )

            self.assertEqual([display.key], safety["safe_to_translate"])
            self.assertEqual([internal.key], safety["keep_original"])
            self.assertIn(
                "database_storage_without_display_sink",
                safety["reasons"][internal.key],
            )

    def test_dynamic_database_sink_requires_same_selector_and_display_only(self):
        item = TranslationItem(
            key="display_payload",
            original="状态提示",
            translation="状态说明",
            code="COMMON-1-0-0",
        )
        selector = {
            "database": "CDB",
            "type": 58,
            "field": 1,
            "selector": "event-input:1600022",
            "auto_file": "CommonEvent.dat.Auto.txt",
            "event_type": "common",
            "event_id": 44,
            "page": 0,
        }
        writer = {
            "kind": "resource",
            "resource_role": "database_string_write",
            "source_keys": [item.key],
            "right_source_keys": [],
            "condition_keys": [],
            "command": 3,
            "target_database_cells": [],
            "target_database_selectors": [selector],
        }
        display = {
            "kind": "display",
            "source_keys": [],
            "right_source_keys": [],
            "condition_keys": [],
            "command": 4,
            "database_cells": [],
            "database_selectors": [selector],
        }

        usage, _ = _translation_usage_report((), [item], [writer, display])
        self.assertTrue(writer["display_sink_proven"])
        self.assertEqual(["display_only", "display_storage"], usage[item.key])

        condition = {
            "kind": "condition",
            "source_keys": [],
            "right_source_keys": [],
            "condition_keys": [],
            "command": 5,
            "database_cells": [],
            "database_selectors": [selector],
        }
        _translation_usage_report((), [item], [writer, display, condition])
        self.assertFalse(writer["display_sink_proven"])

    def test_dynamic_database_address_terms_cross_events_without_name_rules(self):
        base = _NumberValue(None, identity="event-input:writer:1600000")
        offset = _calculate_numbers(base, _NumberValue(frozenset({2})), 0)
        self.assertEqual("add:event-input:writer:1600000:2", offset.identity)
        self.assertEqual("", _merge_numbers(offset, base).identity)
        self.assertEqual(
            "loop:const:2:1",
            _loop_identity(
                _NumberValue(frozenset({2}), identity="const:2"),
                _NumberValue(frozenset({3}), identity="add:const:2:1"),
            ),
        )

        item = TranslationItem(
            key="dynamic_payload",
            original="状态提示",
            translation="状态说明",
            code="COMMON-44-0-0",
        )
        selector = {
            "database": "CDB",
            "type": 58,
            "field": 1,
            "selector": offset.identity,
            "auto_file": "",
            "event_type": "address-expression",
            "event_id": -1,
            "page": -1,
        }
        writer = {
            "kind": "resource",
            "resource_role": "database_string_write",
            "source_keys": [item.key],
            "right_source_keys": [],
            "condition_keys": [],
            "command": 99,
            "target_database_cells": [],
            "target_database_selectors": [selector],
        }
        display = {
            "kind": "display",
            "source_keys": [],
            "right_source_keys": [],
            "condition_keys": [],
            "command": 1,
            "database_cells": [],
            "database_selectors": [selector],
        }
        usage, _ = _translation_usage_report((), [item], [writer, display])
        self.assertTrue(writer["display_sink_proven"])
        self.assertEqual(["display_only", "display_storage"], usage[item.key])

    def test_dynamic_numeric_database_address_round_trips_only_exact_selector(self):
        database = _DatabaseType(
            "CDB", 0, "", {0: "field"}, {0: 0}, (("0",),), ("row",)
        )
        block = _CommandBlock("Auto", "common", 1, "", 0, ())
        analyzer = _BlockAnalyzer(
            block, {"CDB": {0: database}}, {}, {}, audit=_AnalysisAudit.empty()
        )
        state = _AnalysisState(
            {
                1_600_000: _NumberValue(None, identity="caller-selector"),
                1_600_001: _NumberValue(None, identity="callee-selector"),
            },
            {},
            {},
        )
        writer = _Command(250, (0, 1_600_000, 0, 0, 1_600_001), ("",) * 4, 0, "")
        analyzer._write_database_number(writer, state, "CDB", 0)
        reader = _Command(250, (0, 1_600_000, 0, 0x1000, 1_600_002), ("",) * 4, 0, "")
        analyzer._database(reader, 0, state)
        self.assertEqual("callee-selector", state.numbers[1_600_002].identity)

        wrong_reader = _Command(250, (0, 1_600_001, 0, 0x1000, 1_600_003), ("",) * 4, 0, "")
        analyzer._database(wrong_reader, 1, state)
        self.assertNotEqual("callee-selector", state.numbers[1_600_003].identity)

    def test_dynamic_database_string_round_trip_survives_persistent_state(self):
        database = _DatabaseType(
            "CDB", 0, "", {0: "text"}, {0: 2000}, (("",),), ("row",)
        )
        block = _CommandBlock("Auto", "common", 1, "", 0, ())
        analyzer = _BlockAnalyzer(
            block, {"CDB": {0: database}}, {}, {}, audit=_AnalysisAudit.empty()
        )
        writer_state = _AnalysisState(
            {1_600_000: _NumberValue(None, identity="shared-selector")},
            {
                1_600_001: _StringValue(
                    source_keys=frozenset({"payload"}),
                    literals=frozenset({"translated"}),
                )
            },
            {},
        )
        writer = _Command(250, (0, 1_600_000, 0, 0, 1_600_001), ("",) * 4, 0, "")
        analyzer._database(writer, 0, writer_state)

        persistent = _merge_states([writer_state, _AnalysisState({}, {}, {})])
        reader_state = _AnalysisState(
            {1_600_000: _NumberValue(None, identity="shared-selector")},
            {},
            {},
            dynamic_database_strings=dict(persistent.dynamic_database_strings),
        )
        reader = _Command(250, (0, 1_600_000, 0, 0x1000, 1_600_002), ("",) * 4, 0, "")
        analyzer._database(reader, 1, reader_state)
        self.assertEqual(
            frozenset({"payload"}), reader_state.strings[1_600_002].source_keys
        )

        reader_state.numbers[1_600_000] = _NumberValue(None, identity="other-selector")
        analyzer._database(reader, 2, reader_state)
        self.assertNotIn("payload", reader_state.strings[1_600_002].source_keys)

    def test_dynamic_database_writer_selector_is_an_address_source(self):
        block = _CommandBlock(
            "Auto",
            "common",
            1,
            "",
            0,
            (_Command(250, (0, 1_600_022, 0, 0), ("",) * 4, 0, ""),),
        )
        self.assertEqual(frozenset({1_600_022}), _address_variables_for_block(block))

    def test_translation_safety_allows_unread_loop_file_content_only(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            basic = root / "Auto" / "BasicData"
            basic.mkdir(parents=True)
            (basic / "CommonEvent.dat.Auto.txt").write_text(
                "[COMMON_EVENT_TEXT_OUTPUT]\nCOMMON_EVENT_NUM=0\n",
                encoding="utf-8",
            )
            item = TranslationItem(
                key="credit_line",
                original="制作",
                translation="制作人员",
                code="COMMON-1-0-0",
                category=ImportCategory.DISPLAY,
            )
            editor = EditorInfo(
                root / "Editor.exe",
                "3.713.2026.718",
                (3, 713, 2026, 718),
                "a" * 64,
            )
            analysis = analyze_auto_export(root / "Auto", [item], editor, input_hash="input")
            analysis["usage_by_key"] = {item.key: ["display_only"]}
            analysis["command_catalog"] = {
                "opaque_effects": 0,
                **{
                    field: {"ratio": 1.0}
                    for field in (
                        "semantic_coverage",
                        "cfg_coverage",
                        "call_target_coverage",
                        "data_effect_coverage",
                    )
                },
            }
            analysis["global_string_flow"] = {"converged": True}
            content_write = {
                "kind": "resource",
                "resource_role": "file_content_runtime_write",
                "auto_file": "CommonEvent.dat.Auto.txt",
                "event_type": "common",
                "event_id": 1,
                "page": 0,
                "command": 4,
                "string_index": -1,
                "condition_keys": [],
                "source_keys": [item.key],
                "right_source_keys": [],
                "resource_values": [item.original],
                "resource_path_values": ["credit.txt"],
                "status": "dynamic",
                "reason": "控制流回边扩大为运行时字符串",
            }
            analysis["dependencies"] = [content_write]
            with mock.patch("wolf_editor.analyze_auto_export", return_value=analysis):
                safety = analyze_translation_safety(
                    root / "Auto", [item], {item.key: item.translation}, "warn", analysis=analysis
                )
            self.assertEqual([item.key], safety["safe_to_translate"])

            analysis["dependencies"].append({
                "kind": "resource",
                "resource_role": "file_path_runtime_read",
                "auto_file": "CommonEvent.dat.Auto.txt",
                "event_type": "common",
                "event_id": 1,
                "page": 0,
                "command": 5,
                "string_index": -1,
                "condition_keys": [],
                "source_keys": [],
                "right_source_keys": [],
                "source_values": [".\\credit.txt"],
                "status": "resolved",
                "reason": "",
            })
            with mock.patch("wolf_editor.analyze_auto_export", return_value=analysis):
                safety = analyze_translation_safety(
                    root / "Auto", [item], {item.key: item.translation}, "warn", analysis=analysis
                )
            self.assertEqual([item.key], safety["keep_original"])
            self.assertIn("file_content_not_proven_display_only", safety["reasons"][item.key])

    def test_external_text_flow_approves_only_program_proven_payload(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            basic = root / "Auto" / "BasicData"
            basic.mkdir(parents=True)
            (basic / "CommonEvent.dat.Auto.txt").write_text(
                "[COMMON_EVENT_TEXT_OUTPUT]\nCOMMON_EVENT_NUM=0\n",
                encoding="utf-8",
            )
            item = TranslationItem(
                key="external-story",
                original="@show\nold story\n@image\nface.png",
                translation="@show\nnew story\n@image\nface.png",
                code='TXTFILE-"Data\\story.txt"',
                type="Text File",
                category=ImportCategory.EXTERNAL,
            )
            editor = EditorInfo(
                root / "Editor.exe",
                "3.713.2026.718",
                (3, 713, 2026, 718),
                "a" * 64,
            )
            analysis = analyze_auto_export(
                root / "Auto", [item], editor, input_hash="input"
            )
            analysis["external_text_flows"] = [{
                "path": "Data/story.txt",
                "item_keys": [item.key],
                "mappings": [
                    {
                        "marker": "@show",
                        "dispatch_token": "@1",
                        "safe_display_sink": True,
                        "conditions": [],
                    },
                    {
                        "marker": "@image",
                        "dispatch_token": "@2",
                        "safe_display_sink": False,
                        "conditions": [],
                    },
                ],
                "section_prefixes": [],
                "structural_prefixes": [],
            }]
            analysis["external_text_flow_coverage"] = {
                "data/story.txt": {"readers": 1, "modeled": 1}
            }

            safe = analyze_translation_safety(
                root / "Auto",
                [item],
                {item.key: item.translation},
                "warn",
                analysis=analysis,
            )
            self.assertEqual([item.key], safe["approvals"]["external_text_flow"])
            self.assertEqual({}, safe["translation_overrides"])
            protection = analyze_import_protection(
                [item],
                ImportScope(external=True),
                root,
                ImportProtectionRules(),
                analysis,
                logic_safety=safe,
            )
            self.assertEqual([item.key], protection["safe_to_translate"])
            self.assertNotIn(item.key, protection["protected_keys"])

            mixed = analyze_translation_safety(
                root / "Auto",
                [item],
                {item.key: "@show\nnew story\n@image\nnew-face.png"},
                "warn",
                analysis=analysis,
            )
            self.assertEqual([], mixed["keep_original"])
            self.assertEqual(
                "@show\nnew story\n@image\nface.png",
                mixed["translation_overrides"][item.key],
            )

            protected = analyze_translation_safety(
                root / "Auto",
                [item],
                {item.key: " @show\nnew story\n@image\nface.png"},
                "warn",
                analysis=analysis,
            )
            self.assertEqual([item.key], protected["keep_original"])
            self.assertEqual({}, protected["translation_overrides"])

            analysis["external_text_flow_coverage"]["data/story.txt"]["readers"] = 2
            incomplete = analyze_translation_safety(
                root / "Auto",
                [item],
                {item.key: item.translation},
                "warn",
                analysis=analysis,
            )
            self.assertEqual([item.key], incomplete["keep_original"])
            self.assertIn(
                "external_readers_not_fully_modeled", incomplete["reasons"][item.key]
            )

    def test_external_text_observer_requires_structural_contains(self):
        flow = {
            "path": "Data/story.txt",
            "source_key": "external-file:story",
            "reader": {
                "auto_file": "BasicData/CommonEvent.dat.Auto.txt",
                "event_type": "common",
                "event_id": 2,
                "page": 1,
                "command": 1,
            },
            "section_prefixes": ["section:"],
        }
        read = {
            "resource_role": "file_path_runtime_read",
            "external_file_paths": ["Data/story.txt"],
            "auto_file": "BasicData/CommonEvent.dat.Auto.txt",
            "event_type": "common",
            "event_id": 1,
            "page": 1,
            "command": 4,
        }
        trace = (
            "BasicData/CommonEvent.dat.Auto.txt event=1 page=1 "
            "command=4 opcode=122 external-file-content"
        )
        condition = {
            "kind": "condition",
            "operator": "contains",
            "source_keys": ["external-file:story"],
            "right_source_keys": [],
            "right_templates": [r"section:\cself[0]"],
            "trace": [trace],
            "right_trace": [],
        }
        self.assertEqual(
            1, len(_external_text_observer_report([flow], [read, condition]))
        )
        condition["operator"] = "equals"
        self.assertEqual([], _external_text_observer_report([flow], [read, condition]))

    def test_scoped_workbook_uses_safe_external_translation_override(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "source.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(HEADERS)
            sheet.append([
                'TXTFILE-"Data\\story.txt"',
                "",
                "Text File",
                "",
                "",
                "@show\nold story\n@image\nface.png",
                "",
            ])
            workbook.save(source)
            items = read_translation_items(source)
            items[0].translation = "@show\nnew story\n@image\nnew-face.png"
            full = write_full_workbook(source, root / "full.xlsx", items)
            scoped = write_scoped_workbook(
                full,
                root / "scoped.xlsx",
                ImportScope(external=True),
                root,
                items,
                translation_overrides={
                    items[0].key: "@show\nnew story\n@image\nface.png"
                },
            )
            self.assertEqual(
                "@show\nnew story\n@image\nface.png",
                load_workbook(scoped).active["G2"].value,
            )

    def test_translation_safety_protects_cross_event_database_condition(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            basic = root / "Auto" / "BasicData"
            basic.mkdir(parents=True)
            (basic / "CommonEvent.dat.Auto.txt").write_text(
                "\n".join(
                    (
                        "[COMMON_EVENT_TEXT_OUTPUT]",
                        "COMMON_EVENT_NUM=3",
                        "COMMON_ID=1",
                        "COMMON_NAME=Writer",
                        "COMMAND_NUM=2",
                        "WoditorEvCOMMAND_START",
                        '[122][2,1]<0>(1600000,0)("Selector")',
                        '[250][5,4]<0>(0,0,1,328192,1600000)("","Text","","Value")',
                        "WoditorEvCOMMAND_END",
                        "COMMON_ID=2",
                        "COMMON_NAME=Display",
                        "COMMAND_NUM=2",
                        "WoditorEvCOMMAND_START",
                        '[250][5,4]<0>(0,0,1,332288,1600001)("","Text","","Value")',
                        '[101][0,1]<0>()("\\cself[1]")',
                        "WoditorEvCOMMAND_END",
                        "COMMON_ID=3",
                        "COMMON_NAME=Condition",
                        "COMMAND_NUM=2",
                        "WoditorEvCOMMAND_START",
                        '[250][5,4]<0>(0,0,1,332288,1600001)("","Text","","Value")',
                        '[112][2,1]<0>(1,1600001)("Selector")',
                        "WoditorEvCOMMAND_END",
                    )
                ),
                encoding="utf-8",
            )
            (basic / "DataBase.Auto.txt").write_text(
                "\n".join(
                    (
                        "[DATABASE_TEXT_OUTPUT]",
                        "TYPE_NUM=1",
                        "TYPE_ID=0",
                        "ITEM_NUM=2",
                        "DATATYPE_0=2000",
                        "DATATYPE_1=2000",
                        "DATA_NUM=1",
                        "TYPENAME=Text",
                        "ITEMNAME_NUM=2",
                        "ITEMNAME0=Name",
                        "ITEMNAME1=Value",
                        "<<--CSV_START-->>",
                        '"Name","Value",',
                        '"Entry","",',
                        "<<--CSV_END-->>",
                    )
                ),
                encoding="utf-8",
            )
            item = TranslationItem(
                key="writer",
                original="Selector",
                translation="选择器",
                code="COMMON-1-0-0",
                category=ImportCategory.EXTERNAL,
            )
            editor = EditorInfo(
                root / "Editor.exe",
                "3.713.2026.718",
                (3, 713, 2026, 718),
                "a" * 64,
            )
            analysis = analyze_auto_export(root / "Auto", [item], editor, input_hash="input")

            writer = next(
                dependency
                for dependency in analysis["dependencies"]
                if dependency.get("resource_role") == "database_string_write"
            )
            condition = next(
                dependency
                for dependency in analysis["dependencies"]
                if dependency["kind"] == "condition"
                and dependency["event_id"] == 3
                and dependency["source_keys"] == [item.key]
            )
            self.assertEqual(
                [{"database": "UDB", "type": 0, "data": 0, "field": 1}],
                writer["target_database_cells"],
            )
            self.assertFalse(writer["display_sink_proven"])
            self.assertEqual([item.key], condition["source_keys"])
            self.assertEqual(
                ["display_only", "display_storage", "logic"],
                analysis["usage_by_key"][item.key],
            )

            safety = analyze_translation_safety(
                root / "Auto",
                [item],
                {item.key: item.translation},
                "warn",
                analysis=analysis,
            )
            self.assertEqual([], safety["safe_to_translate"])
            self.assertEqual([item.key], safety["keep_original"])
            self.assertIn(
                "condition_truth_change", safety["reasons"][item.key]
            )

    def test_translation_safety_allows_global_state_only_without_field_names(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            basic = root / "Auto" / "BasicData"
            basic.mkdir(parents=True)
            (basic / "CommonEvent.dat.Auto.txt").write_text(
                "[COMMON_EVENT_TEXT_OUTPUT]\nCOMMON_EVENT_NUM=0\n",
                encoding="utf-8",
            )
            (basic / "DataBase.Auto.txt").write_text(
                "\n".join(
                    (
                        "[DATABASE_TEXT_OUTPUT]",
                        "TYPE_NUM=1",
                        "TYPE_ID=26",
                        "ITEM_NUM=2",
                        "DATATYPE_0=2000",
                        "DATATYPE_1=2000",
                        "DATA_NUM=1",
                        "TYPENAME=Equipment",
                        "ITEMNAME_NUM=2",
                        "ITEMNAME0=識別名",
                        "ITEMNAME1=スクリプト",
                        "<<--CSV_START-->>",
                        '"識別名","スクリプト",',
                        '"錆びたパイプ","run",',
                        "<<--CSV_END-->>",
                    )
                ),
                encoding="utf-8",
            )
            items = [
                TranslationItem(
                    key="item_name",
                    original="錆びたパイプ",
                    translation="生锈的铁管",
                    code="UDB-26-0-0",
                ),
                TranslationItem(
                    key="script_payload",
                    original="run",
                    translation="执行",
                    code="UDB-26-0-1",
                ),
            ]
            editor_path = root / "Editor.exe"
            editor_path.write_bytes(b"editor")
            editor = EditorInfo(
                editor_path,
                "3.713.2026.718",
                (3, 713, 2026, 718),
                "a" * 64,
            )
            analysis = analyze_auto_export(root / "Auto", items, editor, input_hash="input")
            analysis["usage_by_key"].update({
                "item_name": ["logic"],
                "script_payload": ["logic"],
            })
            analysis["dependencies"].extend((
                {
                    "kind": "state",
                    "resource_role": "global_string_write",
                    "global_string_variable": 2_000_000,
                    "condition_keys": [],
                    "source_keys": [],
                    "right_source_keys": [],
                    "unresolved_scopes": ["database:UDB:26:*:0"],
                    "status": "dynamic",
                    "reason": "字符串来源已扩大为可定位的运行时符号范围",
                },
                {
                    "kind": "state",
                    "resource_role": "global_string_write",
                    "global_string_variable": 2_000_001,
                    "condition_keys": [],
                    "source_keys": [],
                    "right_source_keys": [],
                    "unresolved_scopes": ["database:UDB:26:*:1"],
                    "status": "dynamic",
                    "reason": "公共事件返回值为运行时动态值",
                },
            ))

            safety = analyze_translation_safety(
                root / "Auto",
                items,
                {item.key: item.translation for item in items},
                "warn",
                analysis=analysis,
            )

            self.assertEqual(
                ["item_name", "script_payload"], safety["safe_to_translate"]
            )
            self.assertEqual([], safety["keep_original"])

            analysis["dependencies"].append({
                "kind": "condition",
                "condition_keys": [],
                "source_keys": ["item_name"],
                "right_source_keys": [],
                "operator": "equals",
                "literal": "錆びたパイプ",
                "status": "resolved",
            })
            safety = analyze_translation_safety(
                root / "Auto",
                items,
                {item.key: item.translation for item in items},
                "warn",
                analysis=analysis,
            )
            self.assertIn("item_name", safety["keep_original"])
            self.assertIn("condition_truth_change", safety["reasons"]["item_name"])

    def test_global_string_flow_protects_cross_event_condition(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            basic = root / "Auto" / "BasicData"
            basic.mkdir(parents=True)
            (basic / "CommonEvent.dat.Auto.txt").write_text(
                "\n".join(
                    (
                        "[COMMON_EVENT_TEXT_OUTPUT]",
                        "COMMON_EVENT_NUM=2",
                        "COMMON_ID=1",
                        "COMMON_NAME=Writer",
                        "COMMAND_NUM=1",
                        "WoditorEvCOMMAND_START",
                        '[122][2,1]<0>(2000000,0)("Key")',
                        "WoditorEvCOMMAND_END",
                        "COMMON_ID=2",
                        "COMMON_NAME=Reader",
                        "COMMAND_NUM=1",
                        "WoditorEvCOMMAND_START",
                        '[112][2,1]<0>(1,2000000)("Key")',
                        "WoditorEvCOMMAND_END",
                    )
                ),
                encoding="utf-8",
            )
            item = TranslationItem(
                key="writer",
                original="Key",
                translation="NewKey",
                code="COMMON-1-0-0",
            )
            editor = EditorInfo(
                root / "Editor.exe",
                "3.713.2026.718",
                (3, 713, 2026, 718),
                "a" * 64,
            )
            analysis = analyze_auto_export(root / "Auto", [item], editor, input_hash="input")
            condition = next(
                dependency
                for dependency in analysis["dependencies"]
                if dependency["kind"] == "condition"
                and dependency["event_id"] == 2
                and dependency["source_keys"] == ["writer"]
            )

            self.assertTrue(analysis["global_string_flow"]["converged"])
            self.assertGreaterEqual(analysis["global_string_flow"]["iterations"], 1)
            self.assertEqual(["writer"], condition["source_keys"])
            self.assertEqual("resolved", condition["status"])

            safety = analyze_translation_safety(
                root / "Auto",
                [item],
                {item.key: item.translation},
                "warn",
                analysis=analysis,
            )
            self.assertEqual(["writer"], safety["keep_original"])
            self.assertIn("condition_truth_change", safety["reasons"]["writer"])

    def test_editor_database_keeps_zero_field_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "DataBase.Auto.txt"
            path.write_text(
                "\n".join(
                    (
                        "[DATABASE_TEXT_OUTPUT]",
                        "TYPE_NUM=1",
                        "TYPE_ID=0",
                        "ITEM_NUM=0",
                        "DATA_NUM=1",
                        "TYPENAME=Empty",
                        "<<--CSV_START-->>",
                        "",
                        "",
                        "",
                        "<<--CSV_END-->>",
                    )
                ),
                encoding="utf-8",
            )
            databases, summary = _database_index(path, "UDB")
            self.assertEqual(1, summary["csv_rows"])
            self.assertEqual(((),), databases[0].rows)
            self.assertEqual(("",), databases[0].data_names)

    def test_dynamic_database_selector_and_reserved_event_are_conservative(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            basic = root / "Auto" / "BasicData"
            basic.mkdir(parents=True)
            (basic / "CommonEvent.dat.Auto.txt").write_text(
                "\n".join(
                    (
                        "[COMMON_EVENT_TEXT_OUTPUT]",
                        "COMMON_EVENT_NUM=2",
                        "COMMON_ID=1",
                        "COMMON_NAME=Caller",
                        "COMMAND_NUM=6",
                        "WoditorEvCOMMAND_START",
                        '[122][2,1]<0>(1600000,0)("Aya")',
                        '[112][2,4]<0>(1,1075341824)("ya","","","")',
                        '[123][2,0]<0>(1600001,0)()',
                        "[211][2,0]<0>(1600001,0)()",
                        "[121][4,0]<0>(1600002,-1,0,0)()",
                        '[250][5,4]<0>(1600002,0,0,512,1600003)("","","","")',
                        "WoditorEvCOMMAND_END",
                        "COMMON_ID=2",
                        "COMMON_NAME=Callee",
                        "COMMAND_NUM=1",
                        "WoditorEvCOMMAND_START",
                        '[101][0,1]<0>()("display")',
                        "WoditorEvCOMMAND_END",
                    )
                ),
                encoding="utf-8",
            )
            (basic / "DataBase.Auto.txt").write_text(
                "\n".join(
                    (
                        "[DATABASE_TEXT_OUTPUT]",
                        "TYPE_NUM=1",
                        "TYPE_ID=0",
                        "ITEM_NUM=1",
                        "DATATYPE_0=2000",
                        "DATA_NUM=1",
                        "TYPENAME=Names",
                        "ITEMNAME_NUM=1",
                        "ITEMNAME0=Name",
                        "<<--CSV_START-->>",
                        '"Name",',
                        '"Alice",',
                        "<<--CSV_END-->>",
                    )
                ),
                encoding="utf-8",
            )
            report = analyze_auto_export(
                root / "Auto",
                [],
                EditorInfo(
                    root / "Editor.exe",
                    "3.713.2026.718",
                    (3, 713, 2026, 718),
                    "a" * 64,
                ),
                input_hash="input",
            )
            self.assertEqual(0, report["command_catalog"]["opaque_effects"])
            self.assertEqual(1.0, report["command_catalog"]["data_effect_coverage"]["ratio"])
            self.assertTrue(
                any(item["operator"] == "ends_with" for item in report["dependencies"])
            )
            reserved = next(
                item
                for item in report["dependencies"]
                if item["reason"] == "预约公共事件目标为运行时动态值"
            )
            self.assertEqual("dynamic", reserved["status"])
            self.assertEqual(["common:*"], reserved["unresolved_scopes"])
            self.assertEqual("numeric_id", reserved["call_target_kind"])
            self.assertEqual([], report["blocking_issues"])

    def test_dynamic_reserved_numeric_target_does_not_protect_common_display(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            basic = root / "Auto" / "BasicData"
            basic.mkdir(parents=True)
            (basic / "CommonEvent.dat.Auto.txt").write_text(
                "\n".join(
                    (
                        "[COMMON_EVENT_TEXT_OUTPUT]",
                        "COMMON_EVENT_NUM=1",
                        "COMMON_ID=1",
                        "COMMON_NAME=Caller",
                        "COMMAND_NUM=2",
                        "WoditorEvCOMMAND_START",
                        "[211][2,0]<0>(1600001,0)()",
                        '[122][2,1]<0>(1600000,0)("Player text")',
                        "WoditorEvCOMMAND_END",
                    )
                ),
                encoding="utf-8",
            )
            item = TranslationItem(
                key="message",
                original="Player text",
                translation="玩家文本",
                code="COMMON-1-1-0",
                category=ImportCategory.DISPLAY,
            )
            editor = EditorInfo(
                root / "Editor.exe",
                "3.713.2026.718",
                (3, 713, 2026, 718),
                "a" * 64,
            )
            report = analyze_auto_export(root / "Auto", [item], editor, input_hash="input")
            reserved = next(
                dependency
                for dependency in report["dependencies"]
                if dependency.get("call_target_kind") == "numeric_id"
            )
            reserved["reason"] = "reason text must not select this rule"

            safety = analyze_translation_safety(
                root / "Auto",
                [item],
                {item.key: item.translation},
                "warn",
                analysis=report,
            )

            self.assertEqual(["message"], safety["safe_to_translate"])
            self.assertEqual([], safety["keep_original"])

    def test_static_reserved_event_does_not_protect_unrelated_database_display(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            basic = root / "Auto" / "BasicData"
            basic.mkdir(parents=True)
            (basic / "CommonEvent.dat.Auto.txt").write_text(
                "\n".join(
                    (
                        "[COMMON_EVENT_TEXT_OUTPUT]",
                        "COMMON_EVENT_NUM=2",
                        "COMMON_ID=1",
                        "COMMON_NAME=Caller",
                        "COMMAND_NUM=1",
                        "WoditorEvCOMMAND_START",
                        "[211][2,0]<0>(2,0)()",
                        "WoditorEvCOMMAND_END",
                        "COMMON_ID=2",
                        "COMMON_NAME=Reserved",
                        "COMMAND_NUM=1",
                        "WoditorEvCOMMAND_START",
                        "[211][2,0]<0>(1600001,0)()",
                        "WoditorEvCOMMAND_END",
                    )
                ),
                encoding="utf-8",
            )
            (basic / "DataBase.Auto.txt").write_text(
                "\n".join(
                    (
                        "[DATABASE_TEXT_OUTPUT]",
                        "TYPE_NUM=1",
                        "TYPE_ID=0",
                        "ITEM_NUM=1",
                        "DATATYPE_0=2000",
                        "DATA_NUM=1",
                        "TYPENAME=Labels",
                        "ITEMNAME_NUM=1",
                        "ITEMNAME0=Name",
                        "<<--CSV_START-->>",
                        '"Name",',
                        '"Item",',
                        "<<--CSV_END-->>",
                    )
                ),
                encoding="utf-8",
            )
            item = TranslationItem(
                key="database_label",
                original="Item",
                translation="道具",
                code="UDB-0-0-0",
            )
            editor = EditorInfo(
                root / "Editor.exe",
                "3.713.2026.718",
                (3, 713, 2026, 718),
                "a" * 64,
            )
            report = analyze_auto_export(root / "Auto", [item], editor, input_hash="input")
            safety = analyze_translation_safety(
                root / "Auto",
                [item],
                {item.key: item.translation},
                "warn",
                analysis=report,
            )

            self.assertEqual(["database_label"], safety["safe_to_translate"])
            self.assertEqual([], safety["keep_original"])
            self.assertFalse(
                any(
                    dependency.get("reason")
                    == "预约公共事件存在延迟的全局或数据库副作用"
                    for dependency in report["dependencies"]
                )
            )
            self.assertTrue(
                any(
                    dependency.get("reason") == "预约公共事件目标为运行时动态值"
                    for dependency in report["dependencies"]
                )
            )

    def test_editor_execution_waits_for_cross_process_lock(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            editor_path = root / "Editor.exe"
            editor_path.write_bytes(b"editor")
            lock_path = root / "WOLFLator" / "locks" / f"editor-{'a' * 64}.lock"
            code = (
                "import time\n"
                "from safe_io import ResourceLock\n"
                f"with ResourceLock({str(lock_path)!r}, 'test'):\n"
                " print('ready', flush=True)\n"
                " time.sleep(0.5)\n"
            )
            child = subprocess.Popen(
                [sys.executable, "-c", code],
                cwd=Path(__file__).resolve().parents[1],
                stdout=subprocess.PIPE,
                text=True,
            )
            try:
                self.assertEqual("ready", child.stdout.readline().strip())
                logs: list[str] = []
                started = time.monotonic()
                with mock.patch.dict(os.environ, {"LOCALAPPDATA": str(root)}):
                    with _editor_execution_lock(
                        EditorInfo(
                            editor_path,
                            "3.713.2026.718",
                            (3, 713, 2026, 718),
                            "a" * 64,
                        ),
                        cancel_event=None,
                        diagnostic_log=logs.append,
                        warning=None,
                    ):
                        pass
                self.assertGreaterEqual(time.monotonic() - started, 0.3)
                self.assertTrue(any("editor.queue.wait" in item for item in logs))
                self.assertTrue(any("editor.queue.acquired" in item for item in logs))
                self.assertTrue(any("editor.queue.released" in item for item in logs))
            finally:
                child.wait(timeout=5)
                if child.stdout is not None:
                    child.stdout.close()

    def test_editor_roundtrip_masks_only_approved_text_slots(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            before = root / "before" / "BasicData"
            after = root / "after" / "BasicData"
            before.mkdir(parents=True)
            after.mkdir(parents=True)

            def auto(text: str, opcode: int = 101) -> str:
                return "\n".join((
                    "[COMMON_EVENT_TEXT_OUTPUT]",
                    "COMMON_EVENT_NUM=1",
                    "COMMON_ID=1",
                    "COMMON_NAME=Event",
                    "COMMAND_NUM=1",
                    "WoditorEvCOMMAND_START",
                    f'[{opcode}][0,1]<0>()("{text}")',
                    "WoditorEvCOMMAND_END",
                ))

            path_before = before / "CommonEvent.dat.Auto.txt"
            path_after = after / "CommonEvent.dat.Auto.txt"
            path_before.write_text(auto("原文"), encoding="utf-8")
            path_after.write_text(auto("译文"), encoding="utf-8")
            item = TranslationItem(key="display", original="原文", code="COMMON-1-0-0")
            passed = compare_auto_structure(root / "before", root / "after", [item], {item.key})
            self.assertEqual("passed", passed["status"])
            path_after.write_text(auto("译文", opcode=106), encoding="utf-8")
            failed = compare_auto_structure(root / "before", root / "after", [item], {item.key})
            self.assertEqual("failed", failed["status"])
            self.assertEqual("command_structure", failed["differences"][0]["kind"])

            def segmented_auto(name: str, first: str, copied: str) -> str:
                return "\n".join((
                    "[COMMON_EVENT_TEXT_OUTPUT]",
                    "COMMON_EVENT_NUM=1",
                    "COMMON_ID=1",
                    f"COMMON_NAME={name}",
                    "COMMAND_NUM=2",
                    "WoditorEvCOMMAND_START",
                    f'[101][0,1]<0>()("{first}")',
                    f'[101][0,1]<0>()("{copied}")',
                    "WoditorEvCOMMAND_END",
                ))

            path_before.write_text(
                segmented_auto("原名", r"前<\n>后", r"前<\n>后"),
                encoding="utf-8",
            )
            path_after.write_text(
                segmented_auto("译名", r"前<\n>译", r"前<\n>译"),
                encoding="utf-8",
            )
            base = TranslationItem(
                key="base",
                original="前\n",
                translation="不应采用",
                code="COMMON-1-0-0",
                flag="NEXT=SEGMENT_1-COMMON-1-0-0",
            )
            segment = TranslationItem(
                key="segment",
                original="后",
                translation="译",
                code="SEGMENT_1-COMMON-1-0-0",
            )
            copied = TranslationItem(
                key="copied",
                original="前\n后",
                code="COMMON-1-1-0",
                flag=(
                    "<Half-Width Characters Only>\n"
                    "NEXT=SEGMENT_1-COMMON-1-1-0\n"
                    "COPY-FROM-COMMON-1-0-0"
                ),
            )
            copied_segment = TranslationItem(
                key="copied-segment",
                original="后",
                translation="错误",
                code="SEGMENT_1-COMMON-1-1-0",
            )
            name_source = TranslationItem(
                key="name-source",
                original="原名",
                translation="译名",
                code="UDB-1-0-0",
            )
            name_copy = TranslationItem(
                key="name-copy",
                original="原名",
                code="COMMON-1-Name",
                flag="COPY-FROM-UDB-1-0-0",
            )
            segmented = compare_auto_structure(
                root / "before",
                root / "after",
                [base, segment, copied, copied_segment, name_source, name_copy],
                {segment.key, copied_segment.key, name_source.key},
            )
            self.assertEqual("passed", segmented["status"])
            path_after.write_text(
                segmented_auto("译名", "错误", r"前<\n>译"),
                encoding="utf-8",
            )
            segmented_failed = compare_auto_structure(
                root / "before",
                root / "after",
                [base, segment, copied, copied_segment, name_source, name_copy],
                {segment.key, copied_segment.key, name_source.key},
            )
            self.assertEqual("segmented_string", segmented_failed["differences"][0]["kind"])

    def test_editor_roundtrip_maps_named_map_files_through_sdb_coordinates(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)

            def common() -> str:
                return "\n".join((
                    "[COMMON_EVENT_TEXT_OUTPUT]",
                    "COMMON_EVENT_NUM=0",
                ))

            def map_auto(name: str, text: str) -> str:
                return "\n".join((
                    "[MAPDATA_TEXT_OUTPUT]",
                    "EVENT_NUM=1",
                    "EVENT_ID=7",
                    f"EVENT_NAME={name}",
                    "COMMAND_NUM=1",
                    "WoditorEvCOMMAND_START",
                    f'[101][0,1]<0>()("{text}")',
                    "WoditorEvCOMMAND_END",
                ))

            sdb = "\n".join((
                "[DATABASE_TEXT_OUTPUT]",
                "TYPE_NUM=1",
                "TYPE_ID=0",
                "ITEM_NUM=1",
                "DATATYPE_0=2000",
                "DATA_NUM=7",
                "TYPENAME=Maps",
                "ITEMNAME_NUM=1",
                "ITEMNAME0=Path",
                "<<--CSV_START-->>",
                '"Path",',
                '"",',
                '"",',
                '"",',
                '"",',
                '"",',
                '"MapData/Fancy.mps",',
                '"MapData/Fancy.mps",',
                "<<--CSV_END-->>",
            ))
            for side, name, text in (
                ("before", "原事件", "原文"),
                ("after", "译事件", "译文"),
            ):
                basic = root / side / "BasicData"
                maps = root / side / "MapData"
                basic.mkdir(parents=True)
                maps.mkdir(parents=True)
                (basic / "CommonEvent.dat.Auto.txt").write_text(common(), encoding="utf-8")
                (basic / "SysDataBase.Auto.txt").write_text(sdb, encoding="utf-8")
                (maps / "Fancy.mps.Auto.txt").write_text(
                    map_auto(name, text), encoding="utf-8"
                )
            text_item = TranslationItem(
                key="map-text",
                original="原文",
                translation="译文",
                code="MAP-6-Ev007-Page1-0-0",
            )
            name_item = TranslationItem(
                key="map-name",
                original="原事件",
                translation="译事件",
                code="MAP-5-Ev007-Name",
            )
            result = compare_auto_structure(
                root / "before",
                root / "after",
                [text_item, name_item],
                {text_item.key, name_item.key},
            )
            self.assertEqual("passed", result["status"])

    def test_editor_version_and_sandbox_contract(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            editor = root / "Editor.exe"
            editor.write_bytes(b"editor")
            with mock.patch(
                "wolf_editor._windows_version_resource",
                return_value=("3.220", (3, 220, 0, 0), "WOLF RPG Editor"),
            ):
                with self.assertRaisesRegex(ValueError, "版本过旧"):
                    inspect_wolf_editor(editor)
            for version in ("3.631", "3.713"):
                parts = tuple(int(value) for value in version.split(".")) + (0, 0)
                with mock.patch(
                    "wolf_editor._windows_version_resource",
                    return_value=(version, parts[:4], "WOLF RPG Editor"),
                ):
                    self.assertEqual(version, inspect_wolf_editor(editor).version)

            game = root / "game"
            basic = game / "Data" / "BasicData"
            basic.mkdir(parents=True)
            (basic / "Game.dat").write_bytes(b"data")
            (basic / "icon.png").write_bytes(b"image")
            (game / "Data" / "MapData").mkdir()
            (game / "Data" / "MapData" / "Map001.mps").write_bytes(b"map")
            (game / "Data" / "story.txt").write_text("story", encoding="utf-8")
            sandbox = root / "sandbox"
            sandbox.mkdir()
            maps_found = _copy_editor_sandbox(editor, game, sandbox)
            self.assertEqual([Path("MapData/Map001.mps")], maps_found)
            self.assertTrue((sandbox / "Data" / "BasicData" / "Game.dat").is_file())
            sandbox_map = (
                sandbox / "Data" / "MapData" / "WOLFLatorMap00000000.mps"
            )
            self.assertEqual(b"map", sandbox_map.read_bytes())
            self.assertFalse((sandbox / "Data" / maps_found[0]).exists())
            generated = (
                sandbox
                / "Auto"
                / "MapData"
                / "WOLFLatorMap00000000.mps.Auto.txt"
            )
            generated.parent.mkdir(parents=True)
            generated.write_bytes(b"auto")
            _restore_editor_map_paths(sandbox / "Auto", maps_found)
            self.assertEqual(
                b"auto",
                (sandbox / "Auto" / "MapData" / "Map001.mps.Auto.txt").read_bytes(),
            )
            self.assertFalse((sandbox / "Data" / "BasicData" / "icon.png").exists())
            self.assertFalse((sandbox / "Data" / "story.txt").exists())


    def test_managed_editor_install_validates_and_repairs_package(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            archive = root / "editor.zip"
            editor_bytes = b"verified editor"
            with zipfile.ZipFile(archive, "w") as package:
                package.writestr("Editor.exe", editor_bytes)
            archive_bytes = archive.read_bytes()
            archive_hash = hashlib.sha256(archive_bytes).hexdigest()
            release = EditorRelease(
                "3.800",
                (3, 800),
                "https://silversecond.com/WolfRPGEditor/Data/WolfRPGEditor_3.800mini.zip",
                True,
            )

            def download(actual_release, target, *, progress=None):
                self.assertEqual(release, actual_release)
                target.write_bytes(archive_bytes)
                if progress:
                    progress(len(archive_bytes), len(archive_bytes))
                return archive_hash, len(archive_bytes)

            with mock.patch(
                "wolf_editor.discover_latest_editor_release",
                return_value=release,
            ), mock.patch(
                "wolf_editor._windows_version_resource",
                return_value=("3.800.2026.800", (3, 800, 2026, 800), "WOLF RPG Editor"),
            ), mock.patch("wolf_editor._download_editor_archive", side_effect=download) as fetch:
                executable = install_supported_editor(root / "packages")
                self.assertEqual(editor_bytes, executable.read_bytes())
                executable.write_bytes(b"damaged")
                repaired = install_supported_editor(root / "packages")
                self.assertEqual(editor_bytes, repaired.read_bytes())
                self.assertEqual(2, fetch.call_count)

    def test_merge_and_scoped_workbook_preserve_table(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = make_workbook(root / "source.xlsx")
            items = read_translation_items(source)
            payload = to_paratranz(items, full_export_scope())
            output = []
            for row in payload:
                translation = "译文"
                if chr(0xE100) in row["original"]:
                    translation += chr(0xE100)
                output.append({**row, "translation": translation, "stage": 1})
            merge_ainiee_output(items, output, full_export_scope())
            full = write_full_workbook(source, root / "full.xlsx", items)
            game = root / "game"
            (game / "Data").mkdir(parents=True)
            scoped = write_scoped_workbook(full, root / "scoped.xlsx", ImportScope(), game, items)
            workbook = load_workbook(scoped)
            sheet = workbook.active
            values = [sheet.cell(row, 7).value for row in range(2, 10)]
            self.assertTrue(values[0])
            self.assertIsNone(values[1])
            self.assertIsNone(values[2])
            self.assertIsNone(values[3])
            self.assertEqual(r"译文\C[1]", values[4])
            self.assertIsNone(values[5])
            self.assertTrue(values[6])
            self.assertTrue(values[7])
            self.assertIn("WolfTranslation", sheet.tables)
            self.assertTrue(sheet["A1"].font.bold)

    def test_filename_scope_requires_real_target(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = make_workbook(root / "source.xlsx")
            items = read_translation_items(source)
            for item in items:
                if item.category is ImportCategory.FILENAME:
                    item.translation = "Picture/face.png"
                elif item.category is not ImportCategory.COPY:
                    item.translation = "译文" + (r"\C[1]" if item.control_signature else "")
            full = write_full_workbook(source, root / "full.xlsx", items)
            game = root / "game"
            (game / "Data" / "Picture").mkdir(parents=True)
            (game / "Data" / "Other").mkdir(parents=True)
            (game / "Data" / "Other" / "face.png").write_bytes(b"wrong path")
            scope = ImportScope(filename=True)
            with self.assertRaisesRegex(ValueError, "没有对应真实文件"):
                write_scoped_workbook(full, root / "bad.xlsx", scope, game, items)
            (game / "Data" / "Picture" / "face.png").write_bytes(b"png")
            write_scoped_workbook(full, root / "good.xlsx", scope, game, items)

            for item in items:
                if item.category is ImportCategory.FILENAME:
                    item.translation = item.original
            no_op_full = write_full_workbook(source, root / "no-op-full.xlsx", items)
            no_op = write_scoped_workbook(
                no_op_full,
                root / "no-op.xlsx",
                scope,
                root / "empty-game",
                items,
            )
            no_op_sheet = load_workbook(no_op).active
            filename_rows = [
                row
                for row in range(2, no_op_sheet.max_row + 1)
                if "<FILENAME>" in str(no_op_sheet.cell(row, 2).value or "")
            ]
            self.assertTrue(filename_rows)
            self.assertTrue(
                all(no_op_sheet.cell(row, 7).value is None for row in filename_rows)
            )


    def test_copy_source_requires_matching_original(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "source.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(HEADERS)
            sheet.append(["COMMON-1", "", "Event", "Message", "", "原文甲", ""])
            sheet.append(["COMMON-2", "COPY-FROM-COMMON-1", "Event", "Copy", "", "原文乙", ""])
            workbook.save(path)
            with self.assertRaisesRegex(ValueError, "找不到唯一来源"):
                to_paratranz(read_translation_items(path), full_export_scope())

    def test_item_file_roundtrip_and_rejects_malformed_item(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            items = read_translation_items(make_workbook(root / "source.xlsx"))
            versioned = dump_items(root / "items.json", items)
            self.assertEqual(len(items), len(load_items(versioned)))
            malformed = json.loads(versioned.read_text(encoding="utf-8"))
            malformed["items"][0]["stage"] = "1"
            versioned.write_text(json.dumps(malformed), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "stage 不是整数"):
                load_items(versioned)

    def test_incremental_ambiguity_is_not_guessed(self):
        with tempfile.TemporaryDirectory() as directory:
            path = make_workbook(Path(directory) / "source.xlsx")
            previous = read_translation_items(path)
            duplicates = [item for item in previous if item.original == "重複"]
            duplicates[0].translation = "译法一"
            duplicates[1].translation = "译法二"
            current = read_translation_items(path)
            moved = [item for item in current if item.original == "重複"][0]
            moved.key = "new-location"
            current = [moved]
            reconciled, conflicts = reconcile_incremental(previous, current)
            self.assertEqual("", reconciled[0].translation)
            self.assertEqual(["译法一", "译法二"], conflicts[0]["candidates"])


class ControlTests(unittest.TestCase):
    def test_round_trip_and_reorder_rejection(self):
        protected, tokens = protect_control_tokens(r"\C[1]名前\V[2]")
        self.assertEqual([r"\C[1]", r"\V[2]"], tokens)
        self.assertEqual(r"\C[1]姓名\V[2]", restore_control_tokens(protected.replace("名前", "姓名"), tokens))
        swapped = protected.replace(chr(0xE100), "X").replace(chr(0xE101), chr(0xE100)).replace("X", chr(0xE101))
        with self.assertRaisesRegex(ValueError, "占位序列"):
            restore_control_tokens(swapped, tokens)

    def test_ruby_transport_keeps_base_visible_and_restores_chinese_order(self):
        cases = (
            (
                r"\r[鷲見,すみ]です。",
                [r"\r[鷲見,すみ]"],
                "鷲見[[WOLFLATOR_RUBY_0]]です。",
                "我是鹫见[[WOLFLATOR_RUBY_0]]。",
                r"我是\r[鹫见,すみ]。",
            ),
            (
                r"\r[風,ふう]\r[紀,き]\r[院,いん]家の朝は早い。",
                [r"\r[風,ふう]", r"\r[紀,き]", r"\r[院,いん]"],
                "風紀院[[WOLFLATOR_RUBY_0]]家の朝は早い。",
                "风纪院[[WOLFLATOR_RUBY_0]]家的清晨总是很早。",
                r"\r[风,ふう]\r[纪,き]\r[院,いん]家的清晨总是很早。",
            ),
            (
                r"生徒会長の\r[雄,ゆう]\r[弁,べん]\r[寺,じ]　\r[言,こと]\r[継,つぐ]です。",
                [
                    r"\r[雄,ゆう]",
                    r"\r[弁,べん]",
                    r"\r[寺,じ]",
                    r"\r[言,こと]",
                    r"\r[継,つぐ]",
                ],
                "生徒会長の雄弁寺[[WOLFLATOR_RUBY_0]]　言継[[WOLFLATOR_RUBY_1]]です。",
                "我是学生会长雄辩寺[[WOLFLATOR_RUBY_0]]　言继[[WOLFLATOR_RUBY_1]]。",
                r"我是学生会长\r[雄,ゆう]\r[辩,べん]\r[寺,じ]　\r[言,こと]\r[继,つぐ]。",
            ),
            (
                r"\r[天,あま]\r[翔,かける]ツバサです。",
                [r"\r[天,あま]", r"\r[翔,かける]"],
                "天翔[[WOLFLATOR_RUBY_0]]ツバサです。",
                "我是天翔[[WOLFLATOR_RUBY_0]]翼。",
                r"我是\r[天,あま]\r[翔,かける]翼。",
            ),
            (
                r"夜空の\r[星,ほし]を見上げた。",
                [r"\r[星,ほし]"],
                "夜空の星[[WOLFLATOR_RUBY_0]]を見上げた。",
                "我仰望了夜空中的星星[[WOLFLATOR_RUBY_0]]。",
                r"我仰望了夜空中的\r[星星,ほし]。",
            ),
            (
                r"\r[東,ひがし]の門から\r[西,にし]の塔へ向かう。",
                [r"\r[東,ひがし]", r"\r[西,にし]"],
                "東[[WOLFLATOR_RUBY_0]]の門から西[[WOLFLATOR_RUBY_1]]の塔へ向かう。",
                "从东[[WOLFLATOR_RUBY_0]]门向西[[WOLFLATOR_RUBY_1]]塔前进。",
                r"从\r[东,ひがし]门向\r[西,にし]塔前进。",
            ),
        )
        scope = ImportScope(external=True)
        for index, (original, signature, transport, translated, expected) in enumerate(cases):
            with self.subTest(original=original):
                item = TranslationItem(
                    key=f"ruby-{index}",
                    original=original,
                    code=f'SEGMENT_{index}-TXTFILE-"Data\\ノベル\\シーン1.txt"',
                    type="Text File",
                    category=ImportCategory.EXTERNAL,
                    control_signature=signature,
                )
                row = to_paratranz([item], scope)[0]
                self.assertEqual(transport, row["original"])
                self.assertNotIn("すみ", transport)
                self.assertNotIn("ふう", transport)
                self.assertNotIn("あま", transport)
                merged = merge_ainiee_output(
                    [item], [{**row, "translation": translated, "stage": 1}], scope
                )
                self.assertEqual(expected, merged[0].translation)

    def test_ruby_transport_rejects_missing_or_invalid_anchor(self):
        original = r"\r[天,あま]\r[翔,かける]ツバサです。"
        item = TranslationItem(
            key="ruby",
            original=original,
            code='SEGMENT_37-TXTFILE-"Data\\ノベル\\シーン1.txt"',
            type="Text File",
            category=ImportCategory.EXTERNAL,
            control_signature=[r"\r[天,あま]", r"\r[翔,かける]"],
        )
        scope = ImportScope(external=True)
        row = to_paratranz([item], scope)[0]
        translated = "我是天翔[[WOLFLATOR_RUBY_0]]翼。"

        with self.assertRaisesRegex(ValueError, "ruby 锚点序列"):
            merge_ainiee_output(
                [item], [{**row, "translation": translated.replace("[[WOLFLATOR_RUBY_0]]", "")}], scope
            )
        with self.assertRaisesRegex(ValueError, "ruby 标注正文包含非法字符"):
            merge_ainiee_output(
                [item], [{**row, "translation": "我 [[WOLFLATOR_RUBY_0]]翼。"}], scope
            )

    def test_empty_ruby_definition_remains_a_generic_control(self):
        item = TranslationItem(
            key="ruby-color",
            original=r"[Sys]\r[]ルビ用文字色",
            code="NAME-D-SDB-12-13",
            type="SDB info",
            category=ImportCategory.DISPLAY,
            control_signature=[r"\r[]"],
        )
        row = to_paratranz([item], ImportScope(display=True))[0]
        self.assertIn(chr(0xE100), row["original"])
        self.assertNotIn("WOLFLATOR_RUBY", row["original"])

    def test_external_script_translates_display_payload_and_preserves_structure(self):
        original = (
            "@背景：1\n"
            "鷲見A普通\n"
            "@文章：0\n"
            "空を飛んで、落ちて、目が覚める。\n"
            "@文章：0\n"
            r"\c[6]次の台詞。"
        )
        item = TranslationItem(
            key="script",
            original=original,
            code='TXTFILE-"Data\\ノベル\\シーン1.txt"',
            type="Text File",
            category=ImportCategory.EXTERNAL,
            control_signature=[r"\c[6]"],
        )
        scope = ImportScope(external=True)
        payload = to_paratranz([item], scope)
        protected = str(payload[0]["original"])
        self.assertIn("空を飛んで、落ちて、目が覚める。", protected)
        self.assertIn("次の台詞。", protected)
        self.assertNotIn("@文章", protected)
        self.assertNotIn("鷲見A普通", protected)
        self.assertNotIn(r"\c[6]", protected)
        self.assertEqual(0, protected.count("\n"))

        translated = protected.replace(
            "空を飛んで、落ちて、目が覚める。", "飞过天空，坠落，然后醒来。"
        ).replace("次の台詞。", "下一句台词。")
        merged = merge_ainiee_output(
            [item], [{**payload[0], "translation": translated}], scope
        )
        self.assertEqual(
            original.replace(
                "空を飛んで、落ちて、目が覚める。", "飞过天空，坠落，然后醒来。"
            ).replace("次の台詞。", "下一句台词。"),
            merged[0].translation,
        )

        broken = translated.replace(chr(0xE100), "", 1)
        with self.assertRaisesRegex(ValueError, "占位序列"):
            merge_ainiee_output([item], [{**payload[0], "translation": broken}], scope)

    def test_external_script_continuation_preserves_line_breaks_without_command_header(self):
        original = "一行。\r\n\r\n二行。\n三行。"
        item = TranslationItem(
            key="script-continuation",
            original=original,
            code='SEGMENT_51-TXTFILE-"Data\\ノベル\\シーン1.txt"',
            type="Text File",
            category=ImportCategory.EXTERNAL,
        )
        scope = ImportScope(external=True)
        payload = to_paratranz([item], scope)
        protected = str(payload[0]["original"])

        self.assertNotEqual(original, protected)
        self.assertEqual(0, protected.count("\n"))
        self.assertEqual(3, sum("\ue100" <= char <= "\uf7ff" for char in protected))
        translated = (
            protected.replace("一行。", "第一行。")
            .replace("二行。", "第二行。")
            .replace("三行。", "第三行。")
        )
        merged = merge_ainiee_output(
            [item], [{**payload[0], "translation": translated}], scope
        )
        self.assertEqual("第一行。\r\n\r\n第二行。\n第三行。", merged[0].translation)

        broken = translated.replace(chr(0xE100), "", 1)
        with self.assertRaisesRegex(ValueError, "占位序列"):
            merge_ainiee_output([item], [{**payload[0], "translation": broken}], scope)

        changed_line_ending = translated.replace(chr(0xE100), chr(0xE101), 1)
        with self.assertRaisesRegex(ValueError, "占位序列"):
            merge_ainiee_output(
                [item], [{**payload[0], "translation": changed_line_ending}], scope
            )

        moved_blank_line = translated.replace("第二行。", "第二行。\n", 1)
        with self.assertRaisesRegex(ValueError, "译文脚本结构"):
            merge_ainiee_output(
                [item], [{**payload[0], "translation": moved_blank_line}], scope
            )

        item.translation = moved_blank_line
        cached = to_paratranz([item], scope)[0]
        self.assertEqual("", cached["translation"])
        self.assertEqual(0, cached["stage"])

    def test_external_script_does_not_normalize_resource_middle_dots(self):
        original = "@立ち絵：1\n画像・通常\n@文章：0\n台詞。"
        item = TranslationItem(
            key="script",
            original=original,
            code="TXTFILE-1",
            type="Text File",
            category=ImportCategory.EXTERNAL,
        )
        scope = ImportScope(external=True)
        payload = to_paratranz([item], scope)
        merged = merge_ainiee_output(
            [item], [{**payload[0], "translation": payload[0]["original"]}], scope
        )
        self.assertEqual(original, merged[0].translation)


class ProcessTests(unittest.TestCase):
    def test_official_map_failure_parser_requires_corruption_evidence(self):
        output = (
            "Map 86 : Data\\MapData/Map001.mps が読み込めませんでした\n"
            "該当ファイルが 破損しているか アクセス権限 => Failed..."
        )
        failures = parse_official_map_failures(output)
        self.assertEqual(1, len(failures))
        self.assertIn("Map 86", failures[0])
        self.assertEqual(
            [],
            parse_official_map_failures(
                "Map 1 : Data/MapData/A.mps が読み込めませんでした => Failed..."
            ),
        )

    @unittest.skipUnless(os.name == "nt", "Windows dialog contract")
    def test_hidden_official_dialog_is_captured_and_dismissed(self):
        import ctypes

        thread = threading.Thread(
            target=lambda: ctypes.windll.user32.MessageBoxW(
                0, "Map Size Error! [Size 0 Error]", "Error", 0
            )
        )
        thread.start()
        dialogs = []
        for _ in range(50):
            dialogs = _dismiss_process_dialogs(os.getpid())
            if dialogs:
                break
            time.sleep(0.02)
        thread.join(timeout=2)
        self.assertFalse(thread.is_alive())
        self.assertIn("Map Size Error! [Size 0 Error]", dialogs[0])

    def test_slow_process_warning_fires_once_without_stopping_process(self):
        warnings = []
        result = run_process(
            [sys.executable, "-c", "import time; time.sleep(0.35)"],
            timeout=5,
            slow_warning_after=0.05,
            slow_warning=warnings.append,
        )
        self.assertEqual(0, result.return_code)
        self.assertEqual(1, len(warnings))
        self.assertGreaterEqual(warnings[0], 0.05)


    def test_silent_official_executable_does_not_modify_source(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "official.exe"
            original = b"prefix-MessageBeep\0-suffix"
            source.write_bytes(original)
            with mock.patch("wolf_tools._pe_import_name_offset", return_value=7):
                silent = _silent_official_executable(source)
            self.assertEqual(original, source.read_bytes())
            self.assertEqual(b"prefix-IsWindow\0\0\0\0-suffix", silent)






    def test_nonzero_and_cancel(self):
        with self.assertRaisesRegex(RuntimeError, "退出码 3"):
            run_process([sys.executable, "-c", "raise SystemExit(3)"], timeout=10)
        event = threading.Event()
        event.set()
        with self.assertRaises(CancelledError):
            run_process([sys.executable, "-c", "import time; time.sleep(5)"], cancel_event=event)



    def test_detached_console_cannot_invalidate_completed_process(self):
        def detached_console(_message: str) -> None:
            raise OSError(22, "invalid console handle")

        result = run_process(
            [sys.executable, "-c", "print('completed')"],
            timeout=10,
            log=detached_console,
        )
        self.assertEqual(0, result.return_code)


class AiNieeTests(unittest.TestCase):
    def make_runtime(self, root: Path) -> Path:
        (root / "Resource" / "profiles").mkdir(parents=True)
        (root / "Resource" / "rules_profiles").mkdir(parents=True)
        (root / "Resource" / "Version").mkdir(parents=True)
        (root / "Resource" / "Version" / "version.json").write_text(
            '{"version":"test"}', encoding="utf-8"
        )
        assets = root / "Tools" / "WebServer" / "dist" / "assets"
        assets.mkdir(parents=True)
        (assets.parent / "index.html").write_text("<html></html>", encoding="utf-8")
        (assets / "index.js").write_text("", encoding="utf-8")
        (root / "ainiee_cli.py").write_text("# test runtime\n", encoding="utf-8")
        (root / "pyproject.toml").write_text("[project]\nname='x'\n", encoding="utf-8")
        (root / "uv.lock").write_text("lock", encoding="utf-8")
        patcher = mock.patch.object(ainiee, "AINIEE_SOURCE_SHA256", ainiee._source_code_hash(root))
        patcher.start()
        self.addCleanup(patcher.stop)
        return root

    def test_source_validation_rejects_changed_code(self):
        with tempfile.TemporaryDirectory() as directory:
            root = self.make_runtime(Path(directory) / "runtime")
            (root / "ainiee_cli.py").write_text("# changed\n", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "源码版本不兼容"):
                ainiee.validate_ainiee_source(root)

    def test_safe_extract_rejects_traversal(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            archive = root / "bad.zip"
            with zipfile.ZipFile(archive, "w") as package:
                package.writestr("../outside.txt", "bad")
            with self.assertRaisesRegex(ValueError, "越界路径"):
                ainiee._safe_extract(archive, root / "out")




    def test_managed_package_requires_install_metadata(self):
        with tempfile.TemporaryDirectory() as directory:
            source = self.make_runtime(Path(directory) / "source")
            with self.assertRaisesRegex(ValueError, "缺少安装元数据"):
                ainiee._validate_managed_package(source)





    def test_only_verified_ainiee_exclusions_are_restored(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory)
            cache = output / "cache"
            cache.mkdir()
            input_rows = [{"key": "k", "original": "x", "translation": "", "stage": 0}]
            cache_data = {
                "files": {
                    "input.json": {
                        "items": [
                            {
                                "source_text": "x",
                                "translation_status": 7,
                                "extra": {"key": "k"},
                            }
                        ]
                    }
                }
            }
            (cache / "AinieeCacheData.json").write_text(
                json.dumps(cache_data, ensure_ascii=False), encoding="utf-8"
            )
            diagnostics = []
            restored = ainiee._restore_excluded_rows(input_rows, [], output, diagnostics.append)
            self.assertEqual("x", restored[0]["translation"])
            self.assertTrue(restored[0]["wolflator_excluded"])
            self.assertIn("restored=1 unresolved=0", diagnostics[0])

            cache_data["files"]["input.json"]["items"][0]["source_text"] = "changed"
            (cache / "AinieeCacheData.json").write_text(
                json.dumps(cache_data, ensure_ascii=False), encoding="utf-8"
            )
            with self.assertRaisesRegex(ValueError, "原文不一致"):
                ainiee._restore_excluded_rows(input_rows, [], output, None)

            with self.assertRaisesRegex(ValueError, "重复键"):
                ainiee._restore_excluded_rows(input_rows * 2, [], output, None)






    def test_deepseek_request_matches_managed_ainiee_profile(self):
        response = mock.MagicMock()
        response.read.return_value = json.dumps(
            {
                "choices": [{"message": {"content": "ok"}, "finish_reason": "stop"}],
                "usage": {"prompt_tokens": 10, "completion_tokens": 2},
            }
        ).encode("utf-8")
        response.status = 200
        response.__enter__.return_value = response
        diagnostics = []
        client = ainiee.OpenAICompatibleClient(
            "https://api.deepseek.com/v1/chat/completions",
            "secret",
            "deepseek-chat",
            diagnostic_log=diagnostics.append,
        )
        with mock.patch("urllib.request.urlopen", return_value=response) as urlopen:
            self.assertEqual(
                "ok",
                client.chat("hello", max_tokens=None, system_prompt="system"),
            )
        request = urlopen.call_args.args[0]
        body = json.loads(request.data.decode("utf-8"))
        self.assertEqual("https://api.deepseek.com/v1/chat/completions", request.full_url)
        self.assertNotIn("max_tokens", body)
        self.assertEqual({"type": "disabled"}, body["thinking"])
        self.assertEqual(["system", "user"], [message["role"] for message in body["messages"]])
        joined = "\n".join(diagnostics)
        self.assertIn("api.request id=1", joined)
        self.assertIn("api.response id=1 status=200", joined)
        self.assertIn("finish_reason=stop", joined)
        self.assertIn('"prompt_tokens": 10', joined)
        self.assertNotIn("secret", joined)


    def test_api_request_always_disables_thinking(self):
        response = mock.MagicMock()
        response.read.return_value = json.dumps(
            {
                "choices": [{"message": {"content": "ok"}, "finish_reason": "stop"}],
                "usage": {"prompt_tokens": 10, "completion_tokens": 2},
            }
        ).encode("utf-8")
        response.status = 200
        response.__enter__.return_value = response
        client = ainiee.OpenAICompatibleClient(
            "https://gateway.example/v1/chat/completions",
            "secret",
            "qwen3",
        )
        with mock.patch("urllib.request.urlopen", return_value=response) as urlopen:
            self.assertEqual("ok", client.chat("hello"))
        body = json.loads(urlopen.call_args.args[0].data.decode("utf-8"))
        self.assertEqual({"type": "disabled"}, body["thinking"])


    def test_api_timeout_covers_the_complete_response(self):
        body = json.dumps(
            {"choices": [{"message": {"content": "ok"}, "finish_reason": "stop"}]}
        ).encode("utf-8")

        class SlowHandler(http.server.BaseHTTPRequestHandler):
            def do_POST(self):
                self.rfile.read(int(self.headers.get("Content-Length", "0")))
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                try:
                    for part in (body[:1], body[1:2], body[2:]):
                        self.wfile.write(part)
                        self.wfile.flush()
                        time.sleep(0.15)
                except (BrokenPipeError, ConnectionResetError):
                    pass

            def log_message(self, _format, *_args):
                pass

        server = http.server.ThreadingHTTPServer(("127.0.0.1", 0), SlowHandler)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        diagnostics = []
        client = ainiee.OpenAICompatibleClient(
            f"http://127.0.0.1:{server.server_port}/v1",
            "secret",
            "model",
            diagnostic_log=diagnostics.append,
        )
        client.timeout = 0.2
        started = time.monotonic()
        try:
            with self.assertRaisesRegex(ainiee.ApiError, "总时限"):
                client.chat("hello")
        finally:
            server.shutdown()
            server.server_close()
            thread.join(timeout=1)
        self.assertLess(time.monotonic() - started, 1.0)
        self.assertIn("kind=timeout", "\n".join(diagnostics))



    def test_glossary_first_failure_cancels_queued_chunks(self):
        client = mock.Mock()
        client.chat.return_value = '[{"src":"broken"}'
        diagnostics = []
        with mock.patch.object(ainiee.time, "sleep"):
            with self.assertRaisesRegex(RuntimeError, "Expecting"):
                ainiee._parallel_stage(
                    client,
                    "prompt",
                    ["bad", "must not start", "also must not start"],
                    workers=1,
                    cancel_event=None,
                    log=None,
                    diagnostic_log=diagnostics.append,
                    label="角色分析",
                    max_tokens=None,
                )
        self.assertEqual(3, client.chat.call_count)
        self.assertFalse(any("角色分析:2/3" in line for line in diagnostics))

        aborted = threading.Event()
        aborted.set()
        client.reset_mock()
        with self.assertRaises(CancelledError):
            ainiee._request_chunk(
                client,
                "prompt",
                "chunk",
                cancel_event=None,
                abort_event=aborted,
            )
        client.chat.assert_not_called()

    def test_zero_exit_without_artifact_is_failure_and_profile_is_cleaned(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            runtime = self.make_runtime(root / "runtime")
            input_path = root / "input.json"
            input_path.write_text("[]", encoding="utf-8")
            settings = AppSettings(api_base_url="https://example.com/v1", api_model="model")
            fake_result = ToolResult([], 0)
            output = root / "output"
            output.mkdir()
            (output / "input_translated.json").write_text("[]", encoding="utf-8")
            with mock.patch.object(ainiee, "run_process", return_value=fake_result) as run, mock.patch.object(
                ainiee, "locate_uv", return_value=Path(sys.executable)
            ):
                with self.assertRaisesRegex(RuntimeError, "没有生成"):
                    ainiee.run_translation(
                        runtime, input_path, output, dict(ainiee.RULE_DEFAULTS), "project", settings, "secret"
                    )
            self.assertEqual(
                ["run", "--frozen", "--no-sync", "ainiee_cli.py"],
                run.call_args.args[0][1:5],
            )
            self.assertNotIn("-p", run.call_args.args[0])
            self.assertNotIn("--rules-profile", run.call_args.args[0])
            self.assertFalse((runtime / "Resource" / "profiles" / "WOLFLator_session.json").exists())

    def test_translation_activates_profiles_restores_config_and_reads_v275_output(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            runtime = self.make_runtime(root / "runtime")
            config_path = runtime / "Resource" / "config.json"
            original_config = b'{"active_profile":"default","active_rules_profile":"default","keep":1}\n'
            config_path.write_bytes(original_config)
            input_path = root / "input.json"
            input_path.write_text('[{"key":"k","original":"x","translation":"","stage":0}]', encoding="utf-8")
            output = root / "output"
            settings = AppSettings(
                api_base_url="https://api.deepseek.com/v1",
                api_model="deepseek-v4-flash",
            )
            progress = []

            def fake_process(command, *, cwd, **_kwargs):
                self.assertNotIn("-p", command)
                self.assertNotIn("--rules-profile", command)
                self.assertEqual("6", command[command.index("--rounds") + 1])
                self.assertEqual("256", command[command.index("--tokens") + 1])
                self.assertNotIn("--lines", command)
                self.assertIn("--web-mode", command)
                self.assertEqual("1", _kwargs["env"]["PYTHONUTF8"])
                self.assertEqual("utf-8", _kwargs["env"]["PYTHONIOENCODING"])
                _kwargs["output_line"](
                    "stdout",
                    "[STATS] RPM: 1.00 | Progress: 1/4 | Tokens: 10",
                )
                active = json.loads(config_path.read_text(encoding="utf-8"))
                self.assertEqual("WOLFLator_session", active["active_profile"])
                self.assertEqual("WOLFLator_project", active["active_rules_profile"])
                profile = json.loads(
                    (runtime / "Resource" / "profiles" / "WOLFLator_session.json").read_text(encoding="utf-8")
                )
                rules = json.loads(
                    (runtime / "Resource" / "rules_profiles" / "WOLFLator_project.json").read_text(encoding="utf-8")
                )
                self.assertEqual("secret", profile["platforms"]["deepseek"]["api_key"])
                self.assertTrue(profile["tokens_limit_switch"])
                self.assertEqual(256, profile["tokens_limit"])
                self.assertEqual(6, profile["round_limit"])
                self.assertFalse(profile["enable_smart_round_limit"])
                self.assertFalse(
                    profile["response_check_switch"]["newline_character_count_check"]
                )
                self.assertEqual([], rules["prompt_dictionary_data"])
                self.assertTrue(rules["prompt_dictionary_switch"])
                self.assertTrue(rules["exclusion_list_switch"])
                output.mkdir(parents=True, exist_ok=True)
                (output / input_path.name).write_text(
                    '[{"key":"k","original":"x","translation":"译文"}]', encoding="utf-8"
                )
                return ToolResult(command, 0)

            with mock.patch.object(ainiee, "run_process", side_effect=fake_process), mock.patch.object(
                ainiee, "locate_uv", return_value=Path(sys.executable)
            ):
                translated = ainiee.run_translation(
                    runtime,
                    input_path,
                    output,
                    {"prompt_dictionary_data": []},
                    "project",
                    settings,
                    "secret",
                    progress=progress.append,
                )
            self.assertEqual("译文", translated[0]["translation"])
            self.assertEqual(1, progress[-1]["current"])
            self.assertEqual(4, progress[-1]["total"])
            self.assertEqual(original_config, config_path.read_bytes())
            self.assertFalse((runtime / "Resource" / "profiles" / "WOLFLator_session.json").exists())

    def test_translation_failure_includes_ainiee_session_log_tail(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            runtime = self.make_runtime(root / "runtime")
            input_path = root / "input.json"
            input_path.write_text("[]", encoding="utf-8")
            output = root / "output"
            diagnostics = []

            def fail_process(*_args, **_kwargs):
                logs = output / "logs"
                logs.mkdir(parents=True)
                (logs / "session.log").write_text("API 429 rate limited\nretry exhausted", encoding="utf-8")
                raise RuntimeError("translation failed")

            settings = AppSettings(api_base_url="https://example.com/v1", api_model="model")
            with mock.patch.object(ainiee, "run_process", side_effect=fail_process), mock.patch.object(
                ainiee, "locate_uv", return_value=Path(sys.executable)
            ):
                with self.assertRaisesRegex(RuntimeError, "translation failed"):
                    ainiee.run_translation(
                        runtime,
                        input_path,
                        output,
                        dict(ainiee.RULE_DEFAULTS),
                        "project",
                        settings,
                        "secret",
                        diagnostic_log=diagnostics.append,
                    )
            joined = "\n".join(diagnostics)
            self.assertIn("ainiee.session_log.tail", joined)
            self.assertIn("API 429 rate limited", joined)


if __name__ == "__main__":
    unittest.main()
