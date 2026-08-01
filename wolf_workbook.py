from __future__ import annotations

import hashlib
import json
import os
import re
import xml.etree.ElementTree as ET
import zipfile
from collections import Counter
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from fonts import FONT_CODES
from formats import ARTIFACT_EPOCH, require_format
from models import ImportCategory, ImportScope, TranslationItem
from safe_io import atomic_output_path, atomic_write_json, read_text_with_retry, replace_with_retry


CODE_HEADER = "Code (No Change)"

FLAG_HEADER = "Flag (No Change)"

TYPE_HEADER = "Type"

INFO_HEADER = "Info"

ORIGINAL_HEADER = "Original text (No Change)"

TARGET_PREFIX = "Translated text 1 /"

EXPECTED_TARGET = "Chinese (Simplified)"

SUPPORT_DIR = "WOLF_Translation_Support_Tool_Data"

WORKBOOK_NAME = "WOLF_Translation_Text.xlsx"

PUA_START = 0xE100

PUA_END = 0xF7FF

SPECIAL_ESCAPES = set("!.|^<>${}\\")

COPY_FROM_RE = re.compile(r"(?:^|\r?\n)COPY-FROM-([^\r\n]+)", re.IGNORECASE)

def full_export_scope() -> ImportScope:
    # ponytail: this is the complete internal WOLF structure, not the user-facing export default.
    return ImportScope(display=True, external=False, optional_name=True, halfwidth=True, filename=True)

def name_baseline_scope(scope: ImportScope | None = None) -> ImportScope:
    source = scope or full_export_scope()
    return ImportScope(
        display=source.display,
        external=source.external,
        optional_name=False,
        halfwidth=source.halfwidth,
        filename=source.filename,
    )

def _header_map(worksheet) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        if CODE_HEADER in row and ORIGINAL_HEADER in row:
            mapping = {str(value): index + 1 for index, value in enumerate(row) if value is not None}
            missing = {CODE_HEADER, FLAG_HEADER, TYPE_HEADER, INFO_HEADER, ORIGINAL_HEADER} - mapping.keys()
            if missing:
                raise ValueError(f"官方工作簿缺少列: {', '.join(sorted(missing))}")
            targets = [name for name in mapping if name.startswith(TARGET_PREFIX)]
            if not targets or EXPECTED_TARGET not in targets[0]:
                raise ValueError("官方工作簿第一译文列不是简体中文。")
            mapping["__target__"] = mapping[targets[0]]
            return row_index, mapping
    raise ValueError("不是受支持的 WOLF Translation Support Tool 工作簿。")

def locate_workbook(game_root: str | Path) -> Path:
    support = Path(game_root) / SUPPORT_DIR
    workbook_path = support / WORKBOOK_NAME
    if not workbook_path.is_file():
        raise FileNotFoundError(f"官方工具没有生成 {workbook_path}。")
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    try:
        _header_map(workbook.active)
    finally:
        workbook.close()
    return workbook_path

def _content_category(code: str, flag: str, type_name: str) -> ImportCategory:
    upper_flag = flag.upper()
    upper_code = code.upper()
    upper_type = type_name.upper()
    if "<FILENAME>" in upper_flag:
        return ImportCategory.FILENAME
    if "<HALF-WIDTH CHARACTERS ONLY>" in upper_flag:
        return ImportCategory.HALFWIDTH
    if upper_code.startswith("NAME-") or upper_code.endswith("-NAME"):
        return ImportCategory.OPTIONAL_NAME
    if upper_code.startswith(("TXT-", "CSV-", "TXTFILE-", "CSVFILE-")) or any(
        marker in upper_type for marker in ("TXT", "CSV", "TEXT FILE")
    ):
        return ImportCategory.EXTERNAL
    return ImportCategory.DISPLAY

def _category(code: str, flag: str, type_name: str) -> ImportCategory:
    if "COPY-FROM-" in flag.upper():
        return ImportCategory.COPY
    return _content_category(code, flag, type_name)

def stable_key(code: str, flag: str, original: str, ordinal: int) -> str:
    payload = "\0".join((code, flag, original, str(ordinal))).encode("utf-8", "surrogatepass")
    return hashlib.sha256(payload).hexdigest()

def _iter_data_rows(worksheet) -> Iterable[tuple[int, dict[str, str], int]]:
    header_row, headers = _header_map(worksheet)
    counts: Counter[tuple[str, str, str]] = Counter()
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        def value(column: str):
            index = headers[column] - 1
            return row[index] if index < len(row) else None

        original = value(ORIGINAL_HEADER)
        if original is None:
            continue
        values = {
            "code": str(value(CODE_HEADER) or ""),
            "flag": str(value(FLAG_HEADER) or ""),
            "type": str(value(TYPE_HEADER) or ""),
            "info": str(value(INFO_HEADER) or ""),
            "original": str(original),
            "translation": str(value("__target__") or ""),
        }
        identity = (values["code"], values["flag"], values["original"])
        counts[identity] += 1
        yield row_index, values, counts[identity]

def _scan_control_spans(text: str) -> list[tuple[int, int]]:
    spans: list[tuple[int, int]] = []
    index = 0
    while index < len(text):
        if text[index] != "\\":
            index += 1
            continue
        start = index
        index += 1
        if index >= len(text):
            spans.append((start, index))
            break
        char = text[index]
        if char in SPECIAL_ESCAPES:
            index += 1
        elif char.isascii() and (char.isalnum() or char == "_"):
            index += 1
            while index < len(text) and text[index].isascii() and (text[index].isalnum() or text[index] == "_"):
                index += 1
            while index < len(text) and text[index] == "[":
                depth = 0
                while index < len(text):
                    if text[index] == "[":
                        depth += 1
                    elif text[index] == "]":
                        depth -= 1
                        if depth == 0:
                            index += 1
                            break
                    elif text[index] in "\r\n":
                        break
                    index += 1
        else:
            # ponytail: Unknown backslash forms protect only the slash; upgrade the scanner if WOLF documents more syntax.
            index = start + 1
        spans.append((start, index))
    return spans

def _scan_control_tokens(text: str) -> list[str]:
    return [text[start:end] for start, end in _scan_control_spans(text)]

_EXTERNAL_DISPLAY_COMMAND_RE = re.compile(r"^[ \t]*@(?:文章|連続文章|タイトルコール)(?:[：:].*)?$")

def _external_script_control_spans(text: str) -> list[tuple[int, int]]:
    lines = list(re.finditer(r".*?(?:\r\n|\n|\r|$)", text))
    if not any(
        _EXTERNAL_DISPLAY_COMMAND_RE.match(match.group(0).rstrip("\r\n"))
        for match in lines
    ):
        # ponytail: AiNiee validates one logical row at a time; keep physical newlines as transport tokens.
        return [match.span() for match in re.finditer(r"\r\n|\n|\r", text)]

    first = next(
        (match.group(0).rstrip("\r\n").strip() for match in lines if match.group(0).strip()),
        "",
    )
    display_payload = not first.startswith(("@", "●", "-", "//"))
    spans: list[tuple[int, int]] = []
    for match in lines:
        line = match.group(0)
        if not line:
            continue
        content = line.rstrip("\r\n")
        stripped = content.strip()
        line_end = match.start() + len(content)
        if stripped.startswith("@"):
            if line_end > match.start():
                spans.append((match.start(), match.end()))
            display_payload = bool(_EXTERNAL_DISPLAY_COMMAND_RE.match(content))
        elif stripped.startswith("●") or (stripped and set(stripped) == {"-"}):
            if line_end > match.start():
                spans.append((match.start(), match.end()))
            display_payload = False
        elif stripped.startswith("//") or not display_payload:
            if line_end > match.start():
                spans.append((match.start(), match.end()))
        elif line_end < match.end():
            spans.append((line_end, match.end()))

    # ponytail: This recognizes the observed line-oriented WOLF script dialect.
    # Register a dialect parser if an external format needs different payload rules.
    merged: list[tuple[int, int]] = []
    for start, end in spans:
        if merged and start <= merged[-1][1]:
            merged[-1] = (merged[-1][0], max(merged[-1][1], end))
        else:
            merged.append((start, end))
    return merged

def _item_control_spans(item: TranslationItem, text: str) -> list[tuple[int, int]]:
    structural = (
        _external_script_control_spans(text)
        if _content_category(item.code, item.flag, item.type) is ImportCategory.EXTERNAL
        else []
    )
    controls = [
        span
        for span in _scan_control_spans(text)
        if not any(span[0] < end and start < span[1] for start, end in structural)
    ]
    return sorted((*structural, *controls))

def _protect_spans(text: str, spans: list[tuple[int, int]]) -> tuple[str, list[str]]:
    if not spans:
        return text, []
    output: list[str] = []
    tokens: list[str] = []
    cursor = 0
    for offset, (start, end) in enumerate(spans):
        codepoint = PUA_START + offset
        if codepoint > PUA_END:
            raise ValueError("单条文本的控制符数量超过占位符容量。")
        output.append(text[cursor:start])
        output.append(chr(codepoint))
        tokens.append(text[start:end])
        cursor = end
    output.append(text[cursor:])
    return "".join(output), tokens

def protect_control_tokens(text: str) -> tuple[str, list[str]]:
    return _protect_spans(text, _scan_control_spans(text))

def _protect_item_tokens(item: TranslationItem, text: str) -> tuple[str, list[str]]:
    return _protect_spans(text, _item_control_spans(item, text))

def _restore_tokens(text: str, tokens: list[str]) -> str:
    expected = [chr(PUA_START + index) for index in range(len(tokens))]
    actual = [char for char in text if PUA_START <= ord(char) <= PUA_END]
    if actual != expected:
        raise ValueError(
            "控制符占位序列不一致: "
            f"expected={[f'U+{ord(c):04X}' for c in expected]}, "
            f"actual={[f'U+{ord(c):04X}' for c in actual]}"
        )
    restored = text
    for placeholder, token in zip(expected, tokens):
        restored = restored.replace(placeholder, token, 1)
    return restored

def restore_control_tokens(text: str, tokens: list[str]) -> str:
    restored = _restore_tokens(text, tokens)
    if _scan_control_tokens(restored) != tokens:
        raise ValueError("译文控制符序列与原文不一致。")
    return restored

def _restore_item_tokens(item: TranslationItem, text: str) -> str:
    tokens = [item.original[start:end] for start, end in _item_control_spans(item, item.original)]
    restored = _restore_tokens(text, tokens)
    if [restored[start:end] for start, end in _item_control_spans(item, restored)] != tokens:
        raise ValueError("译文脚本结构与原文不一致。")
    if _scan_control_tokens(restored) != item.control_signature:
        raise ValueError("译文控制符序列与原文不一致。")
    return restored


def _line_structure(text: str) -> tuple[list[str], list[tuple[int, str]]]:
    endings = [match.group(0) for match in re.finditer(r"\r\n|\n|\r", text)]
    lines = re.split(r"\r\n|\n|\r", text)
    blank_lines = [
        (index, line)
        for index, line in enumerate(lines)
        if not line.strip()
    ]
    return endings, blank_lines


def _validate_line_structure(original: str, translated: str) -> None:
    expected_endings, expected_blank_lines = _line_structure(original)
    actual_endings, actual_blank_lines = _line_structure(translated)
    if actual_endings != expected_endings:
        raise ValueError("译文换行序列与原文不一致。")
    if actual_blank_lines != expected_blank_lines:
        raise ValueError("译文空行结构与原文不一致。")

def read_translation_items(workbook_path: str | Path) -> list[TranslationItem]:
    # Normal mode releases the underlying ZIP deterministically on Windows;
    # read-only iterators can retain the workbook handle until garbage collection.
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    worksheet = workbook.active
    items: list[TranslationItem] = []
    for _row, values, ordinal in _iter_data_rows(worksheet):
        signature = _scan_control_tokens(values["original"])
        category = _category(values["code"], values["flag"], values["type"])
        copy_category = (
            _content_category(values["code"], values["flag"], values["type"])
            if category is ImportCategory.COPY
            else None
        )
        items.append(
            TranslationItem(
                key=stable_key(values["code"], values["flag"], values["original"], ordinal),
                original=values["original"],
                translation=values["translation"],
                context=" | ".join(
                    part for part in (values["type"], values["info"], f"Code={values['code']}", values["flag"])
                    if part
                ),
                stage=1 if values["translation"] else 0,
                code=values["code"],
                flag=values["flag"],
                type=values["type"],
                info=values["info"],
                category=category,
                copy_category=copy_category,
                control_signature=signature,
            )
        )
    workbook.close()
    return items

def is_font_setting(item: TranslationItem) -> bool:
    return item.code.upper() in FONT_CODES

def _location_identities(items: list[TranslationItem]) -> list[tuple[str, str, int]]:
    counts: Counter[tuple[str, str]] = Counter()
    identities: list[tuple[str, str, int]] = []
    for item in items:
        identity = (item.code, item.original)
        counts[identity] += 1
        identities.append((item.code, item.original, counts[identity]))
    return identities

def classify_optional_name_delta(
    full_items: list[TranslationItem],
    baseline_items: list[TranslationItem],
) -> int:
    full_identities = _location_identities(full_items)
    full_set = set(full_identities)
    baseline_set = set(_location_identities(baseline_items))
    missing_from_full = baseline_set - full_set
    if missing_from_full:
        raise ValueError(f"基准导出包含全量导出中不存在的行，共 {len(missing_from_full)} 条。")

    optional_count = 0
    for item, identity in zip(full_items, full_identities):
        if identity not in baseline_set:
            category = item.copy_category if item.category is ImportCategory.COPY else item.category
            if category is ImportCategory.EXTERNAL:
                continue
            if item.category is ImportCategory.COPY:
                item.copy_category = ImportCategory.OPTIONAL_NAME
            else:
                item.category = ImportCategory.OPTIONAL_NAME
            optional_count += 1
        elif item.category is ImportCategory.OPTIONAL_NAME:
            item.category = ImportCategory.DISPLAY
        elif item.category is ImportCategory.COPY and item.copy_category is ImportCategory.OPTIONAL_NAME:
            item.copy_category = ImportCategory.DISPLAY
    return optional_count

def _copy_source(item: TranslationItem, by_code: dict[str, list[TranslationItem]]) -> TranslationItem:
    current = item
    visited: set[str] = set()
    while current.category is ImportCategory.COPY:
        match = COPY_FROM_RE.search(current.flag)
        if not match:
            raise ValueError(f"COPY-FROM 行缺少来源代码: {current.code}")
        source_code = match.group(1)
        marker = f"{current.code}\0{source_code}\0{current.original}"
        if marker in visited:
            raise ValueError(f"COPY-FROM 出现循环引用: {current.code}")
        visited.add(marker)
        candidates = by_code.get(source_code, [])
        exact = [candidate for candidate in candidates if candidate.original == current.original]
        if len(exact) == 1:
            current = exact[0]
        else:
            raise ValueError(f"COPY-FROM 找不到唯一来源: {current.code} -> {source_code}")
    return current

def selected_translation_requirements(
    items: list[TranslationItem],
    scope: ImportScope,
    *,
    allow_copy_condition_groups: bool = False,
) -> dict[str, set[ImportCategory]]:
    by_code: dict[str, list[TranslationItem]] = {}
    for item in items:
        by_code.setdefault(item.code, []).append(item)

    groups: dict[str, set[ImportCategory]] = {}
    sources: dict[str, TranslationItem] = {}
    for item in items:
        category = item.copy_category if item.category is ImportCategory.COPY else item.category
        if category is None:
            continue
        source = _copy_source(item, by_code) if item.category is ImportCategory.COPY else item
        if (
            allow_copy_condition_groups
            and item.category is ImportCategory.COPY
            and item.copy_category is ImportCategory.OPTIONAL_NAME
        ):
            category = _content_category(source.code, source.flag, source.type)
        sources[source.key] = source
        categories = groups.setdefault(source.key, set())
        categories.add(category)
        intrinsic = _content_category(source.code, source.flag, source.type)
        if intrinsic in {ImportCategory.FILENAME, ImportCategory.HALFWIDTH}:
            categories.add(intrinsic)

    requirements: dict[str, set[ImportCategory]] = {}
    for key, categories in groups.items():
        if is_font_setting(sources[key]):
            continue
        # ponytail: COPY-FROM is one shared value in WOLF. Mixed-scope groups stay
        # original; enable every category in the group to translate it atomically.
        if all(scope.allows(category) for category in categories):
            requirements[key] = categories
    return requirements

def selected_translation_items(
    items: list[TranslationItem],
    scope: ImportScope,
    *,
    allow_copy_condition_groups: bool = False,
) -> list[TranslationItem]:
    required = selected_translation_requirements(
        items,
        scope,
        allow_copy_condition_groups=allow_copy_condition_groups,
    )
    return [item for item in items if item.key in required]

def to_paratranz(
    items: list[TranslationItem],
    scope: ImportScope,
    *,
    allow_copy_condition_groups: bool = False,
) -> list[dict[str, object]]:
    output: list[dict[str, object]] = []
    for item in selected_translation_items(
        items,
        scope,
        allow_copy_condition_groups=allow_copy_condition_groups,
    ):
        protected, tokens = _protect_item_tokens(item, item.original)
        if _scan_control_tokens(item.original) != item.control_signature:
            raise ValueError(f"控制符签名发生变化: {item.code}")
        translation = ""
        if item.translation:
            protected_translation, translated_tokens = _protect_item_tokens(item, item.translation)
            if translated_tokens == tokens:
                try:
                    _validate_line_structure(item.original, item.translation)
                except ValueError:
                    pass
                else:
                    translation = protected_translation
        output.append(
            {
                "key": item.key,
                "original": protected,
                "translation": translation,
                "context": item.context,
                "stage": 1 if translation else 0,
            }
        )
    return output

def _index_ainiee_rows(
    translated: list[dict[str, object]],
) -> dict[str, dict[str, object]]:
    actual: dict[str, dict[str, object]] = {}
    for row in translated:
        key = str(row.get("key", ""))
        if not key or key in actual:
            raise ValueError(f"AiNiee 输出包含空键或重复键: {key!r}")
        actual[key] = row
    return actual

def _validated_ainiee_translation(
    item: TranslationItem,
    row: dict[str, object],
) -> str:
    if row.get("wolflator_excluded") is True:
        protected, _tokens = _protect_item_tokens(item, item.original)
        if str(row.get("translation", "")) != protected:
            raise ValueError(f"AiNiee 排除项不能安全原样回填: {item.code}")
        return item.original
    raw = str(row.get("translation", ""))
    if not raw.strip():
        raise ValueError(f"AiNiee 没有生成译文: {item.code} / {item.original[:80]}")
    try:
        restored = _restore_item_tokens(item, raw)
        _validate_line_structure(item.original, restored)
        return restored
    except ValueError as exc:
        raise ValueError(f"AiNiee 译文结构校验失败: {item.code}: {exc}") from exc

def retryable_translation_errors(
    items: list[TranslationItem],
    translated: list[dict[str, object]],
    scope: ImportScope,
    *,
    allow_copy_condition_groups: bool = False,
) -> dict[str, str]:
    expected = {
        item.key: item
        for item in selected_translation_items(
            items,
            scope,
            allow_copy_condition_groups=allow_copy_condition_groups,
        )
    }
    actual = _index_ainiee_rows(translated)
    extra = set(actual) - set(expected)
    if extra:
        raise ValueError(f"AiNiee 输出包含不属于当前输入的键: extra={len(extra)}")
    errors: dict[str, str] = {}
    for key, item in expected.items():
        if key not in actual:
            errors[key] = f"AiNiee 缺少输出: {item.code}"
            continue
        try:
            _validated_ainiee_translation(item, actual[key])
        except ValueError as exc:
            errors[key] = str(exc)
    return errors

def merge_ainiee_output(
    items: list[TranslationItem],
    translated: list[dict[str, object]],
    scope: ImportScope,
    *,
    allow_copy_condition_groups: bool = False,
) -> list[TranslationItem]:
    expected = {
        item.key: item
        for item in selected_translation_items(
            items,
            scope,
            allow_copy_condition_groups=allow_copy_condition_groups,
        )
    }
    actual = _index_ainiee_rows(translated)
    missing = set(expected) - set(actual)
    extra = set(actual) - set(expected)
    if missing or extra:
        raise ValueError(f"AiNiee 输出键集合不一致: missing={len(missing)}, extra={len(extra)}")
    for key, item in expected.items():
        item.translation = _validated_ainiee_translation(item, actual[key])
        item.stage = 1
    for item in items:
        if item.category is ImportCategory.COPY:
            item.translation = ""
    return items

def normalize_import_display_middle_dots(
    items: list[TranslationItem],
    scope: ImportScope,
    *,
    allow_copy_condition_groups: bool = False,
    eligible_keys: set[str] | None = None,
) -> set[str]:
    requirements = selected_translation_requirements(
        items,
        scope,
        allow_copy_condition_groups=allow_copy_condition_groups,
    )
    eligible = (
        set(requirements)
        if eligible_keys is None
        else set(requirements) & eligible_keys
    )
    unsafe = {
        ImportCategory.EXTERNAL,
        ImportCategory.FILENAME,
        ImportCategory.HALFWIDTH,
    }
    changed: set[str] = set()
    for item in items:
        categories = requirements.get(item.key, set())
        if (
            item.key not in eligible
            or ImportCategory.DISPLAY not in categories
            or categories & unsafe
            or not item.translation
        ):
            continue
        normalized = item.translation.replace("・", "·")
        if normalized != item.translation:
            item.translation = normalized
            changed.add(item.key)
    return changed

def reconcile_incremental(
    previous: list[TranslationItem],
    current: list[TranslationItem],
) -> tuple[list[TranslationItem], list[dict[str, object]]]:
    previous_by_key = {item.key: item.translation for item in previous if item.translation}
    previous_by_original: dict[str, set[str]] = {}
    for item in previous:
        if item.translation and item.category is not ImportCategory.COPY:
            previous_by_original.setdefault(item.original, set()).add(item.translation)
    conflicts: list[dict[str, object]] = []
    for item in current:
        if item.category is ImportCategory.COPY or is_font_setting(item):
            item.translation = ""
            item.stage = 0
            continue
        exact = previous_by_key.get(item.key)
        if exact:
            item.translation = exact
            item.stage = 1
            continue
        candidates = sorted(previous_by_original.get(item.original, set()))
        if len(candidates) == 1:
            item.translation = candidates[0]
            item.stage = 1
        elif len(candidates) > 1:
            # ponytail: Ambiguous moved duplicates are retranslated instead of guessing; a future review UI may map candidates.
            item.translation = ""
            item.stage = 0
            conflicts.append(
                {
                    "key": item.key,
                    "code": item.code,
                    "original": item.original,
                    "candidates": candidates,
                }
            )
    return current, conflicts

def _save_workbook_atomic(workbook, output_path: str | Path) -> Path:
    output = Path(output_path)
    with atomic_output_path(output) as temporary:
        workbook.save(temporary)
        _normalize_xlsx_shared_strings(temporary)
    return output

def _normalize_xlsx_shared_strings(path: Path) -> None:
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    content_ns = "http://schemas.openxmlformats.org/package/2006/content-types"
    q = lambda namespace, name: f"{{{namespace}}}{name}"
    converted = path.with_name(f"{path.name}.{os.getpid()}.shared")
    try:
        with zipfile.ZipFile(path, "r") as source:
            entries = {info.filename: info for info in source.infolist()}
            payloads: dict[str, bytes] = {}
            shared_name = "xl/sharedStrings.xml"
            if shared_name in entries:
                shared = ET.fromstring(source.read(shared_name))
            else:
                shared = ET.Element(q(main_ns, "sst"))
            unique: dict[bytes, int] = {
                ET.tostring(item, encoding="utf-8"): index
                for index, item in enumerate(shared.findall(q(main_ns, "si")))
            }
            total = int(shared.get("count", "0"))
            for name, info in entries.items():
                if not name.startswith("xl/worksheets/") or not name.endswith(".xml"):
                    continue
                root = ET.fromstring(source.read(info))
                changed = False
                for cell in root.iter(q(main_ns, "c")):
                    if cell.get("t") != "inlineStr":
                        continue
                    inline = cell.find(q(main_ns, "is"))
                    if inline is None:
                        continue
                    item = ET.Element(q(main_ns, "si"))
                    for child in list(inline):
                        item.append(child)
                    key = ET.tostring(item, encoding="utf-8")
                    index = unique.get(key)
                    if index is None:
                        index = len(unique)
                        unique[key] = index
                        shared.append(item)
                    cell.remove(inline)
                    cell.set("t", "s")
                    ET.SubElement(cell, q(main_ns, "v")).text = str(index)
                    total += 1
                    changed = True
                if changed:
                    payloads[name] = ET.tostring(
                        root, encoding="utf-8", xml_declaration=True
                    )
            if not total:
                return

            shared.set("count", str(total))
            shared.set("uniqueCount", str(len(unique)))
            payloads[shared_name] = ET.tostring(
                shared, encoding="utf-8", xml_declaration=True
            )

            rels_name = "xl/_rels/workbook.xml.rels"
            rels = ET.fromstring(source.read(rels_name))
            if not any(
                relation.get("Type", "").endswith("/sharedStrings")
                for relation in rels
            ):
                used = {
                    int(match.group(1))
                    for relation in rels
                    if (match := re.fullmatch(r"rId(\d+)", relation.get("Id", "")))
                }
                next_id = max(used, default=0) + 1
                ET.SubElement(
                    rels,
                    q(rel_ns, "Relationship"),
                    {
                        "Id": f"rId{next_id}",
                        "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings",
                        "Target": "sharedStrings.xml",
                    },
                )
                payloads[rels_name] = ET.tostring(
                    rels, encoding="utf-8", xml_declaration=True
                )

            types_name = "[Content_Types].xml"
            content_types = ET.fromstring(source.read(types_name))
            if not any(
                item.get("PartName") == "/xl/sharedStrings.xml"
                for item in content_types
            ):
                ET.SubElement(
                    content_types,
                    q(content_ns, "Override"),
                    {
                        "PartName": "/xl/sharedStrings.xml",
                        "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml",
                    },
                )
                payloads[types_name] = ET.tostring(
                    content_types, encoding="utf-8", xml_declaration=True
                )

            with zipfile.ZipFile(converted, "w") as target:
                for name, info in entries.items():
                    if name == shared_name and name not in payloads:
                        continue
                    target.writestr(info, payloads.get(name, source.read(info)))
                if shared_name not in entries:
                    target.writestr(
                        shared_name,
                        payloads[shared_name],
                        compress_type=zipfile.ZIP_DEFLATED,
                    )
        replace_with_retry(converted, path)
    finally:
        converted.unlink(missing_ok=True)

def _set_literal_cell(cell, value: str) -> None:
    cell.value = value
    if value.startswith("="):
        cell.data_type = "s"

def write_full_workbook(
    template_path: str | Path,
    output_path: str | Path,
    items: list[TranslationItem],
) -> Path:
    translations = {item.key: item.translation for item in items}
    workbook = load_workbook(template_path)
    worksheet = workbook.active
    _header_row, headers = _header_map(worksheet)
    for row_index, values, ordinal in _iter_data_rows(worksheet):
        key = stable_key(values["code"], values["flag"], values["original"], ordinal)
        category = _category(values["code"], values["flag"], values["type"])
        _set_literal_cell(
            worksheet.cell(row_index, headers["__target__"]),
            ""
            if category is ImportCategory.COPY or values["code"].upper() in FONT_CODES
            else translations.get(key, ""),
        )
    return _save_workbook_atomic(workbook, output_path)

def read_font_slots(items: list[TranslationItem], *, translated: bool = False) -> list[str]:
    by_code: dict[str, list[TranslationItem]] = {code: [] for code in FONT_CODES}
    for item in items:
        code = item.code.upper()
        if code in by_code:
            by_code[code].append(item)
    invalid = [code for code, matches in by_code.items() if len(matches) > 1]
    if invalid:
        raise ValueError("字体字段重复: " + ", ".join(invalid))
    return [
        (
            by_code[code][0].translation if translated else by_code[code][0].original
        )
        if by_code[code]
        else ""
        for code in FONT_CODES
    ]

def write_font_workbook(
    template_path: str | Path,
    output_path: str | Path,
    slots: list[str],
) -> Path:
    if len(slots) != len(FONT_CODES) or not all(isinstance(value, str) for value in slots):
        raise ValueError("字体工作簿必须提供四个字体槽位")
    items = read_translation_items(template_path)
    by_code: dict[str, list[TranslationItem]] = {}
    for item in items:
        by_code.setdefault(item.code, []).append(item)
    font_copy_keys = {
        item.key
        for item in items
        if item.category is ImportCategory.COPY
        and _copy_source(item, by_code).code.upper() in FONT_CODES
    }
    workbook = load_workbook(template_path)
    worksheet = workbook.active
    _header_row, headers = _header_map(worksheet)
    found = Counter()
    slot_by_code = dict(zip(FONT_CODES, slots, strict=True))
    for row_index, values, ordinal in _iter_data_rows(worksheet):
        code = values["code"].upper()
        cell = worksheet.cell(row_index, headers["__target__"])
        key = stable_key(values["code"], values["flag"], values["original"], ordinal)
        _set_literal_cell(cell, "")
        if key in font_copy_keys:
            # COPY-FROM treats an identity translation as empty, so detach only
            # font-dependent copies before pinning their original text.
            flag = COPY_FROM_RE.sub("", values["flag"]).strip("\r\n")
            _set_literal_cell(worksheet.cell(row_index, headers[FLAG_HEADER]), flag)
            _set_literal_cell(cell, values["original"])
        if code in slot_by_code:
            found[code] += 1
            _set_literal_cell(cell, slot_by_code[code])
    invalid = [code for code in FONT_CODES if found[code] > 1]
    if invalid:
        workbook.close()
        raise ValueError("字体工作簿字段重复: " + ", ".join(invalid))
    return _save_workbook_atomic(workbook, output_path)

def dump_items(path: str | Path, items: list[TranslationItem]) -> Path:
    output = Path(path)
    atomic_write_json(
        output,
        {
            "kind": "translation-items",
            "epoch": ARTIFACT_EPOCH,
            "items": [item.to_dict() for item in items],
        },
    )
    return output

def load_items(path: str | Path) -> list[TranslationItem]:
    data = json.loads(read_text_with_retry(path, encoding="utf-8"))
    if not isinstance(data, dict) or set(data) != {"kind", "epoch", "items"}:
        raise ValueError("翻译条目文件结构不匹配。")
    require_format(
        data,
        kind="translation-items",
        version_key="epoch",
        version=ARTIFACT_EPOCH,
        label="翻译条目文件",
    )
    if not isinstance(data["items"], list):
        raise ValueError("翻译条目 items 不是数组。")
    return [TranslationItem.from_dict(item) for item in data["items"]]
