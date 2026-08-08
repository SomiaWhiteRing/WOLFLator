from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

from formats import ARTIFACT_EPOCH, require_format
from models import ImportCategory, ImportProtectionRules, ImportScope, TranslationItem
from safe_io import read_text_with_retry
from wolf_analysis import ANALYSIS_ENGINE, validate_editor_analysis
from wolf_workbook import (
    _category,
    _content_category,
    _copy_source,
    _header_map,
    _iter_data_rows,
    _save_workbook_atomic,
    _scan_control_tokens,
    _set_literal_cell,
    is_font_setting,
    selected_translation_requirements,
    stable_key,
)


EXTERNAL_DIRECTIVE_RE = re.compile(
    r"^[ \t]*@(?P<command>[slr])(?P<value>[^\s#]+)",
    re.MULTILINE | re.IGNORECASE,
)

PATH_OR_COMMAND_RE = re.compile(
    r"(?:^[ \t]*@[A-Za-z]+)|(?:\bData[\\/][^\r\n]+)|(?:[A-Za-z]:[\\/][^\r\n]+)|"
    r"(?:[^\s<>]+\.(?:png|jpe?g|webp|ogg|wav|mp3|md|txt|csv|json|ini)$)",
    re.IGNORECASE,
)

def _external_reference_evidence(game_root: str | Path) -> dict[str, str]:
    # ponytail: This recognizes the compact @s/@l/@r DSL only; add a parser when another grammar is observed.
    text_root = Path(game_root) / "Data" / "textfile"
    evidence: dict[str, str] = {}
    if not text_root.is_dir():
        return evidence
    for path in sorted(text_root.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in {".md", ".txt", ".csv"}:
            continue
        try:
            content = path.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError:
            content = path.read_text(encoding="cp932", errors="replace")
        relative = path.relative_to(Path(game_root)).as_posix()
        for match in EXTERNAL_DIRECTIVE_RE.finditer(content):
            value = match.group("value").strip().strip('"\'')
            if value and value not in evidence:
                line = content.count("\n", 0, match.start()) + 1
                evidence[value] = f"{relative}:{line}"
    return evidence

def _looks_like_path_or_command(item: TranslationItem) -> bool:
    if item.category is ImportCategory.FILENAME:
        return False
    return bool(PATH_OR_COMMAND_RE.search(item.original.strip()))

def _looks_like_identifier(value: str) -> bool:
    # ponytail: This is warning-only by default; promote project-specific patterns through exact rules, not a larger heuristic.
    text = value.strip()
    return (
        1 < len(text) <= 200
        and not any(char.isspace() for char in text)
        and "_" in text
        and any(char.isascii() and char.isalnum() for char in text)
        and any(not char.isascii() for char in text)
    )

def _logic_predicate(operator: str, value: str, literal: str) -> bool:
    if operator == "equals":
        return value == literal
    if operator == "not_equals":
        return value != literal
    if operator == "contains":
        return literal in value
    if operator == "starts_with":
        return value.startswith(literal)
    raise ValueError(f"Editor 分析报告包含未知比较操作符：{operator}")

def analyze_import_protection(
    items: list[TranslationItem],
    scope: ImportScope,
    game_root: str | Path,
    rules: ImportProtectionRules,
    logic_analysis: dict[str, object] | None = None,
    *,
    logic_safety: dict[str, object] | None = None,
    block_on_logic_issue: bool | None = None,
) -> dict[str, object]:
    requirements = selected_translation_requirements(
        items,
        scope,
        allow_copy_condition_groups=rules.allow_copy_condition_groups,
    )
    by_code: dict[str, list[TranslationItem]] = {}
    for item in items:
        by_code.setdefault(item.code, []).append(item)
    source_by_item = {
        item.key: _copy_source(item, by_code) if item.category is ImportCategory.COPY else item
        for item in items
    }
    protected: set[str] = set()
    entries: list[dict[str, object]] = []
    seen: set[tuple[str, str, str]] = set()
    logic_dependencies = (
        logic_analysis.get("dependencies", []) if isinstance(logic_analysis, dict) else []
    )
    logic_blocking = (
        logic_analysis.get("blocking_issues", []) if isinstance(logic_analysis, dict) else []
    )
    relevant_logic_blocking: list[dict[str, object]] = []
    unresolved_scope_entries: list[dict[str, object]] = []
    proven_safe: set[str] = set()
    if logic_safety is not None:
        if logic_safety.get("engine") != ANALYSIS_ENGINE:
            raise ValueError("WOLF 候选译文安全报告格式错误。")
        safe_values = logic_safety.get("safe_to_translate")
        if not isinstance(safe_values, list):
            raise ValueError("WOLF 候选译文安全报告缺少批准键集合。")
        proven_safe = {str(value) for value in safe_values}
    elif isinstance(logic_analysis, dict):
        safe_values = logic_analysis.get("safe_to_translate", [])
        if isinstance(safe_values, list):
            proven_safe = {str(value) for value in safe_values}

    def add(
        item: TranslationItem,
        action: str,
        reason: str,
        evidence: str = "",
        details: dict[str, object] | None = None,
    ) -> None:
        source = source_by_item[item.key]
        if source.key not in requirements:
            return
        if action != "atomic_translate" and not source.translation:
            return
        marker = (source.key, action, reason)
        if marker in seen:
            return
        seen.add(marker)
        if action == "keep_original":
            protected.add(source.key)
        entry: dict[str, object] = {
            "action": action,
            "reason": reason,
            "key": source.key,
            "code": source.code,
            "matched_code": item.code,
            "original": source.original,
            "translation": source.translation,
            "evidence": evidence,
        }
        if details:
            entry.update(details)
        entries.append(entry)

    for item in items:
        if (
            item.category is ImportCategory.COPY
            and rules.allow_copy_condition_groups
            and (
                item.copy_category is ImportCategory.OPTIONAL_NAME
                or "(Condition[String])" in item.context
            )
        ):
            add(item, "atomic_translate", "copy_mixed_scope_group")

    if rules.protect_external_references:
        references = _external_reference_evidence(game_root)
        for item in items:
            evidence = references.get(item.original)
            if evidence:
                add(item, "keep_original", "external_reference", evidence)

    source_items = list({source.key: source for source in source_by_item.values()}.values())
    common_items: dict[str, list[TranslationItem]] = {}
    database_items: list[tuple[TranslationItem, tuple[str, str, str, str]]] = []
    for item in source_items:
        common_match = re.match(r"COMMON-(\d+)-", item.code, re.IGNORECASE)
        if common_match:
            common_items.setdefault(common_match.group(1), []).append(item)
        database_match = re.fullmatch(
            r"(UDB|CDB|SDB)-(\d+)-(\d+)-(\d+)", item.code, re.IGNORECASE
        )
        if database_match:
            database_items.append((item, database_match.groups()))
    scope_cache: dict[str, list[TranslationItem]] = {}

    def items_for_scopes(scopes: list[object]) -> list[TranslationItem]:
        selected: dict[str, TranslationItem] = {}
        for raw_scope in scopes:
            scope_name = str(raw_scope)
            if scope_name in scope_cache:
                candidates = scope_cache[scope_name]
                for item in candidates:
                    if item.key in requirements and item.translation:
                        selected[item.key] = item
                continue
            if scope_name == "project":
                candidates = source_items
            elif scope_name == "common:*":
                candidates = [item for values in common_items.values() for item in values]
            elif scope_name.startswith("common:"):
                event_id = scope_name.split(":", 1)[1]
                candidates = common_items.get(event_id, [])
            elif scope_name.startswith("map:"):
                _, map_id, event_id, page = scope_name.split(":", 3)
                prefix = f"MAP-{map_id}-EV{int(event_id):03d}-PAGE{page}-"
                candidates = [
                    item for item in source_items if item.code.upper().startswith(prefix)
                ]
            elif scope_name.startswith("database:"):
                parts = scope_name.split(":")
                if len(parts) != 5:
                    candidates = source_items
                else:
                    _, database, type_id, data_id, field_id = parts
                    candidates = [
                        item
                        for item, coordinate in database_items
                        if coordinate[0].upper() == database.upper()
                        and (type_id == "*" or coordinate[1] == type_id)
                        and (data_id == "*" or coordinate[2] == data_id)
                        and (field_id == "*" or coordinate[3] == field_id)
                    ]
            else:
                candidates = source_items
            candidates = list(candidates)
            scope_cache[scope_name] = candidates
            for item in candidates:
                if item.key in requirements and item.translation:
                    selected[item.key] = item
        return list(selected.values())
    if rules.protect_paths_and_commands:
        for item in source_items:
            # A program proof outranks this regex fallback; explicit safety
            # rejections below can still protect the item.
            if item.key not in proven_safe and _looks_like_path_or_command(item):
                add(item, "keep_original", "path_or_command")

    if rules.protect_logic_references and logic_safety is not None:
        validate_editor_analysis(logic_analysis)
        keep_values = logic_safety.get("keep_original")
        safety_reasons = logic_safety.get("reasons")
        if not isinstance(keep_values, list) or not isinstance(safety_reasons, dict):
            raise ValueError("WOLF 候选译文安全报告缺少保护键或原因。")
        by_key = {item.key: item for item in items}
        for raw_key in keep_values:
            key = str(raw_key)
            item = by_key.get(key)
            if item is None:
                continue
            raw_reasons = safety_reasons.get(key, ["logic_safety"])
            reasons_for_key = (
                [str(value) for value in raw_reasons]
                if isinstance(raw_reasons, list)
                else [str(raw_reasons)]
            )
            add(
                item,
                "keep_original",
                "logic_safety",
                "；".join(reasons_for_key),
                {"safety_reasons": reasons_for_key},
            )
        unresolved_scope_entries = [
            {"scopes": [str(scope)], "reason": "candidate_safety", "evidence": ""}
            for scope in logic_safety.get("unresolved_scopes", ())
        ]
        relevant_logic_blocking = [
            issue
            for issue in logic_blocking
            if isinstance(issue, dict)
        ]
        should_block = (
            rules.logic_unknown_policy == "block"
            if block_on_logic_issue is None
            else block_on_logic_issue
        )
        if should_block and keep_values:
            raise RuntimeError(
                "WOLF 静态安全分析需要保留风险原文，严格模式已阻止导入。"
            )
    elif rules.protect_logic_references:
        validate_editor_analysis(logic_analysis)
        dependencies = logic_analysis.get("dependencies")
        blocking_issues = logic_analysis.get("blocking_issues")
        if not isinstance(dependencies, list) or not isinstance(blocking_issues, list):
            raise ValueError("Editor 分析报告缺少条件依赖或阻断问题。")
        by_key = {item.key: item for item in items}
        relevant_blocking: list[dict[str, object]] = []
        for dependency in dependencies:
            if not isinstance(dependency, dict):
                raise ValueError("Editor 分析报告的条件依赖格式错误。")
            operator = str(dependency.get("operator", ""))
            dependency_kind = str(dependency.get("kind", "condition"))
            literal = str(dependency.get("literal", ""))
            source_keys = dependency.get("source_keys", [])
            right_source_keys = dependency.get("right_source_keys", [])
            condition_keys = dependency.get("condition_keys", [])
            cells = dependency.get("database_cells", [])
            scopes = dependency.get("unresolved_scopes", [])
            left_values = dependency.get("left_values", [])
            right_values = dependency.get("right_values", [])
            if (
                not isinstance(source_keys, list)
                or not isinstance(right_source_keys, list)
                or not isinstance(condition_keys, list)
                or not isinstance(cells, list)
                or not isinstance(scopes, list)
                or not isinstance(left_values, list)
                or not isinstance(right_values, list)
            ):
                raise ValueError("Editor 分析报告的条件来源格式错误。")
            coordinates = ", ".join(
                f"{cell.get('database')}[{cell.get('type')},{cell.get('data')},{cell.get('field')}]"
                for cell in cells[:5]
                if isinstance(cell, dict)
            )
            evidence = (
                f"{dependency.get('auto_file', '')} event={dependency.get('event_id', '')} "
                f"page={dependency.get('page', '')} command={dependency.get('command', '')}: "
                f"{operator} {literal!r}"
                + (f" <- {coordinates}" if coordinates else "")
            )
            details = {
                "dependency_kind": dependency_kind,
                "auto_file": dependency.get("auto_file", ""),
                "event_type": dependency.get("event_type", ""),
                "event_id": dependency.get("event_id", -1),
                "event_name": dependency.get("event_name", ""),
                "page": dependency.get("page", -1),
                "command": dependency.get("command", -1),
                "operator": operator,
                "literal": literal,
                "database_cells": cells,
                "right_database_cells": dependency.get("right_database_cells", []),
                "left_values": left_values,
                "right_values": right_values,
                "dependency_status": dependency.get("status", ""),
                "unresolved_scopes": scopes,
                "resource_role": dependency.get("resource_role", ""),
            }
            if dependency_kind in {"display", "flow"}:
                continue
            if dependency_kind == "state":
                state_items = [
                    by_key[str(key)] for key in source_keys if str(key) in by_key
                ]
                if dependency.get("status") != "resolved":
                    state_items.extend(items_for_scopes(scopes))
                for item in state_items:
                    add(item, "keep_original", "logic_state_write", evidence, details)
                continue
            if dependency_kind == "resource":
                resource_items = [
                    by_key[str(key)] for key in source_keys if str(key) in by_key
                ]
                if dependency.get("status") == "blocking":
                    resource_items.extend(items_for_scopes(scopes or ["project"]))
                for item in resource_items:
                    add(item, "keep_original", "resource_reference", evidence, details)
                if dependency.get("status") == "blocking" and resource_items:
                    relevant_blocking.append(dependency)
                    unresolved_scope_entries.append(
                        {"scopes": scopes or ["project"], "reason": dependency.get("reason", ""), "evidence": evidence}
                    )
                continue
            if dependency_kind in {"call", "control_flow"}:
                call_items = [
                    by_key[str(key)] for key in [*source_keys, *right_source_keys]
                    if str(key) in by_key
                ]
                if dependency.get("status") != "resolved":
                    call_items.extend(items_for_scopes(scopes or ["common:*"]))
                for item in call_items:
                    add(item, "keep_original", "event_call_target", evidence, details)
                if dependency.get("status") == "blocking" and call_items:
                    relevant_blocking.append(dependency)
                    unresolved_scope_entries.append(
                        {
                            "scopes": scopes or ["common:*"],
                            "reason": dependency.get("reason", ""),
                            "evidence": evidence,
                        }
                    )
                continue
            for key in condition_keys:
                item = by_key.get(str(key))
                if item:
                    add(item, "keep_original", "logic_condition", evidence, details)
                    if dependency.get("status") == "untracked":
                        add(item, "warn", "logic_untracked", str(dependency.get("reason", "")), details)
            source_candidates: dict[str, TranslationItem] = {}
            relevant_sources = []
            for key in source_keys:
                item = by_key.get(str(key))
                if item is None:
                    continue
                source = source_by_item[item.key]
                source_candidates[source.key] = item
                if source.key not in requirements or not source.translation:
                    continue
                relevant_sources.append(item)
            right_candidates: dict[str, TranslationItem] = {}
            relevant_right_sources = []
            for key in right_source_keys:
                item = by_key.get(str(key))
                if item is None:
                    continue
                source = source_by_item[item.key]
                right_candidates[source.key] = item
                if source.key in requirements and source.translation:
                    relevant_right_sources.append(item)
            scoped_sources = items_for_scopes(scopes)
            if dependency.get("status") == "blocking" and (
                relevant_sources or relevant_right_sources or scoped_sources
            ):
                relevant_blocking.append(dependency)
                unresolved_scope_entries.append(
                    {"scopes": scopes or ["project"], "reason": dependency.get("reason", ""), "evidence": evidence}
                )
                if rules.logic_unknown_policy == "warn":
                    affected = [*relevant_sources, *relevant_right_sources, *scoped_sources]
                    if not affected:
                        affected = items_for_scopes(["project"])
                    for item in affected:
                        add(item, "keep_original", "logic_unresolved_scope", evidence, details)
                continue
            if dependency.get("status") == "untracked" and dependency.get("right_is_variable"):
                for item in [*relevant_sources, *relevant_right_sources]:
                    add(item, "keep_original", "logic_untracked_source", evidence, details)
            if dependency.get("status") != "resolved":
                continue
            left_originals = {
                source_by_item[item.key].original for item in source_candidates.values()
            }
            right_originals = {
                source_by_item[item.key].original for item in right_candidates.values()
            }
            derived_left = bool(left_values) and set(map(str, left_values)) != left_originals
            derived_right = bool(right_values) and set(map(str, right_values)) != right_originals
            if derived_left or derived_right:
                for item in (
                    ([*relevant_sources] if derived_left else [])
                    + ([*relevant_right_sources] if derived_right else [])
                ):
                    add(item, "keep_original", "logic_derived_value", evidence, details)
                continue
            if right_source_keys:
                # ponytail: Auto dependencies are small in practice; if a project
                # proves this O(n^2) comparison costly, preserve selector correlation.
                for left in source_candidates.values():
                    left_source = source_by_item[left.key]
                    left_final = (
                        left_source.translation
                        if left_source.key in requirements and left_source.translation
                        else left_source.original
                    )
                    for right in right_candidates.values():
                        right_source = source_by_item[right.key]
                        right_final = (
                            right_source.translation
                            if right_source.key in requirements and right_source.translation
                            else right_source.original
                        )
                        if left_final == left_source.original and right_final == right_source.original:
                            continue
                        if _logic_predicate(operator, left_source.original, right_source.original) != _logic_predicate(
                            operator, left_final, right_final
                        ):
                            add(left, "keep_original", "logic_value_change", evidence, details)
                            add(right, "keep_original", "logic_value_change", evidence, details)
                continue
            for item in relevant_sources:
                source = source_by_item[item.key]
                if _logic_predicate(operator, source.original, literal) != _logic_predicate(
                    operator, source.translation, literal
                ):
                    add(item, "keep_original", "logic_value_change", evidence, details)
        relevant_logic_blocking = relevant_blocking
        should_block = (
            rules.logic_unknown_policy == "block"
            if block_on_logic_issue is None
            else block_on_logic_issue
        )
        if relevant_blocking and should_block:
            first = relevant_blocking[0]
            raise RuntimeError(
                "WOLF 事件逻辑依赖在分析途中失去可证明来源，已阻止导入："
                f"{first.get('auto_file', '')} event={first.get('event_id', '')} "
                f"page={first.get('page', '')} command={first.get('command', '')}，"
                f"{first.get('reason', '')}。可改用“保守：保留风险原文后继续”或关闭逻辑保护。"
            )

    for item in source_items:
        if (
            item.key in requirements
            and item.translation
            and item.translation != item.original
            and item.key not in proven_safe
            and item.key not in protected
        ):
            add(
                item,
                "keep_original",
                "not_proven_safe",
                "该文本没有显示用途证明或候选译文语义等价证明。",
            )

    if rules.suspicious_identifiers != "ignore":
        action = "keep_original" if rules.suspicious_identifiers == "protect" else "warn"
        for item in source_items:
            if item.key not in protected and _looks_like_identifier(item.original):
                add(item, action, "suspicious_identifier")

    action_order = {"keep_original": 0, "warn": 1, "atomic_translate": 2}
    entries.sort(key=lambda entry: (action_order[entry["action"]], entry["code"], entry["original"]))
    return {
        "kind": "import-protection",
        "epoch": ARTIFACT_EPOCH,
        "protected_keys": sorted(protected),
        "safe_to_translate": sorted((set(requirements) & proven_safe) - protected),
        "keep_original": sorted(protected),
        "translation_overrides": (
            dict(logic_safety.get("translation_overrides", {}))
            if isinstance(logic_safety, dict)
            and isinstance(logic_safety.get("translation_overrides"), dict)
            else {}
        ),
        "approvals": (
            dict(logic_safety.get("approvals", {}))
            if isinstance(logic_safety, dict)
            else {}
        ),
        "unresolved_scopes": unresolved_scope_entries,
        "translated_replay": (
            dict(logic_safety.get("replay", {}))
            if isinstance(logic_safety, dict)
            else {
                "iterations": 0,
                "candidate_changes": 0,
                "safe_changes": 0,
                "protected_changes": 0,
                "control_flow_equivalent": False,
            }
        ),
        "structural_diff": {"status": "pending_official_roundtrip"},
        "entries": entries,
        "summary": {
            "protected": len(protected),
            "warnings": sum(entry["action"] == "warn" for entry in entries),
            "atomic_groups": sum(entry["action"] == "atomic_translate" for entry in entries),
            "logic_dependencies": len(
                logic_dependencies
            ),
            "logic_protected": len({
                entry["key"]
                for entry in entries
                if entry["reason"] in {
                    "logic_condition", "logic_value_change", "logic_safety"
                }
            }),
            "logic_blocking": len(
                logic_blocking
            ),
            "logic_blocking_relevant": len(relevant_logic_blocking),
            "logic_protection_enabled": rules.protect_logic_references,
            "logic_unknown_policy": rules.logic_unknown_policy,
            "logic_permissive_warnings": (
                len(relevant_logic_blocking)
                if rules.protect_logic_references and rules.logic_unknown_policy == "warn"
                else 0
            ),
            "logic_auto_preserved": (
                len({
                    entry["key"] for entry in entries
                    if entry["reason"] in {"logic_unresolved_scope", "logic_safety"}
                })
                if rules.protect_logic_references and rules.logic_unknown_policy == "warn"
                else 0
            ),
            "logic_proven_safe": len((set(requirements) & proven_safe) - protected),
            "logic_direct_display": len(
                logic_safety.get("approvals", {}).get("direct_display", ())
                if isinstance(logic_safety, dict)
                and isinstance(logic_safety.get("approvals"), dict)
                else ()
            ),
            "logic_display_contract": len(
                logic_safety.get("approvals", {}).get("official_display_contract", ())
                if isinstance(logic_safety, dict)
                and isinstance(logic_safety.get("approvals"), dict)
                else ()
            ),
            "logic_semantic_equivalence": len(
                logic_safety.get("approvals", {}).get("semantic_equivalence", ())
                if isinstance(logic_safety, dict)
                and isinstance(logic_safety.get("approvals"), dict)
                else ()
            ),
            "logic_external_text_flow": len(
                logic_safety.get("approvals", {}).get("external_text_flow", ())
                if isinstance(logic_safety, dict)
                and isinstance(logic_safety.get("approvals"), dict)
                else ()
            ),
            "logic_external_partial_merge": len(
                logic_safety.get("translation_overrides", {})
                if isinstance(logic_safety, dict)
                and isinstance(logic_safety.get("translation_overrides"), dict)
                else {}
            ),
            "logic_not_proven": sum(
                entry["reason"] == "not_proven_safe" for entry in entries
            ),
            "logic_risk": (
                len(logic_blocking)
                if not rules.protect_logic_references
                else 0
            ),
            "unknown_logic_semantics": len(
                logic_analysis.get("unknown_commands", [])
                if isinstance(logic_analysis, dict)
                else []
            ),
        },
    }

def validate_import_protection(value: object) -> dict[str, object]:
    report = require_format(
        value,
        kind="import-protection",
        version_key="epoch",
        version=ARTIFACT_EPOCH,
        label="导入保护报告",
    )
    expected = {
        "kind",
        "epoch",
        "protected_keys",
        "safe_to_translate",
        "keep_original",
        "translation_overrides",
        "approvals",
        "unresolved_scopes",
        "translated_replay",
        "structural_diff",
        "entries",
        "summary",
        "middle_dot_normalized",
    }
    if set(report) != expected:
        raise ValueError("导入保护报告字段不匹配。")
    if not all(
        isinstance(report[name], list)
        for name in (
            "protected_keys",
            "safe_to_translate",
            "keep_original",
            "unresolved_scopes",
            "entries",
            "middle_dot_normalized",
        )
    ) or not all(
        isinstance(report[name], dict)
        for name in (
            "translation_overrides",
            "approvals",
            "translated_replay",
            "structural_diff",
            "summary",
        )
    ):
        raise ValueError("导入保护报告字段类型不匹配。")
    return report

def load_import_protection(path: str | Path) -> dict[str, object]:
    try:
        value = json.loads(read_text_with_retry(path, encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError("导入保护报告损坏。") from exc
    return validate_import_protection(value)

def imported_display_texts(
    items: list[TranslationItem],
    scope: ImportScope,
    *,
    allow_copy_condition_groups: bool = False,
    protected_keys: set[str] | None = None,
) -> list[str]:
    """Return translated display targets that the scoped import can write."""
    requirements = selected_translation_requirements(
        items,
        scope,
        allow_copy_condition_groups=allow_copy_condition_groups,
    )
    requirements = {
        key: categories
        for key, categories in requirements.items()
        if key not in (protected_keys or set())
    }
    by_code: dict[str, list[TranslationItem]] = {}
    for item in items:
        by_code.setdefault(item.code, []).append(item)
    result: list[str] = []
    for item in items:
        if is_font_setting(item):
            continue
        intrinsic = _content_category(item.code, item.flag, item.type)
        if intrinsic in {ImportCategory.FILENAME, ImportCategory.HALFWIDTH}:
            continue
        if item.category is ImportCategory.COPY:
            source = _copy_source(item, by_code)
            text = source.translation if source.key in requirements else ""
        else:
            text = (
                item.translation
                if item.key in requirements and item.translation != item.original
                else ""
            )
        if not text:
            continue
        for token in _scan_control_tokens(text):
            text = text.replace(token, "")
        result.append(text)
    return result

def _filename_target_exists(game_root: Path, translated_name: str) -> bool:
    name = translated_name.strip().replace("\\", "/").lstrip("/")
    if not name or ".." in Path(name).parts:
        return False
    data_root = game_root / "Data"
    direct = data_root.joinpath(*name.split("/"))
    return direct.is_file()

def write_scoped_workbook(
    full_path: str | Path,
    output_path: str | Path,
    scope: ImportScope,
    game_root: str | Path,
    items: list[TranslationItem],
    *,
    allow_copy_condition_groups: bool = False,
    protected_keys: set[str] | None = None,
    translation_overrides: dict[str, str] | None = None,
) -> Path:
    workbook = load_workbook(full_path)
    worksheet = workbook.active
    _header_row, headers = _header_map(worksheet)
    requirements = selected_translation_requirements(
        items,
        scope,
        allow_copy_condition_groups=allow_copy_condition_groups,
    )
    requirements = {
        key: categories
        for key, categories in requirements.items()
        if key not in (protected_keys or set())
    }
    items_by_key = {item.key: item for item in items}
    by_code: dict[str, list[TranslationItem]] = {}
    for item in items:
        by_code.setdefault(item.code, []).append(item)
    missing_filenames: list[str] = []
    for row_index, values, ordinal in _iter_data_rows(worksheet):
        key = stable_key(values["code"], values["flag"], values["original"], ordinal)
        category = _category(values["code"], values["flag"], values["type"])
        cell = worksheet.cell(row_index, headers["__target__"])
        item = items_by_key.get(key)
        required_categories = requirements.get(key, set())
        keep = bool(required_categories)
        if category is ImportCategory.COPY:
            if item is None:
                raise ValueError(f"范围工作簿找不到 COPY-FROM 条目: {values['code']}")
            source = _copy_source(item, by_code)
            copy_enabled = source.key in requirements
            _set_literal_cell(
                cell,
                source.translation if copy_enabled and source.translation else item.original,
            )
            continue
        if not keep:
            _set_literal_cell(cell, "")
            continue
        if item is None:
            raise ValueError(f"范围工作簿找不到翻译条目: {values['code']}")
        target = (translation_overrides or {}).get(
            item.key,
            item.translation if item.translation != item.original else "",
        )
        # ponytail: rewrite every selected cell from the item model so a stale
        # workbook formula cannot leak into the official tool input.
        _set_literal_cell(cell, target)
        if (
            ImportCategory.FILENAME in required_categories
            and target
            and not _filename_target_exists(Path(game_root), target)
        ):
            missing_filenames.append(target)
    if missing_filenames:
        sample = ", ".join(missing_filenames[:5])
        raise ValueError(f"文件名译文没有对应真实文件，共 {len(missing_filenames)} 项，例如: {sample}")
    return _save_workbook_atomic(workbook, output_path)
